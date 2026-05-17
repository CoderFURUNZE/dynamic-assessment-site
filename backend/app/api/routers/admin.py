import json
import logging
import re
import shutil
import csv
from collections import Counter
from io import BytesIO, StringIO
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Response
from sqlmodel import Session, select

from app.api.deps import require_role, teacher_has_course_access
from app.core.config import settings
from app.core.security import hash_password
from app.db.models import (
    CourseLifecycleStatus,
    CourseEnrollStatus,
    ChapterEdge,
    Course,
    CourseApplication,
    CourseCompletionRecord,
    CourseTeacherActivation,
    CoursePortraitIndicatorSelection,
    CoursePrerequisite,
    CourseStage,
    EvalConfig,
    KnowledgeEdge,
    KnowledgePoint,
    KpTask,
    KpTaskType,
    LearnerPersonaOverride,
    LearningBehaviorEvent,
    LearningResource,
    Mastery,
    PersonaType,
    RecommendationLog,
    RelationType,
    ResourceType,
    QuestionnairePortraitIndicatorInput,
    KpQuestionAssignment,
    PracticeAttempt,
    Enrollment,
    EnrollmentStatus,
    Question,
    Quiz,
    QuizAttempt,
    QuizItem,
    ReviewSchedule,
    Note,
    ExpressionEvent,
    StageImportBatch,
    StageImportRecord,
    StageEvaluationSnapshot,
    TeacherFinalScoreConfirmation,
    TeacherPortraitIndicatorInput,
    StageTeacherFeedback,
    TeacherCourseStatus,
    User,
    UserRole,
    VideoProgress,
    AuditLog,
)
from app.db.session import get_session
from sqlalchemy import Integer, false, func, or_, delete
from app.schemas.admin import (
    KnowledgeEdgeIn,
    KnowledgeEdgeOut,
    KnowledgePointIn,
    KnowledgePointUpdateIn,
    KpResourceIn,
    KpResourceUpdateIn,
    KpTaskIn,
    KpTaskUpdateIn,
    CourseIn,
    CourseOut,
    CourseUpdateIn,
    QuestionIn,
    QuestionOut,
    PersonaOverrideIn,
    PersonaOverrideOut,
    PersonaRuleOut,
    AdminAnalyticsOut,
    AdminPracticeReportOut,
    AuditLogOut,
    TeacherFinalScoreConfirmIn,
    UserImportPreviewOut,
    UserImportResultOut,
    UserOut,
    UserUpdateIn,
)
from app.schemas.paging import PageOut
from app.services.learner_profile import (
    _aggregate_stage_portrait_summary,
    _json_load,
    build_cohort_ability_practice_summary,
    canonical_ability_subtags_str,
    clear_persona_override,
    get_or_create_persona_rule,
    get_latest_profile_snapshot,
    get_latest_stage_snapshot,
    get_stage_snapshot_trend,
    get_stage_teacher_feedback,
    persona_label,
    recalculate_profile_snapshot,
    recalculate_profiles_for_subject,
    recalculate_stage_snapshots_for_stage,
    resolve_persona_thresholds,
    resolve_persona_weights,
    sync_profile_snapshot_from_stage,
    normalize_question_cognitive_level,
    upsert_persona_override,
    upsert_stage_teacher_feedback,
)
from app.services.resource_files import (
    _preview_type_for_detected,
    _resource_type_from_detected,
    build_resource_payload,
    inspect_uploaded_file_stream,
    maybe_prepare_preview,
    store_uploaded_file_stream,
    store_video_file_stream,
)
from app.services.kp_tagging import auto_tag_knowledge_points

router = APIRouter(prefix="/admin", tags=["admin"])
logger = logging.getLogger("app.audit")

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependency guard
    load_workbook = None


def _log_action(session: Session | None, user: User | None, action: str, detail: str = "") -> None:
    if user is None:
        try:
            logger.info("actor=system action=%s detail=%s", action, detail)
            if session is None:
                return
            session.add(AuditLog(actor="system", role="system", action=action, detail=detail))
            session.commit()
        except Exception:
            if session is not None:
                session.rollback()
            logger.info("action=%s detail=%s", action, detail)
        return
    try:
        logger.info("actor=%s role=%s action=%s detail=%s", user.username, user.role.value, action, detail)
        if session is not None:
            session.add(
                AuditLog(actor=user.username, role=user.role.value, action=action, detail=detail)
            )
            session.commit()
    except Exception:
        if session is not None:
            session.rollback()
        logger.info("action=%s detail=%s", action, detail)


USER_IMPORT_FIELD_ALIASES: dict[str, list[str]] = {
    "username": ["用户名", "账号"],
    "password": ["密码", "初始密码"],
    "full_name": ["姓名", "教师姓名", "学生姓名"],
    "student_no": ["学号", "工号", "编号"],
    "class_name": ["班级", "所属班级", "行政班"],
    "phone": ["手机号", "电话", "手机"],
    "active": ["状态", "是否启用", "启用"],
}


def _normalize_user_import_row(row: dict[str, str]) -> dict[str, str]:
    normalized = {str(key).strip(): "" if value is None else str(value).strip() for key, value in row.items()}
    for canonical, aliases in USER_IMPORT_FIELD_ALIASES.items():
        if canonical in normalized and normalized[canonical]:
            continue
        for alias in aliases:
            if alias in normalized and normalized[alias]:
                normalized[canonical] = normalized[alias]
                break
    return normalized


def _user_rows_from_upload(file: UploadFile, payload: bytes) -> list[dict[str, str]]:
    name = (file.filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [_normalize_user_import_row(row) for row in reader]
    if name.endswith(".xlsx"):
        if load_workbook is None:
            raise HTTPException(status_code=400, detail="xlsx import requires openpyxl")
        wb = load_workbook(BytesIO(payload), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(item).strip() if item is not None else "" for item in values[0]]
        rows: list[dict[str, str]] = []
        for value_row in values[1:]:
            item: dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                cell = value_row[idx] if idx < len(value_row) else ""
                item[header] = "" if cell is None else str(cell).strip()
            rows.append(_normalize_user_import_row(item))
        return rows
    raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")


def _to_active_flag(raw_value: str | None) -> bool:
    raw = str(raw_value or "").strip().lower()
    if not raw:
        return True
    return raw in {"1", "true", "yes", "y", "启用", "active", "enabled"}


def _user_import_template_csv(role: str) -> str:
    if role == UserRole.teacher.value:
        rows = [
            "username,password,full_name,student_no,class_name,phone,active",
            "teacher_demo,Temp1234,张老师,T0001,教研组A,13800138000,true",
        ]
    else:
        rows = [
            "username,password,full_name,student_no,class_name,phone,active",
            "student_demo,Temp1234,张同学,20260001,软件工程1班,13800138001,true",
        ]
    return "\n".join(rows) + "\n"


def _preview_user_import_rows(
    *,
    session: Session,
    role_value: str,
    rows: list[dict[str, str]],
) -> UserImportPreviewOut:
    required_fields = ["username", "password", "full_name"]
    detected_fields = sorted({key for row in rows for key in row.keys() if key})
    warnings: list[str] = []
    errors: list[str] = []
    valid_rows = 0
    seen_usernames: set[str] = set()
    seen_phones: set[str] = set()
    matched_courses: list[dict] = []

    missing_required = [field for field in required_fields if field not in detected_fields]
    if missing_required:
        warnings.append(f"缺少模板字段：{', '.join(missing_required)}")

    for index, row in enumerate(rows, start=2):
        username = (row.get("username") or "").strip()
        password = (row.get("password") or "").strip()
        full_name = (row.get("full_name") or "").strip()
        phone = (row.get("phone") or "").strip()
        row_errors: list[str] = []

        if not username:
            row_errors.append("username is required")
        if not full_name:
            row_errors.append("full_name is required")
        existing = session.exec(select(User).where(User.username == username)).first() if username else None
        if existing is None and not password:
            row_errors.append("password is required for new user")
        if username:
            if username in seen_usernames:
                row_errors.append(f"duplicate username in file: {username}")
            seen_usernames.add(username)
        if phone:
            phone_owner = session.exec(select(User).where(User.phone == phone)).first()
            if phone_owner is not None and (existing is None or int(phone_owner.id) != int(existing.id)):
                row_errors.append(f"phone already exists: {phone}")
            if phone in seen_phones:
                row_errors.append(f"duplicate phone in file: {phone}")
            seen_phones.add(phone)

        if row_errors:
            errors.append(f"row {index}: {'; '.join(row_errors)}")
        else:
            valid_rows += 1

    if role_value == UserRole.teacher.value:
        warnings.append("教师导入会统一写入 teacher 角色，已有同名账号会被更新。")
    else:
        warnings.append("学生导入会统一写入 student 角色，已有同名账号会被更新。")
        class_names = sorted({(row.get("class_name") or "").strip() for row in rows if (row.get("class_name") or "").strip()})
        if class_names:
            courses = session.exec(
                select(Course).where(
                    Course.active == True,  # noqa: E712
                    Course.lifecycle_status == CourseLifecycleStatus.active,
                    Course.target_class.in_(class_names),
                )
            ).all()
            matched_courses = [
                {
                    "course_id": int(course.id),
                    "course_title": course.title,
                    "course_code": course.code,
                    "target_class": course.target_class,
                }
                for course in courses
                if course.id is not None
            ]
            if matched_courses:
                warnings.append(f"检测到 {len(matched_courses)} 门课程将按班级自动分配学生。")

    return UserImportPreviewOut(
        role=role_value,
        total_rows=len(rows),
        valid_rows=valid_rows,
        invalid_rows=max(0, len(rows) - valid_rows),
        required_fields=required_fields,
        detected_fields=detected_fields,
        matched_courses=matched_courses,
        warnings=warnings[:20],
        errors=errors[:20],
    )


def _bilibili_embed_url(*, bvid: str, page: int) -> str:
    p = max(1, int(page))
    return f"https://player.bilibili.com/player.html?bvid={bvid}&page={p}"


def _replace_kp_video(*, session: Session, kp: KnowledgePoint, title: str, url: str) -> LearningResource:
    existing = session.exec(
        select(LearningResource).where(LearningResource.kp_id == kp.id, LearningResource.type == ResourceType.video)
    ).all()
    for r in existing:
        session.delete(r)
    session.commit()

    r = LearningResource(subject=kp.subject, grade=kp.grade, kp_id=kp.id, title=title, url=url, type=ResourceType.video)
    session.add(r)
    session.commit()
    session.refresh(r)
    return r


def _ensure_teacher_id(session: Session, teacher_id: int | None) -> int | None:
    if teacher_id in {None, 0}:
        return None
    teacher = session.get(User, teacher_id)
    if teacher is None or teacher.role != UserRole.teacher:
        raise HTTPException(status_code=400, detail="teacher_id must reference a teacher user")
    return int(teacher.id)


def _normalize_teacher_ids(session: Session, teacher_ids: list[int] | None, fallback_teacher_id: int | None = None) -> list[int]:
    raw_ids = list(teacher_ids or [])
    if not raw_ids and fallback_teacher_id is not None:
        raw_ids = [fallback_teacher_id]
    normalized: list[int] = []
    for raw_id in raw_ids:
        teacher_id = _ensure_teacher_id(session, int(raw_id) if raw_id is not None else None)
        if teacher_id is not None and teacher_id not in normalized:
            normalized.append(int(teacher_id))
    return normalized


def _course_teacher_ids(session: Session, course: Course) -> list[int]:
    if course.id is None:
        return [int(course.teacher_id)] if course.teacher_id is not None else []
    rows = session.exec(
        select(CourseTeacherActivation).where(CourseTeacherActivation.course_id == int(course.id))
    ).all()
    teacher_ids = [int(row.teacher_id) for row in rows if row.teacher_id is not None]
    if course.teacher_id is not None and int(course.teacher_id) not in teacher_ids:
        teacher_ids.insert(0, int(course.teacher_id))
    return teacher_ids


def _sync_course_teacher_assignments(session: Session, course: Course, teacher_ids: list[int]) -> None:
    if course.id is None:
        return
    course_id = int(course.id)
    existing_rows = session.exec(
        select(CourseTeacherActivation).where(CourseTeacherActivation.course_id == course_id)
    ).all()
    existing_map = {int(row.teacher_id): row for row in existing_rows if row.teacher_id is not None}
    target_ids = set(teacher_ids)
    for teacher_id in teacher_ids:
        existing = existing_map.get(int(teacher_id))
        if existing is None:
            session.add(
                CourseTeacherActivation(
                    course_id=course_id,
                    teacher_id=int(teacher_id),
                    teaching_status=TeacherCourseStatus.not_started,
                    finished_at=None,
                    updated_at=datetime.utcnow(),
                )
            )
    for teacher_id, existing in existing_map.items():
        if teacher_id not in target_ids and existing.teaching_status == TeacherCourseStatus.not_started:
            session.delete(existing)


def _sync_course_class_enrollments(session: Session, course: Course) -> None:
    if course.id is None:
        return
    course_id = int(course.id)
    lifecycle = course.lifecycle_status.value if hasattr(course.lifecycle_status, "value") else str(course.lifecycle_status or "draft")
    target_class = str(course.target_class or "").strip()
    if not course.active or lifecycle != CourseLifecycleStatus.active.value or not target_class:
        return
    students = session.exec(
        select(User).where(
            User.role == UserRole.student,
            User.active == True,  # noqa: E712
            User.class_name == target_class,
        )
    ).all()
    if not students:
        return
    existing = session.exec(
        select(Enrollment).where(
            Enrollment.course_id == course_id,
            Enrollment.status == EnrollmentStatus.active,
        )
    ).all()
    existing_ids = {int(item.student_id) for item in existing}
    created = False
    for student in students:
        if student.id is None or int(student.id) in existing_ids:
            continue
        session.add(
            Enrollment(
                student_id=int(student.id),
                course_id=course_id,
                application_id=None,
                status=EnrollmentStatus.active,
            )
        )
        created = True
    if created:
        session.commit()


def _sync_student_class_enrollments_for_user(session: Session, user: User) -> int:
    if user.id is None or user.role != UserRole.student or not bool(user.active):
        return 0
    target_class = str(user.class_name or "").strip()
    if not target_class:
        return 0
    courses = session.exec(
        select(Course).where(
            Course.active == True,  # noqa: E712
            Course.lifecycle_status == CourseLifecycleStatus.active,
            Course.target_class == target_class,
        )
    ).all()
    if not courses:
        return 0
    course_ids = [int(course.id) for course in courses if course.id is not None]
    existing_rows = session.exec(
        select(Enrollment).where(
            Enrollment.student_id == int(user.id),
            Enrollment.course_id.in_(course_ids),
        )
    ).all()
    existing_map = {int(item.course_id): item for item in existing_rows if item.course_id is not None}
    created_or_reactivated = 0
    for course in courses:
        if course.id is None:
            continue
        course_id = int(course.id)
        existing = existing_map.get(course_id)
        if existing is None:
            session.add(
                Enrollment(
                    student_id=int(user.id),
                    course_id=course_id,
                    application_id=None,
                    status=EnrollmentStatus.active,
                )
            )
            created_or_reactivated += 1
            continue
        if existing.status != EnrollmentStatus.active:
            existing.status = EnrollmentStatus.active
            session.add(existing)
            created_or_reactivated += 1
    if created_or_reactivated:
        session.commit()
    return created_or_reactivated


def _snapshot_for_user(session: Session, *, user_id: int, subject: str, grade: str):
    snapshot = get_latest_profile_snapshot(session, user_id=user_id, subject=subject, grade=grade)
    if snapshot is None:
        snapshot = recalculate_profile_snapshot(
            session,
            user_id=user_id,
            subject=subject,
            grade=grade,
            refresh_mastery=False,
            persist=True,
        )
    return snapshot


def _resolve_course_for_subject(
    session: Session,
    *,
    subject: str,
    grade: str | None = None,
    admin: User | None = None,
) -> Course | None:
    stmt = select(Course).where(Course.title == subject)
    if grade:
        stmt = (
            stmt.join(CourseStage, CourseStage.course_id == Course.id)
            .where(
                CourseStage.subject == subject,
                CourseStage.grade == grade,
            )
            .distinct()
        )
    stmt = stmt.order_by(Course.active.desc(), Course.id.desc())
    rows = session.exec(stmt).all()
    course = next(
        (
            item
            for item in rows
            if admin is None or admin.role != UserRole.teacher or teacher_has_course_access(session, int(admin.id), item)
        ),
        None,
    )
    if course is not None or not grade:
        return course
    fallback = select(Course).where(Course.title == subject)
    fallback_rows = session.exec(fallback.order_by(Course.active.desc(), Course.id.desc())).all()
    return next(
        (
            item
            for item in fallback_rows
            if admin is None or admin.role != UserRole.teacher or teacher_has_course_access(session, int(admin.id), item)
        ),
        None,
    )


def _check_teacher_subject_access(*, session: Session, admin: User, course: Course | None) -> None:
    if admin.role == UserRole.teacher and course is not None and not teacher_has_course_access(session, int(admin.id), course):
        raise HTTPException(status_code=403, detail="No permission for this subject")


def _teacher_accessible_subjects(session: Session, admin: User) -> set[str] | None:
    if admin.role != UserRole.teacher:
        return None
    rows = session.exec(select(Course).order_by(Course.id.desc())).all()
    return {
        str(course.title).strip()
        for course in rows
        if course.title and teacher_has_course_access(session, int(admin.id), course)
    }


def _require_teacher_subject_access(
    *,
    session: Session,
    admin: User,
    subject: str | None,
    grade: str | None = None,
) -> None:
    if admin.role != UserRole.teacher:
        return
    normalized_subject = str(subject or "").strip()
    if not normalized_subject:
        raise HTTPException(status_code=403, detail="No permission for this subject")
    course = _resolve_course_for_subject(session, subject=normalized_subject, grade=grade, admin=admin)
    if course is None:
        raise HTTPException(status_code=403, detail="No permission for this subject")
    _check_teacher_subject_access(session=session, admin=admin, course=course)


def _require_teacher_kp_access(*, session: Session, admin: User, kp: KnowledgePoint | None) -> None:
    if kp is None:
        raise HTTPException(status_code=404, detail="Knowledge point not found")
    _require_teacher_subject_access(session=session, admin=admin, subject=kp.subject, grade=kp.grade)


def _require_teacher_kp_id_access(*, session: Session, admin: User, kp_id: int) -> KnowledgePoint:
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    return kp


def _require_teacher_resource_access(*, session: Session, admin: User, row: LearningResource | None) -> LearningResource:
    if row is None:
        raise HTTPException(status_code=404, detail="Resource not found")
    _require_teacher_subject_access(session=session, admin=admin, subject=row.subject, grade=row.grade)
    return row


def _require_teacher_task_access(*, session: Session, admin: User, row: KpTask | None) -> KpTask:
    if row is None:
        raise HTTPException(status_code=404, detail="Task not found")
    _require_teacher_subject_access(session=session, admin=admin, subject=row.subject, grade=row.grade)
    return row


def _require_teacher_question_access(*, session: Session, admin: User, question: Question | None) -> Question:
    if question is None:
        raise HTTPException(status_code=404, detail="Question not found")
    _require_teacher_subject_access(session=session, admin=admin, subject=question.subject, grade=question.grade)
    return question


def _active_course_student_ids(session: Session, *, course_id: int) -> list[int]:
    rows = session.exec(
        select(Enrollment.student_id).where(
            Enrollment.course_id == course_id,
            Enrollment.status == EnrollmentStatus.active,
        )
    ).all()
    return [int(row) for row in rows if row is not None]


def _course_delete_blockers(session: Session, *, course_id: int, subject: str) -> list[str]:
    checks = [
        ("知识点", session.exec(select(func.count()).select_from(KnowledgePoint).where(KnowledgePoint.subject == subject)).one()),
        ("选课申请", session.exec(select(func.count()).select_from(CourseApplication).where(CourseApplication.course_id == course_id)).one()),
        ("选课记录", session.exec(select(func.count()).select_from(Enrollment).where(Enrollment.course_id == course_id)).one()),
        ("课程阶段", session.exec(select(func.count()).select_from(CourseStage).where(CourseStage.course_id == course_id)).one()),
        ("阶段导入批次", session.exec(select(func.count()).select_from(StageImportBatch).where(StageImportBatch.course_id == course_id)).one()),
        ("阶段导入记录", session.exec(select(func.count()).select_from(StageImportRecord).where(StageImportRecord.course_id == course_id)).one()),
        ("阶段评价快照", session.exec(select(func.count()).select_from(StageEvaluationSnapshot).where(StageEvaluationSnapshot.course_id == course_id)).one()),
        ("教师阶段反馈", session.exec(select(func.count()).select_from(StageTeacherFeedback).where(StageTeacherFeedback.course_id == course_id)).one()),
        ("期末确认记录", session.exec(select(func.count()).select_from(TeacherFinalScoreConfirmation).where(TeacherFinalScoreConfirmation.course_id == course_id)).one()),
        ("教师画像输入", session.exec(select(func.count()).select_from(TeacherPortraitIndicatorInput).where(TeacherPortraitIndicatorInput.course_id == course_id)).one()),
        ("问卷画像输入", session.exec(select(func.count()).select_from(QuestionnairePortraitIndicatorInput).where(QuestionnairePortraitIndicatorInput.course_id == course_id)).one()),
        ("课程画像配置", session.exec(select(func.count()).select_from(CoursePortraitIndicatorSelection).where(CoursePortraitIndicatorSelection.course_id == course_id)).one()),
        ("行为记录", session.exec(select(func.count()).select_from(LearningBehaviorEvent).where(LearningBehaviorEvent.course_id == course_id)).one()),
        (
            "先修课程关系",
            session.exec(
                select(func.count()).select_from(CoursePrerequisite).where(
                    or_(
                        CoursePrerequisite.course_id == course_id,
                        CoursePrerequisite.prerequisite_course_id == course_id,
                    )
                )
            ).one(),
        ),
    ]
    return [f"{label}{int(count)}条" for label, count in checks if int(count or 0) > 0]


def _purge_course_related_rows(session: Session, *, course_id: int) -> None:
    session.exec(delete(TeacherPortraitIndicatorInput).where(TeacherPortraitIndicatorInput.course_id == course_id))
    session.exec(delete(StageTeacherFeedback).where(StageTeacherFeedback.course_id == course_id))
    session.exec(delete(StageEvaluationSnapshot).where(StageEvaluationSnapshot.course_id == course_id))
    session.exec(delete(StageImportRecord).where(StageImportRecord.course_id == course_id))
    session.exec(delete(StageImportBatch).where(StageImportBatch.course_id == course_id))
    session.exec(delete(TeacherFinalScoreConfirmation).where(TeacherFinalScoreConfirmation.course_id == course_id))
    session.exec(
        delete(QuestionnairePortraitIndicatorInput).where(
            QuestionnairePortraitIndicatorInput.course_id == course_id
        )
    )
    session.exec(
        delete(CoursePortraitIndicatorSelection).where(CoursePortraitIndicatorSelection.course_id == course_id)
    )
    session.exec(delete(LearningBehaviorEvent).where(LearningBehaviorEvent.course_id == course_id))
    session.exec(delete(CourseCompletionRecord).where(CourseCompletionRecord.course_id == course_id))
    session.exec(delete(Enrollment).where(Enrollment.course_id == course_id))
    session.exec(delete(CourseApplication).where(CourseApplication.course_id == course_id))
    session.exec(
        delete(CoursePrerequisite).where(
            or_(
                CoursePrerequisite.course_id == course_id,
                CoursePrerequisite.prerequisite_course_id == course_id,
            )
        )
    )
    session.exec(delete(CourseTeacherActivation).where(CourseTeacherActivation.course_id == course_id))
    session.exec(delete(CourseStage).where(CourseStage.course_id == course_id))
    session.flush()


def _all_active_students_final_confirmed(session: Session, *, course_id: int) -> bool:
    active_student_ids = set(_active_course_student_ids(session, course_id=course_id))
    if not active_student_ids:
        return False
    confirmed_user_ids = {
        int(item)
        for item in session.exec(
            select(TeacherFinalScoreConfirmation.user_id).where(
                TeacherFinalScoreConfirmation.course_id == course_id,
                TeacherFinalScoreConfirmation.user_id.in_(list(active_student_ids)),
            )
        ).all()
        if item is not None
    }
    return confirmed_user_ids.issuperset(active_student_ids)


def _course_has_active_students(session: Session, *, course_id: int) -> bool:
    return bool(_active_course_student_ids(session, course_id=course_id))


def _mark_course_teaching_finished(session: Session, *, course: Course) -> None:
    if course.id is None:
        return
    now = datetime.utcnow()
    activation_rows = session.exec(
        select(CourseTeacherActivation).where(CourseTeacherActivation.course_id == int(course.id))
    ).all()
    for activation in activation_rows:
        if activation.teaching_status != TeacherCourseStatus.finished:
            activation.teaching_status = TeacherCourseStatus.finished
            activation.finished_at = now
            activation.updated_at = now
            session.add(activation)
    course.enroll_status = CourseEnrollStatus.closed
    session.add(course)


def _build_student_detail_payload(
    session: Session,
    *,
    user_id: int,
    subject: str,
    grade: str,
    course: Course | None,
):
    student = session.get(User, user_id)
    if student is None or student.role != UserRole.student:
        raise HTTPException(status_code=404, detail="Student not found")

    snapshot = _snapshot_for_user(session, user_id=user_id, subject=subject, grade=grade)
    if snapshot is None:
        raise HTTPException(status_code=404, detail="No profile snapshot for this student and subject")
    stage_snapshots = list(reversed(get_stage_snapshot_trend(session, user_id=user_id, subject=subject, grade=grade, limit=12)))
    current_stage = stage_snapshots[-1] if stage_snapshots else None
    feedback = get_stage_teacher_feedback(
        session,
        user_id=user_id,
        subject=subject,
        grade=grade,
        stage_id=int(current_stage.stage_id) if current_stage is not None else None,
    )
    kps = session.exec(
        select(KnowledgePoint)
        .where(KnowledgePoint.subject == subject, KnowledgePoint.grade == grade)
        .order_by(KnowledgePoint.chapter, KnowledgePoint.id)
    ).all()
    kp_ids = [int(kp.id) for kp in kps if kp.id is not None]

    mastery_rows = []
    if kp_ids:
        mastery_rows = session.exec(
            select(Mastery).where(Mastery.user_id == user_id, Mastery.kp_id.in_(kp_ids))
        ).all()
    mastery_map = {int(row.kp_id): row for row in mastery_rows}
    mastery_items = [
        {
            "kp_id": int(kp.id),
            "code": kp.code,
            "title": kp.title,
            "chapter": kp.chapter,
            "mastery": float(mastery_map.get(int(kp.id)).value) if int(kp.id) in mastery_map else 0.0,
            "direct_value": float(mastery_map.get(int(kp.id)).direct_value) if int(kp.id) in mastery_map else 0.0,
            "status": mastery_map.get(int(kp.id)).status if int(kp.id) in mastery_map else "not_started",
            "reason_summary": mastery_map.get(int(kp.id)).reason_summary if int(kp.id) in mastery_map else "",
        }
        for kp in kps
        if kp.id is not None
    ]

    behavior_rows = session.exec(
        select(LearningBehaviorEvent)
        .where(LearningBehaviorEvent.user_id == user_id)
        .order_by(LearningBehaviorEvent.created_at.desc())
        .limit(30)
    ).all()
    behavior_items = [
        {
            "id": int(row.id),
            "event_type": row.event_type,
            "kp_id": row.kp_id,
            "value_json": row.value_json,
            "created_at": row.created_at.isoformat(),
        }
        for row in behavior_rows
        if row.id is not None
    ]

    recommendation_rows = session.exec(
        select(RecommendationLog)
        .where(
            RecommendationLog.user_id == user_id,
            RecommendationLog.subject == subject,
            RecommendationLog.grade == grade,
        )
        .order_by(RecommendationLog.created_at.desc())
        .limit(10)
    ).all()
    recommendation_items = [
        {
            "id": int(row.id),
            "source_kp_id": row.source_kp_id,
            "target_kp_id": row.target_kp_id,
            "persona_type": row.persona_type.value if isinstance(row.persona_type, PersonaType) else str(row.persona_type),
            "reason_summary": row.reason_summary,
            "created_at": row.created_at.isoformat(),
            "payload_json": row.payload_json,
        }
        for row in recommendation_rows
        if row.id is not None
    ]

    practice_rows = session.exec(
        select(PracticeAttempt)
        .where(PracticeAttempt.user_id == user_id, PracticeAttempt.kp_id.in_(kp_ids) if kp_ids else True)
        .order_by(PracticeAttempt.created_at.desc())
        .limit(20)
    ).all()
    quiz_rows = session.exec(
        select(QuizAttempt)
        .where(QuizAttempt.user_id == user_id, QuizAttempt.kp_id.in_(kp_ids) if kp_ids else True)
        .order_by(QuizAttempt.created_at.desc())
        .limit(10)
    ).all()
    video_rows = session.exec(
        select(VideoProgress)
        .where(VideoProgress.user_id == user_id, VideoProgress.kp_id.in_(kp_ids) if kp_ids else True)
        .order_by(VideoProgress.updated_at.desc())
        .limit(10)
    ).all()

    now = datetime.utcnow()
    since_30d = now - timedelta(days=30)
    since_14d = now - timedelta(days=14)

    course_id = int(course.id) if course is not None and course.id is not None else None
    behavior_scope_stmt = select(LearningBehaviorEvent).where(
        LearningBehaviorEvent.user_id == user_id,
        LearningBehaviorEvent.created_at >= since_30d,
    )
    if course_id is not None and kp_ids:
        behavior_scope_stmt = behavior_scope_stmt.where(
            or_(
                LearningBehaviorEvent.course_id == course_id,
                LearningBehaviorEvent.kp_id.in_(kp_ids),
            )
        )
    elif course_id is not None:
        behavior_scope_stmt = behavior_scope_stmt.where(LearningBehaviorEvent.course_id == course_id)
    elif kp_ids:
        behavior_scope_stmt = behavior_scope_stmt.where(LearningBehaviorEvent.kp_id.in_(kp_ids))
    else:
        behavior_scope_stmt = behavior_scope_stmt.where(false())
    scoped_behavior_rows = session.exec(
        behavior_scope_stmt.order_by(LearningBehaviorEvent.created_at.desc()).limit(600)
    ).all()
    scoped_behavior_14d = [row for row in scoped_behavior_rows if row.created_at and row.created_at >= since_14d]

    practice_14d = [row for row in practice_rows if row.created_at and row.created_at >= since_14d]
    quiz_14d = [row for row in quiz_rows if row.created_at and row.created_at >= since_14d]
    video_14d = [row for row in video_rows if row.updated_at and row.updated_at >= since_14d]

    login_rows_30d = [row for row in scoped_behavior_rows if (row.event_type or "").strip().lower() == "login"]
    login_days_30d = {row.created_at.date() for row in login_rows_30d if row.created_at}
    active_days_14d = (
        {row.created_at.date() for row in scoped_behavior_14d if row.created_at}
        | {row.created_at.date() for row in practice_14d if row.created_at}
        | {row.created_at.date() for row in quiz_14d if row.created_at}
        | {row.updated_at.date() for row in video_14d if row.updated_at}
    )
    consecutive_days = 0
    if active_days_14d:
        ordered = sorted(active_days_14d, reverse=True)
        consecutive_days = 1
        for idx in range(1, len(ordered)):
            if (ordered[idx - 1] - ordered[idx]).days == 1:
                consecutive_days += 1
                continue
            break

    practice_attempts_30d = len([row for row in practice_rows if row.created_at and row.created_at >= since_30d])
    practice_correct_30d = len(
        [row for row in practice_rows if row.created_at and row.created_at >= since_30d and bool(row.correct)]
    )
    practice_accuracy_30d = (practice_correct_30d / practice_attempts_30d) if practice_attempts_30d else 0.0

    practice_total_ms_14d = sum(max(0, int(row.duration_ms or 0)) for row in practice_14d)
    quiz_total_ms_14d = sum(max(0, int(row.duration_ms or 0)) for row in quiz_14d)
    video_total_seconds_14d = sum(max(0.0, float(row.watched_seconds or 0.0)) for row in video_14d)
    total_study_seconds_14d = (practice_total_ms_14d + quiz_total_ms_14d) / 1000.0 + video_total_seconds_14d

    video_started_30d = len(
        [
            row
            for row in video_rows
            if row.updated_at and row.updated_at >= since_30d and float(row.watched_seconds or 0.0) > 0
        ]
    )
    video_completed_30d = len(
        [row for row in video_rows if row.updated_at and row.updated_at >= since_30d and bool(row.completed)]
    )
    video_completion_values = [
        min(1.0, max(0.0, float(row.watched_seconds or 0.0) / float(row.duration_seconds)))
        for row in video_rows
        if row.updated_at and row.updated_at >= since_30d and float(row.duration_seconds or 0.0) > 0
    ]
    avg_video_completion_30d = (sum(video_completion_values) / len(video_completion_values)) if video_completion_values else 0.0

    event_type_counter = Counter((row.event_type or "").strip() or "unknown" for row in scoped_behavior_rows)
    top_event_types = [{"event_type": key, "count": int(count)} for key, count in event_type_counter.most_common(10) if key]

    portrait_summary = _json_load(snapshot.portrait_summary_json, {})
    final_dimensions, final_indicators, term_summary = _aggregate_stage_portrait_summary(stage_snapshots)
    if stage_snapshots:
        portrait_summary["final_portrait_dimensions"] = final_dimensions
        portrait_summary["final_portrait_indicators"] = final_indicators
        portrait_summary["term_summary"] = term_summary
    final_confirmation = None
    if course is not None and course.id is not None:
        final_confirmation = session.exec(
            select(TeacherFinalScoreConfirmation).where(
                TeacherFinalScoreConfirmation.user_id == user_id,
                TeacherFinalScoreConfirmation.course_id == int(course.id),
            )
        ).first()
    latest_recommendation = recommendation_items[0] if recommendation_items else None
    latest_target_kp = (
        session.get(KnowledgePoint, int(latest_recommendation["target_kp_id"]))
        if latest_recommendation and latest_recommendation.get("target_kp_id") is not None
        else None
    )

    return {
        "student": {
            "id": int(student.id),
            "username": student.username,
            "full_name": student.full_name,
            "student_no": student.student_no,
            "class_name": student.class_name,
        },
        "profile": {
            "course_id": int(course.id) if course is not None and course.id is not None else None,
            "persona_type": snapshot.persona_type.value,
            "persona_label": persona_label(snapshot.persona_type),
            "engagement": float(snapshot.engagement),
            "achievement": float(snapshot.achievement),
            "efficiency": float(snapshot.efficiency),
            "risk": float(snapshot.risk),
            "course_mastery": float(snapshot.course_mastery),
            "dynamic_score": float(snapshot.dynamic_score),
            "stability": float(snapshot.stability),
            "risk_level": snapshot.risk_level,
            "reason_summary": snapshot.reason_summary,
            "updated_at": snapshot.updated_at.isoformat(),
            "current_stage_title": current_stage.stage_title if current_stage is not None else "",
            "current_stage_trend": current_stage.trend_label if current_stage is not None else "",
            "portrait_dimensions": portrait_summary.get("portrait_dimensions", []),
            "portrait_indicators": portrait_summary.get("portrait_indicators", []),
            "final_portrait_dimensions": portrait_summary.get("final_portrait_dimensions", []),
            "final_portrait_indicators": portrait_summary.get("final_portrait_indicators", []),
            "term_summary": portrait_summary.get("term_summary", {}),
        },
        "stage_history": [
            {
                "stage_id": int(item.stage_id),
                "stage_title": item.stage_title,
                "stage_order": int(item.stage_order),
                "persona_type": item.persona_type.value if isinstance(item.persona_type, PersonaType) else str(item.persona_type),
                "persona_label": persona_label(item.persona_type),
                "engagement": float(item.engagement),
                "achievement": float(item.achievement),
                "habit": float(item.habit),
                "characteristic": float(item.characteristic),
                "dynamic_score": float(item.dynamic_score),
                "course_mastery": float(item.course_mastery),
                "trend_label": item.trend_label,
                "risk_level": item.risk_level,
                "reason_summary": item.reason_summary,
                "portrait_dimensions": _json_load(item.dimension_summary_json, {}).get("portrait_dimensions", []),
                "portrait_indicators": _json_load(item.indicator_summary_json, {}).get("portrait_indicators", []),
                "updated_at": item.updated_at.isoformat(),
            }
            for item in stage_snapshots
        ],
        "teacher_feedback": (
            {
                "stage_id": int(feedback.stage_id),
                "feedback_tag": feedback.feedback_tag,
                "comment": feedback.comment,
                "updated_by": feedback.updated_by,
                "updated_at": feedback.updated_at.isoformat(),
            }
            if feedback is not None and feedback.id is not None
            else None
        ),
        "mastery_map": mastery_items,
        "learning_behavior_overview": {
            "window_days": {"recent": 14, "history": 30},
            "login_count_30d": len(login_rows_30d),
            "login_days_30d": len(login_days_30d),
            "active_days_14d": len(active_days_14d),
            "consecutive_days_14d": int(consecutive_days),
            "study_duration_seconds_14d": round(float(total_study_seconds_14d), 2),
            "study_duration_minutes_14d": round(float(total_study_seconds_14d) / 60.0, 2),
            "video_started_30d": int(video_started_30d),
            "video_completed_30d": int(video_completed_30d),
            "avg_video_completion_30d": round(float(avg_video_completion_30d), 4),
            "practice_attempts_30d": int(practice_attempts_30d),
            "practice_accuracy_30d": round(float(practice_accuracy_30d), 4),
            "top_event_types_30d": top_event_types,
        },
        "behavior_timeline": behavior_items,
        "recommendations": recommendation_items,
        "recommendation_closure": {
            "total_recommendations": len(recommendation_items),
            "latest_target_kp_id": latest_recommendation["target_kp_id"] if latest_recommendation else None,
            "latest_target_kp_title": latest_target_kp.title if latest_target_kp is not None else "",
            "latest_target_kp_code": latest_target_kp.code if latest_target_kp is not None else "",
            "latest_reason_summary": latest_recommendation["reason_summary"] if latest_recommendation else "",
            "latest_created_at": latest_recommendation["created_at"] if latest_recommendation else None,
            "final_summary": (
                final_confirmation.recommendation_summary
                if final_confirmation is not None
                else (latest_recommendation["reason_summary"] if latest_recommendation else "")
            ),
        },
        "final_score_confirmation": (
            {
                "id": int(final_confirmation.id),
                "suggested_score": float(final_confirmation.suggested_score),
                "confirmed_score": float(final_confirmation.confirmed_score),
                "confirmed_level": final_confirmation.confirmed_level,
                "comment": final_confirmation.comment,
                "recommendation_summary": final_confirmation.recommendation_summary,
                "confirmed_by": final_confirmation.confirmed_by,
                "confirmed_at": final_confirmation.confirmed_at.isoformat(),
                "updated_at": final_confirmation.updated_at.isoformat(),
            }
            if final_confirmation is not None and final_confirmation.id is not None
            else None
        ),
        "recent_practice": [
            {
                "id": int(row.id),
                "kp_id": row.kp_id,
                "question_id": row.question_id,
                "correct": bool(row.correct),
                "duration_ms": row.duration_ms,
                "created_at": row.created_at.isoformat(),
            }
            for row in practice_rows
            if row.id is not None
        ],
        "recent_quiz": [
            {
                "id": int(row.id),
                "kp_id": row.kp_id,
                "score": float(row.score),
                "passed": bool(row.passed),
                "duration_ms": row.duration_ms,
                "created_at": row.created_at.isoformat(),
            }
            for row in quiz_rows
            if row.id is not None
        ],
        "recent_video": [
            {
                "id": int(row.id),
                "kp_id": row.kp_id,
                "resource_id": row.resource_id,
                "watched_seconds": float(row.watched_seconds),
                "duration_seconds": float(row.duration_seconds),
                "updated_at": row.updated_at.isoformat(),
            }
            for row in video_rows
            if row.id is not None
        ],
    }


def _relation_type_value(value) -> str:
    if isinstance(value, RelationType):
        return value.value
    if isinstance(value, str) and value:
        return value
    return RelationType.prerequisite.value


def _course_to_out(course: Course, *, teacher_name: str = "", teacher_ids: list[int] | None = None, teacher_names: list[str] | None = None) -> CourseOut:
    return CourseOut(
        id=int(course.id or 0),
        code=course.code,
        title=course.title,
        description=course.description,
        active=bool(course.active),
        lifecycle_status=course.lifecycle_status.value if hasattr(course.lifecycle_status, "value") else str(course.lifecycle_status or "draft"),
        teacher_id=int(course.teacher_id) if course.teacher_id is not None else None,
        teacher_name=teacher_name,
        teacher_ids=teacher_ids or ([int(course.teacher_id)] if course.teacher_id is not None else []),
        teacher_names=teacher_names or ([teacher_name] if teacher_name else []),
        archived_at=course.archived_at,
    )


def _safe_filename(name: str) -> str:
    name = re.sub(r"[^A-Za-z0-9._-]", "_", name)
    return name.strip("._") or "video.mp4"


def _validate_external_url(url: str) -> str:
    normalized = url.strip()
    parsed = urlparse(normalized)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(status_code=400, detail="外部链接必须是有效的 http/https 地址")
    return normalized


def _resource_type_value(value) -> str:
    if isinstance(value, ResourceType):
        return value.value
    if isinstance(value, str) and value:
        return value
    return ResourceType.note.value


def _resource_category_value(row: LearningResource) -> str:
    if getattr(row, "category", ""):
        return str(row.category)
    type_value = _resource_type_value(getattr(row, "type", ""))
    return "recommend" if type_value in {ResourceType.book.value, ResourceType.recommend_book.value} else "learning"


def _task_type_value(value) -> str:
    if isinstance(value, KpTaskType):
        return value.value
    if isinstance(value, str) and value:
        return value
    return KpTaskType.task.value


@router.put("/kp-video/bilibili")
def set_kp_bilibili_video(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp_id = int(payload.get("kp_id"))
    kp = _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    bvid = str(payload.get("bvid", "")).strip()
    page = int(payload.get("page", 1))
    title = str(payload.get("title", "")).strip() or f"B站视频 {bvid} P{page}"
    if not bvid:
        raise HTTPException(status_code=400, detail="bvid required")

    url = _bilibili_embed_url(bvid=bvid, page=page)
    r = _replace_kp_video(session=session, kp=kp, title=title, url=url)
    _log_action(session, admin, "kp_video_bind_bilibili", f"kp_id={kp_id} bvid={bvid} page={page}")
    return {"ok": True, "resource_id": r.id, "url": r.url}


@router.delete("/kp-video")
def clear_kp_video(
    kp_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    existing = session.exec(
        select(LearningResource).where(LearningResource.kp_id == kp_id, LearningResource.type == ResourceType.video)
    ).all()
    for r in existing:
        session.delete(r)
    session.commit()
    _log_action(session, admin, "kp_video_clear", f"kp_id={kp_id} deleted={len(existing)}")
    return {"ok": True, "deleted": len(existing)}


@router.post("/kp-video/local")
def upload_kp_video_local(
    kp_id: int = Form(...),
    title: str = Form(""),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)

    filename = _safe_filename(file.filename or "video.mp4")
    ext = Path(filename).suffix.lower()
    if ext not in {".mp4", ".m3u8", ".m4v", ".mov"}:
        raise HTTPException(status_code=400, detail="Unsupported video format")

    try:
        stored = store_video_file_stream(filename=f"{kp_id}_{filename}", source=file.file)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    stored_name = stored["stored_name"]
    final_title = title.strip() or f"本地视频：{stored_name}"
    url = stored["relative_url"]
    r = _replace_kp_video(session=session, kp=kp, title=final_title, url=url)
    _log_action(session, admin, "kp_video_upload_local", f"kp_id={kp_id} file={stored['stored_name']} size={stored['file_size_bytes']}")
    return {"ok": True, "resource_id": r.id, "url": r.url}


@router.put("/kp-video/url")
def set_kp_video_url(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp_id = int(payload.get("kp_id"))
    kp = _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    url = str(payload.get("url", "")).strip()
    title = str(payload.get("title", "")).strip() or "自托管视频"
    if not url:
        raise HTTPException(status_code=400, detail="url required")

    r = _replace_kp_video(session=session, kp=kp, title=title, url=url)
    _log_action(session, admin, "kp_video_bind_url", f"kp_id={kp_id} url={url}")
    return {"ok": True, "resource_id": r.id, "url": r.url}


@router.get("/kp-resources")
def list_kp_resources(
    kp_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    rows = session.exec(select(LearningResource).where(LearningResource.kp_id == kp_id).order_by(LearningResource.id)).all()
    return [build_resource_payload(row) for row in rows if row.id is not None]


@router.get("/kp-resources/{resource_id}/detail")
def get_kp_resource_detail(
    resource_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    row = session.get(LearningResource, resource_id)
    _require_teacher_resource_access(session=session, admin=admin, row=row)
    kp = session.get(KnowledgePoint, int(row.kp_id))
    payload = build_resource_payload(row)
    payload.update(
        {
            "subject": row.subject,
            "grade": row.grade,
            "kp_code": kp.code if kp else "",
            "kp_title": kp.title if kp else "",
        }
    )
    return payload


@router.post("/kp-resources/detect")
def detect_kp_resource_upload(
    file: UploadFile = File(...),
    _admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    try:
        detected = inspect_uploaded_file_stream(
            filename=file.filename or "resource",
            source=file.file,
            content_type=file.content_type,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {
        **detected,
        "preview_label": {
            "pdf_inline": "PDF 在线预览",
            "video_inline": "页面内视频播放",
            "image_inline": "页面内图片预览",
            "pdf_after_convert": "转换为 PDF 后在线预览",
            "external_link": "新窗口打开",
            "download": "下载查看",
        }.get(detected["preview_type"], "下载查看"),
    }


@router.post("/kp-resources/upload")
def upload_kp_resource(
    kp_id: int = Form(...),
    title: str = Form(""),
    category: str = Form("learning"),
    tags: str = Form(""),
    description: str = Form(""),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    try:
        stored = store_uploaded_file_stream(
            kp_id=kp_id,
            filename=file.filename or "resource",
            content_type=file.content_type,
            source=file.file,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    detected = stored["detected"]
    type_value = _resource_type_from_detected(detected["detected_resource_type"], category=category)
    preview_type = detected["preview_type"]
    preview_status = "processing" if preview_type == "pdf_after_convert" else "ready"
    final_title = title.strip() or Path(file.filename or "resource").stem
    row = LearningResource(
        subject=kp.subject,
        grade=kp.grade,
        kp_id=int(kp.id),
        title=final_title,
        url="",
        type=type_value,
        category=category if category in {"learning", "recommend"} else "learning",
        description=description.strip(),
        tags=tags.strip(),
        original_file_name=detected["original_file_name"],
        file_extension=detected["file_extension"],
        detected_mime_type=detected["detected_mime_type"],
        detected_resource_type=detected["detected_resource_type"],
        preview_type=preview_type,
        preview_status=preview_status,
        converted_preview_url="",
        original_file_url=stored["relative_url"],
        file_size_bytes=stored["file_size_bytes"],
        extension_mismatch=bool(detected["extension_mismatch"]),
        source_kind="upload",
    )
    if preview_type in {"pdf_inline", "video_inline", "image_inline"}:
        row.converted_preview_url = stored["relative_url"]
        row.url = stored["relative_url"]
    else:
        row.url = stored["relative_url"]
    session.add(row)
    session.commit()
    session.refresh(row)
    maybe_prepare_preview(resource=row)
    row.updated_at = datetime.utcnow()
    if row.preview_status == "ready" and row.converted_preview_url:
        row.url = row.converted_preview_url
    session.add(row)
    session.commit()
    session.refresh(row)
    _log_action(session, admin, "kp_resource_upload", f"kp_id={kp.id} resource_id={row.id} type={row.detected_resource_type}")
    return build_resource_payload(row)


@router.post("/kp-resources")
def create_kp_resource(
    payload: KpResourceIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, payload.kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    try:
        resource_type = ResourceType(str(payload.type or "note").strip())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid resource type") from exc
    title = payload.title.strip()
    url = _validate_external_url(payload.url)
    if not title or not url:
        raise HTTPException(status_code=400, detail="title/url required")
    row = LearningResource(
        subject=kp.subject,
        grade=kp.grade,
        kp_id=int(kp.id),
        title=title,
        url=url,
        type=resource_type,
        category=payload.category if payload.category in {"learning", "recommend"} else "learning",
        tags=payload.tags.strip(),
        description=payload.description.strip(),
        detected_resource_type=_resource_type_value(resource_type),
        preview_type="external_link",
        preview_status="ready",
        original_file_url=url,
        converted_preview_url="",
        source_kind="external",
    )
    row.url = url
    session.add(row)
    session.commit()
    session.refresh(row)
    _log_action(session, admin, "kp_resource_create", f"kp_id={kp.id} resource_id={row.id}")
    return build_resource_payload(row)


@router.put("/kp-resources/{resource_id}")
def update_kp_resource(
    resource_id: int,
    payload: KpResourceUpdateIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    row = session.get(LearningResource, resource_id)
    _require_teacher_resource_access(session=session, admin=admin, row=row)
    if payload.title is not None:
        title = payload.title.strip()
        if not title:
            raise HTTPException(status_code=400, detail="title required")
        row.title = title
    if payload.description is not None:
        row.description = payload.description.strip()
    if payload.tags is not None:
        row.tags = payload.tags.strip()
    if payload.category is not None and payload.category in {"learning", "recommend"}:
        row.category = payload.category
    if payload.url is not None:
        url = _validate_external_url(payload.url)
        if not url:
            raise HTTPException(status_code=400, detail="url required")
        row.url = url
        if (row.source_kind or "external") == "external":
            row.original_file_url = url
            row.converted_preview_url = ""
    if payload.type is not None:
        if (row.source_kind or "external") == "upload":
            raise HTTPException(status_code=400, detail="Uploaded file resources cannot change file type manually")
        try:
            row.type = ResourceType(str(payload.type).strip())
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid resource type") from exc
        row.detected_resource_type = _resource_type_value(row.type)
        row.preview_type = "external_link" if row.type == ResourceType.link else _preview_type_for_detected(_resource_type_value(row.type))
    if (row.source_kind or "external") == "external":
        row.preview_status = "ready"
        row.preview_type = "external_link"
        row.converted_preview_url = ""
        row.extension_mismatch = False
    row.updated_at = datetime.utcnow()
    session.add(row)
    session.commit()
    _log_action(session, admin, "kp_resource_update", f"resource_id={resource_id}")
    session.refresh(row)
    return build_resource_payload(row)


@router.delete("/kp-resources/{resource_id}")
def delete_kp_resource(
    resource_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    row = session.get(LearningResource, resource_id)
    _require_teacher_resource_access(session=session, admin=admin, row=row)
    session.delete(row)
    session.commit()
    _log_action(session, admin, "kp_resource_delete", f"resource_id={resource_id}")
    return {"ok": True}


@router.get("/kp-tasks")
def list_kp_tasks(
    kp_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    rows = session.exec(select(KpTask).where(KpTask.kp_id == kp_id).order_by(KpTask.sort_order, KpTask.id)).all()
    return [
        {
            "id": int(row.id),
            "kp_id": int(row.kp_id),
            "type": _task_type_value(row.type),
            "title": row.title,
            "description": row.description,
            "link_url": row.link_url,
            "sort_order": row.sort_order,
        }
        for row in rows
        if row.id is not None
    ]


@router.post("/kp-tasks")
def create_kp_task(
    payload: KpTaskIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, payload.kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    try:
        task_type = KpTaskType(str(payload.type or "task").strip())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid task type") from exc
    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="title required")
    row = KpTask(
        subject=kp.subject,
        grade=kp.grade,
        kp_id=int(kp.id),
        title=title,
        description=payload.description.strip(),
        link_url=payload.link_url.strip(),
        type=task_type,
        sort_order=max(0, int(payload.sort_order)),
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    _log_action(session, admin, "kp_task_create", f"kp_id={kp.id} task_id={row.id}")
    return {"ok": True, "id": row.id}


@router.put("/kp-tasks/{task_id}")
def update_kp_task(
    task_id: int,
    payload: KpTaskUpdateIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    row = session.get(KpTask, task_id)
    _require_teacher_task_access(session=session, admin=admin, row=row)
    if payload.title is not None:
        title = payload.title.strip()
        if not title:
            raise HTTPException(status_code=400, detail="title required")
        row.title = title
    if payload.description is not None:
        row.description = payload.description
    if payload.link_url is not None:
        row.link_url = payload.link_url
    if payload.sort_order is not None:
        row.sort_order = max(0, int(payload.sort_order))
    if payload.type is not None:
        try:
            row.type = KpTaskType(str(payload.type).strip())
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid task type") from exc
    session.add(row)
    session.commit()
    _log_action(session, admin, "kp_task_update", f"task_id={task_id}")
    return {"ok": True}


@router.delete("/kp-tasks/{task_id}")
def delete_kp_task(
    task_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    row = session.get(KpTask, task_id)
    _require_teacher_task_access(session=session, admin=admin, row=row)
    session.delete(row)
    session.commit()
    _log_action(session, admin, "kp_task_delete", f"task_id={task_id}")
    return {"ok": True}


@router.post("/users")
def create_user(
    payload: dict,
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin)),
):
    username = payload.get("username", "").strip()
    password = payload.get("password", "").strip()
    role = payload.get("role", "student")
    full_name = payload.get("full_name", "") or ""
    student_no = payload.get("student_no", "") or ""
    class_name = payload.get("class_name", "") or ""
    phone = payload.get("phone", None)
    active = bool(payload.get("active", True))
    if not username or not password:
        raise HTTPException(status_code=400, detail="Invalid user")
    # Password format validation disabled for testing.
    exists = session.exec(select(User).where(User.username == username)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Username already exists")
    if phone:
        exists_phone = session.exec(select(User).where(User.phone == str(phone).strip())).first()
        if exists_phone:
            raise HTTPException(status_code=400, detail="Phone already exists")
    user = User(
        username=username,
        password_hash=hash_password(password),
        role=UserRole(role),
        active=active,
        full_name=str(full_name),
        student_no=str(student_no),
        class_name=str(class_name),
        phone=str(phone).strip() if phone else None,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    if user.role == UserRole.student:
        _sync_student_class_enrollments_for_user(session, user)
    _log_action(session, _admin, "user_create", f"username={username} role={role}")
    return {"ok": True, "user_id": user.id}


@router.get("/users/import-template")
def download_user_import_template(
    role: str = UserRole.student.value,
    _admin=Depends(require_role(UserRole.admin)),
):
    role_value = str(role or UserRole.student.value).strip().lower()
    if role_value not in {UserRole.student.value, UserRole.teacher.value}:
        raise HTTPException(status_code=400, detail="role must be student or teacher")
    content = _user_import_template_csv(role_value)
    filename = f"{role_value}_import_template.csv"
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/users/import", response_model=UserImportResultOut)
def import_users(
    role: str = Form(UserRole.student.value),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin)),
):
    role_value = str(role or UserRole.student.value).strip().lower()
    if role_value not in {UserRole.student.value, UserRole.teacher.value}:
        raise HTTPException(status_code=400, detail="role must be student or teacher")

    payload = file.file.read()
    rows = _user_rows_from_upload(file, payload)
    total_rows = len(rows)
    success_rows = 0
    failed_rows = 0
    created_rows = 0
    updated_rows = 0
    auto_enrolled_rows = 0
    errors: list[str] = []

    for index, row in enumerate(rows, start=2):
        username = (row.get("username") or "").strip()
        password = (row.get("password") or "").strip()
        phone = (row.get("phone") or "").strip() or None
        try:
            if not username:
                raise ValueError("username is required")
            existing = session.exec(select(User).where(User.username == username)).first()
            phone_owner = session.exec(select(User).where(User.phone == phone)).first() if phone else None
            if phone_owner is not None and (existing is None or int(phone_owner.id) != int(existing.id)):
                raise ValueError(f"phone already exists: {phone}")

            if existing is None:
                if not password:
                    raise ValueError("password is required for new user")
                user = User(
                    username=username,
                    password_hash=hash_password(password),
                    role=UserRole(role_value),
                    active=_to_active_flag(row.get("active")),
                    full_name=(row.get("full_name") or "").strip(),
                    student_no=(row.get("student_no") or "").strip(),
                    class_name=(row.get("class_name") or "").strip(),
                    phone=phone,
                )
                session.add(user)
                session.commit()
                session.refresh(user)
                if user.role == UserRole.student:
                    auto_enrolled_rows += _sync_student_class_enrollments_for_user(session, user)
                created_rows += 1
            else:
                existing.role = UserRole(role_value)
                existing.active = _to_active_flag(row.get("active")) if (row.get("active") or "").strip() else bool(existing.active)
                if row.get("full_name") is not None:
                    existing.full_name = (row.get("full_name") or "").strip()
                if row.get("student_no") is not None:
                    existing.student_no = (row.get("student_no") or "").strip()
                if row.get("class_name") is not None:
                    existing.class_name = (row.get("class_name") or "").strip()
                existing.phone = phone
                if password:
                    existing.password_hash = hash_password(password)
                session.add(existing)
                session.commit()
                session.refresh(existing)
                if existing.role == UserRole.student:
                    auto_enrolled_rows += _sync_student_class_enrollments_for_user(session, existing)
                updated_rows += 1
            success_rows += 1
        except Exception as exc:
            session.rollback()
            failed_rows += 1
            errors.append(f"row {index}: {exc}")

    _log_action(
        session,
        admin,
        "user_import",
        f"role={role_value} total={total_rows} success={success_rows} failed={failed_rows}",
    )
    return UserImportResultOut(
        role=role_value,
        total_rows=total_rows,
        success_rows=success_rows,
        failed_rows=failed_rows,
        created_rows=created_rows,
        updated_rows=updated_rows,
        auto_enrolled_rows=auto_enrolled_rows,
        errors=errors[:20],
    )


@router.post("/users/import/preview", response_model=UserImportPreviewOut)
def preview_user_import(
    role: str = Form(UserRole.student.value),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin)),
):
    role_value = str(role or UserRole.student.value).strip().lower()
    if role_value not in {UserRole.student.value, UserRole.teacher.value}:
        raise HTTPException(status_code=400, detail="role must be student or teacher")
    payload = file.file.read()
    rows = _user_rows_from_upload(file, payload)
    return _preview_user_import_rows(session=session, role_value=role_value, rows=rows)


@router.get("/courses")
def list_courses(
    page: int = 1,
    page_size: int = 15,
    keyword: str | None = None,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin)),
):
    q = select(Course)
    if keyword:
        like = f"%{keyword.strip()}%"
        q = q.where(or_(Course.code.like(like), Course.title.like(like), Course.description.like(like)))
    total = session.exec(select(func.count()).select_from(q.subquery())).one()
    items = session.exec(q.order_by(Course.created_at.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    course_ids = [int(item.id) for item in items if item.id is not None]
    activation_map: dict[int, list[CourseTeacherActivation]] = {}
    teacher_name_map: dict[int, str] = {}
    if course_ids:
        activation_rows = session.exec(
            select(CourseTeacherActivation).where(CourseTeacherActivation.course_id.in_(course_ids))
        ).all()
        activation_map = {course_id: [] for course_id in course_ids}
        teacher_ids = set()
        for item in items:
            if item.teacher_id is not None:
                teacher_ids.add(int(item.teacher_id))
        for row in activation_rows:
            activation_map.setdefault(int(row.course_id), []).append(row)
            teacher_ids.add(int(row.teacher_id))
        if teacher_ids:
            teacher_rows = session.exec(select(User).where(User.id.in_(teacher_ids))).all()
            teacher_name_map = {int(row.id): row.full_name or row.username for row in teacher_rows if row.id is not None}

    payload_items = []
    for item in items:
        activations = activation_map.get(int(item.id or 0), [])
        assigned_teacher_ids = [int(row.teacher_id) for row in activations if row.teacher_id is not None]
        if item.teacher_id is not None and int(item.teacher_id) not in assigned_teacher_ids:
            assigned_teacher_ids.insert(0, int(item.teacher_id))
        assigned_teacher_names = [teacher_name_map.get(teacher_id, f"教师{teacher_id}") for teacher_id in assigned_teacher_ids]
        assigned_teacher_name = assigned_teacher_names[0] if assigned_teacher_names else ""
        base = _course_to_out(
            item,
            teacher_name=assigned_teacher_name,
            teacher_ids=assigned_teacher_ids,
            teacher_names=assigned_teacher_names,
        ).model_dump()
        teaching_names = [
            teacher_name_map.get(int(row.teacher_id), f"教师{row.teacher_id}")
            for row in activations
            if row.teaching_status == TeacherCourseStatus.teaching
        ]
        finished_names = [
            teacher_name_map.get(int(row.teacher_id), f"教师{row.teacher_id}")
            for row in activations
            if row.teaching_status == TeacherCourseStatus.finished
        ]
        base.update(
            {
                "teaching_teacher_count": len(teaching_names),
                "finished_teacher_count": len(finished_names),
                "teaching_teacher_names": teaching_names,
                "finished_teacher_names": finished_names,
            }
        )
        payload_items.append(base)
    return {"items": payload_items, "total": total}


@router.post("/courses", response_model=CourseOut)
def create_course(
    payload: CourseIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin)),
):
    code = payload.code.strip()
    title = payload.title.strip()
    if not code or not title:
        raise HTTPException(status_code=400, detail="code/title required")
    exists = session.exec(select(Course).where(Course.code == code)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Course code exists")
    teacher_ids = _normalize_teacher_ids(session, payload.teacher_ids, payload.teacher_id)
    teacher_id = teacher_ids[0] if teacher_ids else None
    try:
        lifecycle_status = CourseLifecycleStatus(str(payload.lifecycle_status or "draft").strip().lower())
    except ValueError:
        raise HTTPException(status_code=400, detail="无效的课程生命周期状态")
    course = Course(
        code=code,
        title=title,
        description=payload.description or "",
        active=bool(payload.active),
        lifecycle_status=lifecycle_status,
        teacher_id=teacher_id,
        archived_at=datetime.utcnow() if lifecycle_status == CourseLifecycleStatus.archived else None,
    )
    session.add(course)
    session.commit()
    session.refresh(course)
    _sync_course_teacher_assignments(session, course, teacher_ids)
    session.commit()
    _log_action(session, admin, "course_create", f"code={code} title={title}")
    teacher_rows = session.exec(select(User).where(User.id.in_(teacher_ids))).all() if teacher_ids else []
    teacher_name_map = {int(row.id): row.full_name or row.username for row in teacher_rows if row.id is not None}
    teacher_names = [teacher_name_map.get(teacher_id, f"教师{teacher_id}") for teacher_id in teacher_ids]
    return _course_to_out(course, teacher_name=teacher_names[0] if teacher_names else "", teacher_ids=teacher_ids, teacher_names=teacher_names)


@router.put("/courses/{course_id}", response_model=CourseOut)
def update_course(
    course_id: int,
    payload: CourseUpdateIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin)),
):
    course = session.get(Course, course_id)
    if course is None:
        raise HTTPException(status_code=404, detail="Course not found")
    if payload.code is not None:
        code = payload.code.strip()
        if not code:
            raise HTTPException(status_code=400, detail="code required")
        exists = session.exec(select(Course).where(Course.code == code, Course.id != course_id)).first()
        if exists:
            raise HTTPException(status_code=400, detail="Course code exists")
        course.code = code
    if payload.title is not None:
        title = payload.title.strip()
        if not title:
            raise HTTPException(status_code=400, detail="title required")
        course.title = title
    if payload.description is not None:
        course.description = payload.description
    if payload.active is not None:
        course.active = bool(payload.active)
    if payload.lifecycle_status is not None:
        try:
            course.lifecycle_status = CourseLifecycleStatus(str(payload.lifecycle_status).strip().lower())
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的课程生命周期状态")
        if course.lifecycle_status == CourseLifecycleStatus.archived:
            course.archived_at = payload.archived_at or datetime.utcnow()
            course.active = False
        elif course.lifecycle_status == CourseLifecycleStatus.active:
            course.archived_at = None
            course.active = True
    if payload.archived_at is not None:
        course.archived_at = payload.archived_at
    if payload.teacher_ids is not None:
        teacher_ids = _normalize_teacher_ids(session, payload.teacher_ids, payload.teacher_id)
        course.teacher_id = teacher_ids[0] if teacher_ids else None
    elif payload.teacher_id is not None:
        teacher_ids = _normalize_teacher_ids(session, None, payload.teacher_id)
        course.teacher_id = teacher_ids[0] if teacher_ids else None
    else:
        teacher_ids = _course_teacher_ids(session, course)
    session.add(course)
    session.commit()
    session.refresh(course)
    if payload.teacher_ids is not None or payload.teacher_id is not None:
        _sync_course_teacher_assignments(session, course, teacher_ids)
        session.commit()
    _log_action(session, admin, "course_update", f"id={course_id} code={course.code}")
    teacher_ids = _course_teacher_ids(session, course)
    teacher_rows = session.exec(select(User).where(User.id.in_(teacher_ids))).all() if teacher_ids else []
    teacher_name_map = {int(row.id): row.full_name or row.username for row in teacher_rows if row.id is not None}
    teacher_names = [teacher_name_map.get(teacher_id, f"教师{teacher_id}") for teacher_id in teacher_ids]
    return _course_to_out(course, teacher_name=teacher_names[0] if teacher_names else "", teacher_ids=teacher_ids, teacher_names=teacher_names)


@router.delete("/courses/{course_id}")
def delete_course(
    course_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin)),
):
    course = session.get(Course, course_id)
    if course is None:
        raise HTTPException(status_code=404, detail="Course not found")
    course_code = course.code
    activation_rows = session.exec(
        select(CourseTeacherActivation).where(CourseTeacherActivation.course_id == course_id)
    ).all()
    unfinished_rows = [
        row
        for row in activation_rows
        if (row.teaching_status.value if hasattr(row.teaching_status, "value") else str(row.teaching_status or ""))
        != TeacherCourseStatus.finished.value
    ]
    if unfinished_rows and (
        not _course_has_active_students(session, course_id=course_id)
        or _all_active_students_final_confirmed(session, course_id=course_id)
    ):
        _mark_course_teaching_finished(session, course=course)
        session.flush()
        activation_rows = session.exec(
            select(CourseTeacherActivation).where(CourseTeacherActivation.course_id == course_id)
        ).all()
        unfinished_rows = [
            row
            for row in activation_rows
            if (row.teaching_status.value if hasattr(row.teaching_status, "value") else str(row.teaching_status or ""))
            != TeacherCourseStatus.finished.value
        ]
    if unfinished_rows:
        teacher_ids = [int(row.teacher_id) for row in unfinished_rows]
        teacher_rows = session.exec(select(User).where(User.id.in_(teacher_ids))).all() if teacher_ids else []
        teacher_names = [row.full_name or row.username for row in teacher_rows]
        teacher_suffix = f"：{ '、'.join(teacher_names) }" if teacher_names else ""
        raise HTTPException(status_code=400, detail=f"仍有教师未结课，当前课程不能删除{teacher_suffix}")
    try:
        _purge_course_related_rows(session, course_id=course_id)
        session.delete(course)
        session.commit()
    except Exception as exc:
        session.rollback()
        raise HTTPException(status_code=400, detail=f"删除失败：课程仍有关联数据未清理（{exc}）") from exc
    _log_action(session, admin, "course_delete", f"id={course_id} code={course_code}")
    return {"ok": True}

@router.get("/users")
def list_users(
    page: int = 1,
    page_size: int = 15,
    role: str | None = None,
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin)),
):
    page = max(1, page)
    page_size = max(1, min(100, page_size))
    q = select(User)
    q_total = select(func.count()).select_from(User)
    if role:
        role_value = role.strip()
        if role_value not in {UserRole.admin.value, UserRole.teacher.value, UserRole.student.value}:
            raise HTTPException(status_code=400, detail="Invalid role filter")
        q = q.where(User.role == UserRole(role_value))
        q_total = q_total.where(User.role == UserRole(role_value))
    total = session.exec(q_total).one()
    rows = session.exec(
        q.order_by(User.id).offset((page - 1) * page_size).limit(page_size)
    ).all()
    items = [
        UserOut(
            id=u.id,
            username=u.username,
            role=u.role.value,
            active=bool(u.active),
            full_name=u.full_name,
            student_no=u.student_no,
            class_name=u.class_name,
            phone=u.phone,
            wechat_openid=u.wechat_openid,
        )
        for u in rows
    ]
    return {"items": [i.model_dump() for i in items], "total": int(total or 0), "page": page, "page_size": page_size}


@router.get("/practice/report", response_model=AdminPracticeReportOut)
def practice_report(
    user_id: int,
    kp_id: int | None = None,
    days: int = 14,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    days = max(1, min(180, int(days)))
    if admin.role == UserRole.teacher and kp_id is None:
        raise HTTPException(status_code=400, detail="kp_id required for teacher practice report")
    if kp_id is not None:
        _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    since = datetime.utcnow() - timedelta(days=days)

    q = (
        select(PracticeAttempt, Question)
        .join(Question, PracticeAttempt.question_id == Question.id, isouter=True)
        .where(PracticeAttempt.user_id == user_id, PracticeAttempt.created_at >= since)
    )
    if kp_id is not None:
        q = q.where(PracticeAttempt.kp_id == kp_id)
    rows = session.exec(q).all()

    total = len(rows)
    correct = sum(1 for a, _ in rows if a.correct)
    incorrect = total - correct
    accuracy = (correct / total) if total else 0.0

    buckets: dict[str, dict] = {}
    by_kp: dict[int, dict] = {}
    for attempt, _question in rows:
        day = attempt.created_at.date().isoformat()
        if day not in buckets:
            buckets[day] = {"date": day, "total": 0, "correct": 0}
        buckets[day]["total"] += 1
        if attempt.correct:
            buckets[day]["correct"] += 1

        kp = int(attempt.kp_id)
        if kp not in by_kp:
            by_kp[kp] = {"kp_id": kp, "total": 0, "correct": 0}
        by_kp[kp]["total"] += 1
        if attempt.correct:
            by_kp[kp]["correct"] += 1

    daily = []
    for day in sorted(buckets.keys()):
        b = buckets[day]
        daily.append(
            {
                "date": b["date"],
                "total": b["total"],
                "correct": b["correct"],
                "accuracy": (b["correct"] / b["total"]) if b["total"] else 0.0,
            }
        )

    kp_ids = list(by_kp.keys())
    kp_map: dict[int, KnowledgePoint] = {}
    if kp_ids:
        kps = session.exec(select(KnowledgePoint).where(KnowledgePoint.id.in_(kp_ids))).all()
        kp_map = {int(k.id): k for k in kps}
    by_kp_list = []
    for item in by_kp.values():
        kp = kp_map.get(int(item["kp_id"]))
        by_kp_list.append(
            {
                "kp_id": item["kp_id"],
                "kp_code": kp.code if kp else "",
                "kp_title": kp.title if kp else "",
                "total": item["total"],
                "correct": item["correct"],
                "accuracy": (item["correct"] / item["total"]) if item["total"] else 0.0,
            }
        )
    by_kp_list.sort(key=lambda x: x["total"], reverse=True)

    return AdminPracticeReportOut(
        user_id=user_id,
        kp_id=kp_id,
        total=total,
        correct=correct,
        incorrect=incorrect,
        accuracy=accuracy,
        daily=daily,
        by_kp=by_kp_list,
    )
@router.put("/users/{user_id}", response_model=UserOut)
def update_user(
    user_id: int,
    payload: UserUpdateIn,
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin)),
):
    u = session.get(User, user_id)
    if u is None:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.role is not None:
        u.role = UserRole(payload.role)
    if payload.active is not None:
        u.active = bool(payload.active)
    if payload.password is not None and payload.password.strip():
        pw = payload.password.strip()
        if len(pw) < 8 or not re.search(r"[A-Z]", pw) or not re.search(r"[a-z]", pw) or not re.search(r"\d", pw):
            raise HTTPException(status_code=400, detail="密码至少8位，需包含大小写字母和数字")
        u.password_hash = hash_password(pw)
    if payload.full_name is not None:
        u.full_name = payload.full_name
    if payload.student_no is not None:
        u.student_no = payload.student_no
    if payload.class_name is not None:
        u.class_name = payload.class_name
    if payload.phone is not None:
        phone = payload.phone.strip() or None
        u.phone = phone
    session.add(u)
    session.commit()
    session.refresh(u)
    if u.role == UserRole.student:
        _sync_student_class_enrollments_for_user(session, u)
    _log_action(session, _admin, "user_update", f"user_id={user_id}")
    return UserOut(
        id=u.id,
        username=u.username,
        role=u.role.value,
        active=bool(u.active),
        full_name=u.full_name,
        student_no=u.student_no,
        class_name=u.class_name,
        phone=u.phone,
        wechat_openid=u.wechat_openid,
    )


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin)),
):
    u = session.get(User, user_id)
    if u is None:
        raise HTTPException(status_code=404, detail="User not found")
    session.delete(u)
    session.commit()
    _log_action(session, _admin, "user_delete", f"user_id={user_id} username={u.username}")
    return {"ok": True}


@router.get("/kps")
def list_kps_admin(
    subject: str | None = None,
    grade: str | None = None,
    keyword: str | None = None,
    page: int = 1,
    page_size: int = 15,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    page = max(1, page)
    page_size = max(1, min(100, page_size))
    q = select(KnowledgePoint).order_by(KnowledgePoint.id.desc())
    q_total = select(func.count()).select_from(KnowledgePoint)
    allowed_subjects = _teacher_accessible_subjects(session, admin)
    if allowed_subjects is not None:
        if not allowed_subjects:
            return {"items": [], "total": 0, "page": page, "page_size": page_size}
        allowed_subject_list = sorted(allowed_subjects)
        q = q.where(KnowledgePoint.subject.in_(allowed_subject_list))
        q_total = q_total.where(KnowledgePoint.subject.in_(allowed_subject_list))
    if subject:
        _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
        q = q.where(KnowledgePoint.subject == subject)
        q_total = q_total.where(KnowledgePoint.subject == subject)
    if grade:
        q = q.where(KnowledgePoint.grade == grade)
        q_total = q_total.where(KnowledgePoint.grade == grade)
    if keyword:
        kw = keyword.strip()
        if kw:
            q = q.where(
                or_(
                    KnowledgePoint.code.contains(kw),
                    KnowledgePoint.title.contains(kw),
                    KnowledgePoint.description.contains(kw),
                )
            )
            q_total = q_total.where(
                or_(
                    KnowledgePoint.code.contains(kw),
                    KnowledgePoint.title.contains(kw),
                    KnowledgePoint.description.contains(kw),
                )
            )
    total = session.exec(q_total).one()
    rows = session.exec(q.offset((page - 1) * page_size).limit(page_size)).all()
    return {"items": [r.model_dump() for r in rows], "total": int(total or 0), "page": page, "page_size": page_size}


@router.post("/kps")
def create_kp(
    payload: KnowledgePointIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_subject_access(session=session, admin=admin, subject=payload.subject, grade=payload.grade)
    code = payload.code.strip()
    if not code:
        raise HTTPException(status_code=400, detail="code required")
    exists = session.exec(select(KnowledgePoint).where(KnowledgePoint.code == code)).first()
    if exists:
        raise HTTPException(status_code=400, detail="code already exists")
    kp = KnowledgePoint(
        subject=payload.subject.strip(),
        grade=payload.grade.strip(),
        code=code,
        title=payload.title.strip(),
        description=payload.description.strip(),
        chapter=payload.chapter.strip(),
        knowledge_tag=payload.knowledge_tag.strip(),
        ability_tag=payload.ability_tag.strip(),
        literacy_tag=payload.literacy_tag.strip(),
        importance=max(0.0, min(1.0, float(payload.importance))),
        difficulty=max(0.0, min(1.0, float(payload.difficulty))),
        pos_x=float(payload.pos_x) if payload.pos_x is not None else None,
        pos_y=float(payload.pos_y) if payload.pos_y is not None else None,
        is_terminal=bool(payload.is_terminal),
    )
    session.add(kp)
    session.commit()
    session.refresh(kp)
    return kp


@router.put("/kps/{kp_id}")
def update_kp(
    kp_id: int,
    payload: KnowledgePointUpdateIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    if payload.code is not None:
        code = payload.code.strip()
        if not code:
            raise HTTPException(status_code=400, detail="code required")
        exists = session.exec(select(KnowledgePoint).where(KnowledgePoint.code == code, KnowledgePoint.id != kp_id)).first()
        if exists:
            raise HTTPException(status_code=400, detail="code already exists")
        kp.code = code
    if payload.title is not None:
        kp.title = payload.title
    if payload.description is not None:
        kp.description = payload.description
    if payload.chapter is not None:
        kp.chapter = payload.chapter
    if payload.knowledge_tag is not None:
        kp.knowledge_tag = payload.knowledge_tag
    if payload.ability_tag is not None:
        kp.ability_tag = payload.ability_tag
    if payload.literacy_tag is not None:
        kp.literacy_tag = payload.literacy_tag
    if payload.importance is not None:
        kp.importance = max(0.0, min(1.0, float(payload.importance)))
    if payload.difficulty is not None:
        kp.difficulty = max(0.0, min(1.0, float(payload.difficulty)))
    if payload.pos_x is not None:
        kp.pos_x = float(payload.pos_x)
    if payload.pos_y is not None:
        kp.pos_y = float(payload.pos_y)
    if payload.is_terminal is not None:
        kp.is_terminal = bool(payload.is_terminal)
    session.add(kp)
    session.commit()
    session.refresh(kp)
    return kp


@router.put("/kps/{kp_id}/position")
def update_kp_position(
    kp_id: int,
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    if "x" not in payload or "y" not in payload:
        raise HTTPException(status_code=400, detail="x/y required")
    try:
        kp.pos_x = float(payload.get("x"))
        kp.pos_y = float(payload.get("y"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="x/y must be number") from exc
    session.add(kp)
    session.commit()
    return {"ok": True, "kp_id": kp_id, "x": kp.pos_x, "y": kp.pos_y}


@router.put("/graph/chapter-layout")
def save_graph_chapter_layout(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    """Persist chapter (category) anchor positions so student and teacher share the same graph layout."""
    subject = payload.get("subject")
    grade = payload.get("grade")
    chapters = payload.get("chapters")
    if not subject or not grade:
        raise HTTPException(status_code=400, detail="subject and grade required")
    _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
    if not isinstance(chapters, dict):
        raise HTTPException(status_code=400, detail="chapters must be an object")
    normalized: dict[str, dict[str, float]] = {}
    for key, val in chapters.items():
        if not isinstance(val, dict):
            continue
        try:
            x = float(val.get("x"))
            y = float(val.get("y"))
        except (TypeError, ValueError):
            continue
        normalized[str(key)] = {"x": x, "y": y}
    configs = session.exec(
        select(EvalConfig)
        .where(EvalConfig.subject == subject, EvalConfig.grade == grade)
        .order_by(EvalConfig.id.desc())
    ).all()
    if configs:
        cfg = configs[0]
    else:
        cfg = EvalConfig(subject=subject, grade=grade)
        session.add(cfg)
        session.commit()
        session.refresh(cfg)
        configs = [cfg]
    layout_json = json.dumps({"version": 1, "chapters": normalized}, ensure_ascii=False)
    for item in configs:
        item.graph_layout_json = layout_json
        session.add(item)
    session.commit()
    return {"ok": True, "chapters": normalized}


@router.put("/kps/{kp_id}/practice_total")
def update_kp_practice_total(
    kp_id: int,
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    raw = payload.get("practice_total", None)
    if raw is None:
        kp.practice_total = None
    else:
        try:
            val = int(raw)
        except Exception:
            raise HTTPException(status_code=400, detail="practice_total must be an integer")
        if val < 0:
            raise HTTPException(status_code=400, detail="practice_total must be >= 0")
        kp.practice_total = val
    session.add(kp)
    session.commit()
    session.refresh(kp)
    return {"ok": True, "kp_id": kp.id, "practice_total": kp.practice_total}


@router.delete("/kps/{kp_id}")
def delete_kp(
    kp_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    # 级联清理所有可能引用此知识点的数据，避免“表面无引用但仍删除失败”的情况。
    question_ids = [
        int(qid)
        for qid in session.exec(select(Question.id).where(Question.kp_id == kp_id)).all()
        if qid is not None
    ]
    quiz_ids = [
        int(qid)
        for qid in session.exec(select(Quiz.id).where(Quiz.kp_id == kp_id)).all()
        if qid is not None
    ]
    try:
        # 与题目、测验相关
        if question_ids:
            session.exec(delete(PracticeAttempt).where(PracticeAttempt.question_id.in_(question_ids)))
            session.exec(delete(ReviewSchedule).where(ReviewSchedule.question_id.in_(question_ids)))
            session.exec(delete(KpQuestionAssignment).where(KpQuestionAssignment.question_id.in_(question_ids)))
            session.exec(delete(Question).where(Question.id.in_(question_ids)))
        if quiz_ids:
            session.exec(delete(QuizAttempt).where(QuizAttempt.quiz_id.in_(quiz_ids)))
            session.exec(delete(QuizItem).where(QuizItem.quiz_id.in_(quiz_ids)))
            session.exec(delete(Quiz).where(Quiz.id.in_(quiz_ids)))
        # 与知识点直接关联
        session.exec(delete(KnowledgeEdge).where((KnowledgeEdge.prereq_id == kp_id) | (KnowledgeEdge.next_id == kp_id)))
        session.exec(delete(LearningResource).where(LearningResource.kp_id == kp_id))
        session.exec(delete(KpTask).where(KpTask.kp_id == kp_id))
        session.exec(delete(PracticeAttempt).where(PracticeAttempt.kp_id == kp_id))
        session.exec(delete(QuizAttempt).where(QuizAttempt.kp_id == kp_id))
        session.exec(delete(ReviewSchedule).where(ReviewSchedule.kp_id == kp_id))
        session.exec(delete(ExpressionEvent).where(ExpressionEvent.kp_id == kp_id))
        session.exec(delete(Mastery).where(Mastery.kp_id == kp_id))
        session.exec(delete(Note).where(Note.kp_id == kp_id))
        session.exec(delete(VideoProgress).where(VideoProgress.kp_id == kp_id))
        session.exec(delete(KpQuestionAssignment).where(KpQuestionAssignment.kp_id == kp_id))
        session.exec(delete(StageImportRecord).where(StageImportRecord.kp_id == kp_id))
        session.exec(delete(LearningBehaviorEvent).where(LearningBehaviorEvent.kp_id == kp_id))
        session.exec(delete(RecommendationLog).where((RecommendationLog.source_kp_id == kp_id) | (RecommendationLog.target_kp_id == kp_id)))

        # 最后删除知识点
        session.delete(kp)
        session.commit()
    except Exception as exc:
        session.rollback()
        raise HTTPException(status_code=400, detail=f"删除失败：该知识点仍被其他数据引用 ({exc})") from exc

    _log_action(session, admin, "kp_delete", f"kp_id={kp_id}")
    return {"ok": True}


@router.get("/graph/kp-coverage")
def graph_kp_coverage(
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    """各知识点挂载资源数、题库题数、任务数、是否配置小测（教师端图谱缺省提示）。"""
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    kps = session.exec(
        select(KnowledgePoint)
        .where(KnowledgePoint.subject == subject, KnowledgePoint.grade == grade)
        .order_by(KnowledgePoint.id)
    ).all()
    kp_ids = [int(k.id) for k in kps if k.id is not None]
    if not kp_ids:
        return {"items": []}
    r_rows = session.exec(
        select(LearningResource.kp_id, func.count())
        .where(LearningResource.kp_id.in_(kp_ids))
        .group_by(LearningResource.kp_id)
    ).all()
    q_rows = session.exec(
        select(Question.kp_id, func.count())
        .where(Question.kp_id.in_(kp_ids))
        .group_by(Question.kp_id)
    ).all()
    t_rows = session.exec(
        select(KpTask.kp_id, func.count())
        .where(KpTask.kp_id.in_(kp_ids))
        .group_by(KpTask.kp_id)
    ).all()
    quiz_rows = session.exec(select(Quiz).where(Quiz.kp_id.in_(kp_ids))).all()
    has_quiz = {int(q.kp_id) for q in quiz_rows if q.kp_id is not None}
    rmap = {int(r[0]): int(r[1]) for r in r_rows if r[0] is not None}
    qmap = {int(r[0]): int(r[1]) for r in q_rows if r[0] is not None}
    tmap = {int(r[0]): int(r[1]) for r in t_rows if r[0] is not None}
    items = []
    for kid in kp_ids:
        items.append(
            {
                "kp_id": kid,
                "resource_count": rmap.get(kid, 0),
                "question_count": qmap.get(kid, 0),
                "task_count": tmap.get(kid, 0),
                "has_quiz": kid in has_quiz,
            }
        )
    return {"items": items}


def _export_kp_row(kp: KnowledgePoint) -> dict:
    return {
        "id": int(kp.id) if kp.id is not None else None,
        "code": kp.code,
        "title": kp.title,
        "description": kp.description or "",
        "chapter": kp.chapter or "",
        "knowledge_tag": kp.knowledge_tag or "",
        "ability_tag": kp.ability_tag or "",
        "literacy_tag": kp.literacy_tag or "",
        "importance": float(kp.importance or 0),
        "difficulty": float(kp.difficulty or 0),
        "pos_x": kp.pos_x,
        "pos_y": kp.pos_y,
    }


@router.get("/graph/export")
def graph_export(
    subject: str,
    grade: str,
    format: str = "json",
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    """导出当前课程（subject+grade）下的知识点、知识边、章节边。"""
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    format = (format or "json").strip().lower()
    if format not in {"json", "csv"}:
        raise HTTPException(status_code=400, detail="format 须为 json 或 csv")

    kps = session.exec(
        select(KnowledgePoint)
        .where(KnowledgePoint.subject == subject, KnowledgePoint.grade == grade)
        .order_by(KnowledgePoint.chapter, KnowledgePoint.id)
    ).all()
    edges = session.exec(
        select(KnowledgeEdge)
        .where(KnowledgeEdge.subject == subject, KnowledgeEdge.grade == grade)
        .order_by(KnowledgeEdge.id)
    ).all()
    chapter_edges = session.exec(
        select(ChapterEdge)
        .where(ChapterEdge.subject == subject, ChapterEdge.grade == grade)
        .order_by(ChapterEdge.id)
    ).all()

    kp_payload = [_export_kp_row(kp) for kp in kps]
    edge_payload = [
        {
            "prereq_id": int(e.prereq_id),
            "next_id": int(e.next_id),
            "relation_type": _relation_type_value(e.relation_type),
        }
        for e in edges
        if e.prereq_id is not None and e.next_id is not None
    ]
    ch_payload = [
        {
            "source_chapter": ce.source_chapter,
            "target_chapter": ce.target_chapter,
            "relation_type": _relation_type_value(ce.relation_type),
        }
        for ce in chapter_edges
    ]

    # HTTP Content-Disposition 须为 latin-1；课程名可能含中文，此处仅保留 ASCII 文件名
    raw = re.sub(r"[^A-Za-z0-9_.-]+", "_", (subject or "").strip())[:40]
    safe_name = re.sub(r"_+", "_", raw).strip("_") or "graph"

    if format == "json":
        payload = {
            "subject": subject,
            "grade": grade,
            "exported_at": datetime.utcnow().isoformat() + "Z",
            "knowledge_points": kp_payload,
            "edges": edge_payload,
            "chapter_edges": ch_payload,
        }
        raw = json.dumps(payload, ensure_ascii=False, indent=2)
        return Response(
            content=raw,
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}-graph.json"'},
        )

    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(["# knowledge_points"])
    writer.writerow(
        [
            "id",
            "code",
            "title",
            "chapter",
            "knowledge_tag",
            "ability_tag",
            "literacy_tag",
            "importance",
            "difficulty",
            "pos_x",
            "pos_y",
        ],
    )
    for row in kp_payload:
        writer.writerow(
            [
                row["id"],
                row["code"],
                row["title"],
                row["chapter"],
                row["knowledge_tag"],
                row["ability_tag"],
                row["literacy_tag"],
                row["importance"],
                row["difficulty"],
                row["pos_x"] if row["pos_x"] is not None else "",
                row["pos_y"] if row["pos_y"] is not None else "",
            ]
        )
    writer.writerow([])
    writer.writerow(["# edges"])
    writer.writerow(["prereq_id", "next_id", "relation_type"])
    for e in edge_payload:
        writer.writerow([e["prereq_id"], e["next_id"], e["relation_type"]])
    writer.writerow([])
    writer.writerow(["# chapter_edges"])
    writer.writerow(["source_chapter", "target_chapter", "relation_type"])
    for c in ch_payload:
        writer.writerow([c["source_chapter"], c["target_chapter"], c["relation_type"]])

    csv_bytes = "\ufeff" + buf.getvalue()
    return Response(
        content=csv_bytes.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}-graph.csv"'},
    )


@router.get("/edges")
def list_edges_admin(
    subject: str | None = None,
    grade: str | None = None,
    page: int = 1,
    page_size: int = 15,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    page = max(1, page)
    # 教师图谱工作台一次拉取较多边（前端 page_size=500），上限过低会导致画布与保存结果不一致
    page_size = max(1, min(500, page_size))
    q = select(KnowledgeEdge).order_by(KnowledgeEdge.id.desc())
    q_total = select(func.count()).select_from(KnowledgeEdge)
    allowed_subjects = _teacher_accessible_subjects(session, admin)
    if allowed_subjects is not None:
        if not allowed_subjects:
            return {"items": [], "total": 0, "page": page, "page_size": page_size}
        allowed_subject_list = sorted(allowed_subjects)
        q = q.where(KnowledgeEdge.subject.in_(allowed_subject_list))
        q_total = q_total.where(KnowledgeEdge.subject.in_(allowed_subject_list))
    if subject:
        _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
        q = q.where(KnowledgeEdge.subject == subject)
        q_total = q_total.where(KnowledgeEdge.subject == subject)
    if grade:
        q = q.where(KnowledgeEdge.grade == grade)
        q_total = q_total.where(KnowledgeEdge.grade == grade)
    total = session.exec(q_total).one()
    rows = session.exec(q.offset((page - 1) * page_size).limit(page_size)).all()
    items = [
        KnowledgeEdgeOut(
            id=r.id,
            prereq_id=r.prereq_id,
            next_id=r.next_id,
            relation_type=_relation_type_value(r.relation_type),
        )
        for r in rows
    ]
    return {"items": [i.model_dump() for i in items], "total": int(total or 0), "page": page, "page_size": page_size}


@router.post("/edges", response_model=KnowledgeEdgeOut)
def create_edge(
    payload: KnowledgeEdgeIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_subject_access(session=session, admin=admin, subject=payload.subject, grade=payload.grade)
    prereq = _require_teacher_kp_id_access(session=session, admin=admin, kp_id=payload.prereq_id)
    next_kp = _require_teacher_kp_id_access(session=session, admin=admin, kp_id=payload.next_id)
    if prereq.subject != payload.subject or next_kp.subject != payload.subject:
        raise HTTPException(status_code=400, detail="Edge knowledge points must belong to the same subject")
    if payload.prereq_id == payload.next_id:
        raise HTTPException(status_code=400, detail="Invalid edge")
    try:
        relation_type = RelationType(payload.relation_type or "prerequisite")
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="relation_type 须为 prerequisite / related / support / contains 之一",
        )
    exists = session.exec(
        select(KnowledgeEdge).where(KnowledgeEdge.prereq_id == payload.prereq_id, KnowledgeEdge.next_id == payload.next_id)
    ).first()
    if exists:
        exists.relation_type = relation_type
        session.add(exists)
        session.commit()
        session.refresh(exists)
        return KnowledgeEdgeOut(
            id=exists.id,
            prereq_id=exists.prereq_id,
            next_id=exists.next_id,
            relation_type=_relation_type_value(exists.relation_type),
        )
    edge = KnowledgeEdge(
        subject=payload.subject,
        grade=payload.grade,
        prereq_id=payload.prereq_id,
        next_id=payload.next_id,
        relation_type=relation_type,
    )
    session.add(edge)
    session.commit()
    session.refresh(edge)
    return KnowledgeEdgeOut(
        id=edge.id,
        prereq_id=edge.prereq_id,
        next_id=edge.next_id,
        relation_type=_relation_type_value(edge.relation_type),
    )


@router.delete("/edges/{edge_id}")
def delete_edge(
    edge_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    edge = session.get(KnowledgeEdge, edge_id)
    if edge is None:
        raise HTTPException(status_code=404, detail="Edge not found")
    _require_teacher_subject_access(session=session, admin=admin, subject=edge.subject, grade=edge.grade)
    session.delete(edge)
    session.commit()
    return {"ok": True}


@router.get("/chapter-edges")
def list_chapter_edges(
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
    rows = session.exec(
        select(ChapterEdge)
        .where(ChapterEdge.subject == subject, ChapterEdge.grade == grade)
        .order_by(ChapterEdge.id)
    ).all()
    return [
        {
            "id": int(row.id),
            "source_chapter": row.source_chapter,
            "target_chapter": row.target_chapter,
            "relation_type": _relation_type_value(row.relation_type),
        }
        for row in rows
        if row.id is not None
    ]


@router.post("/chapter-edges")
def create_chapter_edge(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    subject = str(payload.get("subject", "")).strip()
    grade = str(payload.get("grade", "")).strip()
    source_chapter = str(payload.get("source_chapter", "")).strip()
    target_chapter = str(payload.get("target_chapter", "")).strip()
    relation_type_raw = str(payload.get("relation_type", RelationType.related.value)).strip() or RelationType.related.value
    if not subject or not grade or not source_chapter or not target_chapter:
        raise HTTPException(status_code=400, detail="subject/grade/source_chapter/target_chapter required")
    _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
    if source_chapter == target_chapter:
        raise HTTPException(status_code=400, detail="source_chapter and target_chapter cannot be same")
    try:
        relation_type = RelationType(relation_type_raw)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="relation_type must be prerequisite or related") from exc
    exists = session.exec(
        select(ChapterEdge).where(
            ChapterEdge.subject == subject,
            ChapterEdge.grade == grade,
            ChapterEdge.source_chapter == source_chapter,
            ChapterEdge.target_chapter == target_chapter,
        )
    ).first()
    if exists is not None:
        return {
            "id": int(exists.id),
            "source_chapter": exists.source_chapter,
            "target_chapter": exists.target_chapter,
            "relation_type": _relation_type_value(exists.relation_type),
        }
    row = ChapterEdge(
        subject=subject,
        grade=grade,
        source_chapter=source_chapter,
        target_chapter=target_chapter,
        relation_type=relation_type,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return {
        "id": int(row.id),
        "source_chapter": row.source_chapter,
        "target_chapter": row.target_chapter,
        "relation_type": _relation_type_value(row.relation_type),
    }


@router.delete("/chapter-edges/{edge_id}")
def delete_chapter_edge(
    edge_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    row = session.get(ChapterEdge, edge_id)
    if row is None:
        raise HTTPException(status_code=404, detail="Chapter edge not found")
    _require_teacher_subject_access(session=session, admin=admin, subject=row.subject, grade=row.grade)
    session.delete(row)
    session.commit()
    return {"ok": True}


@router.get("/questions")
def list_questions(
    kp_id: int | None = None,
    keyword: str | None = None,
    q_type: str | None = None,
    min_difficulty: float | None = None,
    max_difficulty: float | None = None,
    page: int = 1,
    page_size: int = 15,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    page = max(1, page)
    page_size = max(1, min(100, page_size))
    q = select(Question).order_by(Question.id.desc())
    q_total = select(func.count()).select_from(Question)
    allowed_subjects = _teacher_accessible_subjects(session, admin)
    if allowed_subjects is not None:
        if not allowed_subjects:
            return {"items": [], "total": 0, "page": page, "page_size": page_size}
        allowed_subject_list = sorted(allowed_subjects)
        q = q.where(Question.subject.in_(allowed_subject_list))
        q_total = q_total.where(Question.subject.in_(allowed_subject_list))
    if kp_id is not None:
        _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
        q = q.where(Question.kp_id == kp_id)
        q_total = q_total.where(Question.kp_id == kp_id)
    if keyword:
        kw = keyword.strip()
        if kw:
            q = q.where(Question.prompt.contains(kw))
            q_total = q_total.where(Question.prompt.contains(kw))
    if q_type:
        q = q.where(Question.type == q_type)
        q_total = q_total.where(Question.type == q_type)
    if min_difficulty is not None:
        q = q.where(Question.difficulty >= float(min_difficulty))
        q_total = q_total.where(Question.difficulty >= float(min_difficulty))
    if max_difficulty is not None:
        q = q.where(Question.difficulty <= float(max_difficulty))
        q_total = q_total.where(Question.difficulty <= float(max_difficulty))
    total = session.exec(q_total).one()
    rows = session.exec(q.offset((page - 1) * page_size).limit(page_size)).all()

    ids = [r.id for r in rows]
    stats_map: dict[int, dict] = {}
    if ids:
        stats = session.exec(
            select(
                PracticeAttempt.question_id,
                func.count().label("attempts"),
                func.avg(func.cast(PracticeAttempt.correct, Integer)).label("correct_rate"),
            )
            .where(PracticeAttempt.question_id.in_(ids))
            .group_by(PracticeAttempt.question_id)
        ).all()
        for qid, attempts, correct_rate in stats:
            stats_map[int(qid)] = {
                "attempts": int(attempts or 0),
                "correct_rate": float(correct_rate) if correct_rate is not None else None,
            }

    items = [
        QuestionOut(
            id=r.id,
            kp_id=r.kp_id,
            type=r.type,
            prompt=r.prompt,
            options=json.loads(r.options_json),
            answer=r.answer,
            explanation=r.explanation,
            difficulty=r.difficulty,
            source=r.source,
            tags=r.tags,
            version=r.version,
            cognitive_level=getattr(r, "cognitive_level", None) or "understand",
            ability_subtags=getattr(r, "ability_subtags", None) or "",
            attempts=stats_map.get(r.id, {}).get("attempts"),
            correct_rate=stats_map.get(r.id, {}).get("correct_rate"),
        )
        for r in rows
    ]
    return {"items": [i.model_dump() for i in items], "total": int(total or 0), "page": page, "page_size": page_size}


@router.get("/questions/export")
def export_questions(
    kp_id: int | None = None,
    keyword: str | None = None,
    q_type: str | None = None,
    min_difficulty: float | None = None,
    max_difficulty: float | None = None,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    q = select(Question).order_by(Question.id.desc())
    allowed_subjects = _teacher_accessible_subjects(session, admin)
    if allowed_subjects is not None:
        if not allowed_subjects:
            return Response(content="id,kp_id,type,prompt,options,answer,explanation,difficulty,source,tags,version,cognitive_level,ability_subtags", media_type="text/csv")
        q = q.where(Question.subject.in_(sorted(allowed_subjects)))
    if kp_id is not None:
        _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
        q = q.where(Question.kp_id == kp_id)
    if keyword:
        kw = keyword.strip()
        if kw:
            q = q.where(Question.prompt.contains(kw))
    if q_type:
        q = q.where(Question.type == q_type)
    if min_difficulty is not None:
        q = q.where(Question.difficulty >= float(min_difficulty))
    if max_difficulty is not None:
        q = q.where(Question.difficulty <= float(max_difficulty))

    rows = session.exec(q).all()
    lines = [
        "id,kp_id,type,prompt,options,answer,explanation,difficulty,source,tags,version,cognitive_level,ability_subtags"
    ]
    for r in rows:
        prompt = r.prompt.replace('"', "'").replace(",", "，")
        options = r.options_json.replace('"', "'").replace(",", "，")
        answer = r.answer.replace('"', "'").replace(",", "，")
        explanation = r.explanation.replace('"', "'").replace(",", "，")
        source = (r.source or "").replace('"', "'").replace(",", "，")
        tags = (r.tags or "").replace('"', "'").replace(",", "，")
        version = (r.version or "").replace('"', "'").replace(",", "，")
        cognitive = (getattr(r, "cognitive_level", None) or "understand").replace('"', "'").replace(",", "，")
        ability_sub = (getattr(r, "ability_subtags", None) or "").replace('"', "'").replace(",", "，")
        lines.append(
            f"{r.id},{r.kp_id},{r.type},\"{prompt}\",\"{options}\",\"{answer}\",\"{explanation}\",{r.difficulty},\"{source}\",\"{tags}\",\"{version}\",\"{cognitive}\",\"{ability_sub}\""
        )
    csv = "\n".join(lines)
    return Response(content=csv, media_type="text/csv")


@router.post("/questions/recalibrate-difficulty")
def recalibrate_question_difficulty(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp_id = payload.get("kp_id")
    kp_id = int(kp_id) if kp_id is not None else None
    min_attempts = int(payload.get("min_attempts", 5))
    blend = float(payload.get("blend", 0.7))
    step = float(payload.get("step", 0.1))
    min_attempts = max(1, min(1000, min_attempts))
    blend = max(0.0, min(1.0, blend))
    step = max(0.01, min(0.5, step))

    q = select(Question)
    allowed_subjects = _teacher_accessible_subjects(session, admin)
    if allowed_subjects is not None:
        if not allowed_subjects:
            return {"ok": True, "updated": 0}
        q = q.where(Question.subject.in_(sorted(allowed_subjects)))
    if kp_id is not None:
        _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
        q = q.where(Question.kp_id == kp_id)
    questions = session.exec(q).all()
    if not questions:
        return {"ok": True, "updated": 0}

    ids = [qq.id for qq in questions]
    stats = session.exec(
        select(
            PracticeAttempt.question_id,
            func.count().label("attempts"),
            func.avg(func.cast(PracticeAttempt.correct, Integer)).label("correct_rate"),
        )
        .where(PracticeAttempt.question_id.in_(ids))
        .group_by(PracticeAttempt.question_id)
    ).all()
    stats_map = {int(qid): (int(attempts or 0), float(cr) if cr is not None else None) for qid, attempts, cr in stats}

    def clamp01(x: float) -> float:
        return max(0.0, min(1.0, x))

    def quantize(x: float) -> float:
        return round(x / step) * step

    updated = 0
    for qq in questions:
        attempts, correct_rate = stats_map.get(int(qq.id), (0, None))
        if attempts < min_attempts or correct_rate is None:
            continue
        estimated = clamp01(1.0 - float(correct_rate))
        new_value = blend * estimated + (1.0 - blend) * float(qq.difficulty)
        new_value = clamp01(quantize(new_value))
        if abs(float(qq.difficulty) - new_value) >= 1e-6:
            qq.difficulty = float(new_value)
            session.add(qq)
            updated += 1
    session.commit()
    return {"ok": True, "updated": updated, "min_attempts": min_attempts, "blend": blend, "step": step}


@router.get("/kp-questions")
def list_assigned_questions(
    kp_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    assigns = session.exec(
        select(KpQuestionAssignment).where(KpQuestionAssignment.kp_id == kp_id).order_by(KpQuestionAssignment.order)
    ).all()
    if not assigns:
        return []
    ids = [a.question_id for a in assigns]
    qs = session.exec(select(Question).where(Question.id.in_(ids))).all()
    qmap = {q.id: q for q in qs}
    return [
        {
            "id": a.id,
            "kp_id": a.kp_id,
            "question_id": a.question_id,
            "order": a.order,
            "type": qmap[a.question_id].type if a.question_id in qmap else "",
            "prompt": qmap[a.question_id].prompt if a.question_id in qmap else "",
        }
        for a in assigns
        if a.question_id in qmap
    ]


@router.post("/kp-questions")
def assign_questions(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp_id = int(payload.get("kp_id"))
    question_ids = payload.get("question_ids") or []
    if not isinstance(question_ids, list) or not question_ids:
        raise HTTPException(status_code=400, detail="question_ids required")
    kp = session.get(KnowledgePoint, kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    existing = session.exec(select(KpQuestionAssignment).where(KpQuestionAssignment.kp_id == kp_id)).all()
    existing_ids = {e.question_id for e in existing}
    max_order = max([e.order for e in existing], default=0)
    created = 0
    for qid in question_ids:
        qid_i = int(qid)
        q = _require_teacher_question_access(session=session, admin=admin, question=session.get(Question, qid_i))
        if q.kp_id != kp_id:
            raise HTTPException(status_code=400, detail="Question not found for this knowledge point")
        if qid_i in existing_ids:
            continue
        max_order += 1
        session.add(KpQuestionAssignment(kp_id=kp_id, question_id=qid_i, order=max_order))
        created += 1
    session.commit()
    return {"ok": True, "created": created}


@router.put("/kp-questions/reorder")
def reorder_questions(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp_id = int(payload.get("kp_id"))
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    ordered_ids = payload.get("ordered_question_ids") or []
    if not isinstance(ordered_ids, list) or not ordered_ids:
        raise HTTPException(status_code=400, detail="ordered_question_ids required")
    assigns = session.exec(select(KpQuestionAssignment).where(KpQuestionAssignment.kp_id == kp_id)).all()
    amap = {a.question_id: a for a in assigns}
    order = 1
    for qid in ordered_ids:
        qid_i = int(qid)
        if qid_i not in amap:
            continue
        a = amap[qid_i]
        a.order = order
        order += 1
        session.add(a)
    session.commit()
    return {"ok": True}


@router.delete("/kp-questions/{assignment_id}")
def remove_assignment(
    assignment_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    a = session.get(KpQuestionAssignment, assignment_id)
    if a is None:
        raise HTTPException(status_code=404, detail="Assignment not found")
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=int(a.kp_id))
    session.delete(a)
    session.commit()
    return {"ok": True}


@router.post("/questions", response_model=QuestionOut)
def create_question(
    payload: QuestionIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = session.get(KnowledgePoint, payload.kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    qtype = payload.type.strip()
    if qtype not in {"mcq", "blank"}:
        raise HTTPException(status_code=400, detail="Invalid question type")
    try:
        cognitive_level = normalize_question_cognitive_level(payload.cognitive_level, strict=True)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    ability_subtags = canonical_ability_subtags_str(payload.ability_subtags)
    options = payload.options if qtype == "mcq" else []
    question = Question(
        subject=kp.subject,
        grade=kp.grade,
        kp_id=payload.kp_id,
        type=qtype,
        prompt=payload.prompt,
        options_json=json.dumps(options, ensure_ascii=False),
        answer=payload.answer,
        explanation=payload.explanation,
        difficulty=float(payload.difficulty),
        source=payload.source.strip(),
        tags=payload.tags.strip(),
        version=payload.version.strip() or "v1",
        cognitive_level=cognitive_level,
        ability_subtags=ability_subtags,
    )
    session.add(question)
    session.commit()
    session.refresh(question)
    return QuestionOut(
        id=question.id,
        kp_id=question.kp_id,
        type=question.type,
        prompt=question.prompt,
        options=options,
        answer=question.answer,
        explanation=question.explanation,
        difficulty=question.difficulty,
        source=question.source,
        tags=question.tags,
        version=question.version,
        cognitive_level=question.cognitive_level or cognitive_level,
        ability_subtags=question.ability_subtags or "",
    )


@router.post("/questions/import-docx")
def import_questions_docx(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    try:
        from docx import Document
    except Exception as exc:
        raise HTTPException(status_code=500, detail="python-docx 未安装，暂时无法导入 docx 题库") from exc
    filename = (file.filename or "").lower()
    if not filename.endswith(".docx"):
        raise HTTPException(status_code=400, detail="Only .docx files are supported")

    data = file.file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")

    doc = Document(BytesIO(data))
    lines: list[str] = []
    for p in doc.paragraphs:
        if not p.text:
            continue
        for part in p.text.splitlines():
            line = part.strip()
            if line:
                lines.append(line)

    markers = {"题目", "【题目】", "[题目]", "题目开始", "---", "----", "-----"}
    blocks: list[list[str]] = []
    current: list[str] = []
    for line in lines:
        normalized_line = re.sub(r"^[\\s\\-\\*\\•\\d\\._、)]+", "", line).strip()
        if normalized_line in markers:
            if current:
                blocks.append(current)
                current = []
            continue
        current.append(line)
    if current:
        blocks.append(current)
    if not blocks and lines:
        blocks = [lines]

    def split_key_value(text: str) -> tuple[str | None, str]:
        m = re.match(r"^([^:：=]+)[:：=]\s*(.*)$", text)
        if not m:
            return None, text.strip()
        key = re.sub(r"^[\\s\\-\\*\\•\\d\\._、)]+", "", m.group(1)).strip()
        return key, m.group(2).strip()

    def normalize_type(value: str) -> str | None:
        v = value.strip().lower()
        if v in {"选择", "选择题", "单选", "mcq"}:
            return "mcq"
        if v in {"填空", "填空题", "blank"}:
            return "blank"
        return None

    key_map = {
        "知识点编码": "kp_code",
        "知识点": "kp_code",
        "kp": "kp_code",
        "KP": "kp_code",
        "编码": "kp_code",
        "题型": "qtype",
        "类型": "qtype",
        "type": "qtype",
        "TYPE": "qtype",
        "题干": "prompt",
        "题目": "prompt",
        "PROMPT": "prompt",
        "答案": "answer",
        "ANSWER": "answer",
        "解析": "explanation",
        "EXPLANATION": "explanation",
        "难度": "difficulty",
        "DIFFICULTY": "difficulty",
        "选项": "options",
        "认知层级": "cognitive_level",
        "认知层次": "cognitive_level",
        "布鲁姆": "cognitive_level",
        "布鲁姆层级": "cognitive_level",
        "cognitive": "cognitive_level",
        "cognitive_level": "cognitive_level",
        "COGNITIVE_LEVEL": "cognitive_level",
        "能力标签": "ability_subtags",
        "能力二级标签": "ability_subtags",
        "ability": "ability_subtags",
        "ability_subtags": "ability_subtags",
        "ABILITY_SUBTAGS": "ability_subtags",
    }

    created = 0
    skipped = 0
    errors: list[str] = []
    seen: set[tuple[int, str]] = set()

    for idx, block in enumerate(blocks, start=1):
        data_map: dict[str, str] = {}
        options: list[str] = []
        in_options = False

        for line in block:
            key, value = split_key_value(line)
            if key is None:
                if in_options:
                    m = re.match(r"^([A-H])[\.\:：=]\s*(.+)$", line)
                    if m:
                        options.append(m.group(2).strip())
                        continue
                continue

            mapped = key_map.get(key.strip())
            if mapped == "options":
                in_options = True
                if value:
                    m = re.match(r"^([A-H])[\.\:：=]\s*(.+)$", value)
                    if m:
                        options.append(m.group(2).strip())
                continue

            in_options = False
            if mapped:
                data_map[mapped] = value

        kp_code = (data_map.get("kp_code") or "").strip()
        qtype_raw = data_map.get("qtype") or ""
        qtype = normalize_type(qtype_raw) if qtype_raw else None
        prompt = (data_map.get("prompt") or "").strip()
        answer = (data_map.get("answer") or "").strip()
        explanation = (data_map.get("explanation") or "").strip()
        difficulty_str = (data_map.get("difficulty") or "").strip()

        if not kp_code or not qtype or not prompt or not answer:
            errors.append(f"第{idx}题缺少必填字段（知识点编码/题型/题干/答案）")
            continue

        kp = session.exec(select(KnowledgePoint).where(KnowledgePoint.code == kp_code)).first()
        if kp is None:
            errors.append(f"第{idx}题知识点编码不存在: {kp_code}")
            continue
        try:
            _require_teacher_kp_access(session=session, admin=admin, kp=kp)
        except HTTPException:
            errors.append(f"第{idx}题无权导入到知识点: {kp_code}")
            continue

        if qtype == "mcq":
            if len(options) < 2:
                errors.append(f"第{idx}题选项不足（至少2个）")
                continue
            if answer:
                m = re.match(r"^[A-H]", answer, re.I)
                if m:
                    answer = m.group(0).upper()
            if answer not in {"A", "B", "C", "D", "E", "F", "G", "H"}:
                errors.append(f"第{idx}题答案必须是 A/B/C/D...")
                continue
        else:
            options = []

        try:
            difficulty = float(difficulty_str) if difficulty_str else 0.4
        except ValueError:
            difficulty = 0.4
        difficulty = min(1.0, max(0.0, difficulty))

        cog_raw = (data_map.get("cognitive_level") or "").strip()
        _cog_cn = {
            "记忆": "remember",
            "理解": "understand",
            "应用": "apply",
            "分析": "analyze",
            "评价": "evaluate",
            "创造": "create",
        }
        if cog_raw in _cog_cn:
            cog_raw = _cog_cn[cog_raw]
        cognitive_level = normalize_question_cognitive_level(cog_raw or None, strict=False)
        ability_subtags = canonical_ability_subtags_str(data_map.get("ability_subtags"))

        key = (kp.id, prompt)
        if key in seen:
            skipped += 1
            continue
        seen.add(key)

        exists = session.exec(select(Question).where(Question.kp_id == kp.id, Question.prompt == prompt)).first()
        if exists:
            skipped += 1
            continue

        q = Question(
            subject=kp.subject,
            grade=kp.grade,
            kp_id=kp.id,
            type=qtype,
            prompt=prompt,
            options_json=json.dumps(options, ensure_ascii=False),
            answer=answer,
            explanation=explanation,
            difficulty=difficulty,
            cognitive_level=cognitive_level,
            ability_subtags=ability_subtags,
        )
        session.add(q)
        created += 1

    session.commit()
    _log_action(session, admin, "questions_import_docx", f"created={created} skipped={skipped} errors={len(errors)}")
    return {"ok": True, "created": created, "skipped": skipped, "errors": errors[:50]}


@router.put("/questions/{question_id}", response_model=QuestionOut)
def update_question(
    question_id: int,
    payload: QuestionIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    q = session.get(Question, question_id)
    _require_teacher_question_access(session=session, admin=admin, question=q)
    qtype = payload.type.strip()
    if qtype not in {"mcq", "blank"}:
        raise HTTPException(status_code=400, detail="Invalid question type")
    try:
        cognitive_level = normalize_question_cognitive_level(payload.cognitive_level, strict=True)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    ability_subtags = canonical_ability_subtags_str(payload.ability_subtags)
    q.kp_id = payload.kp_id
    kp = session.get(KnowledgePoint, payload.kp_id)
    _require_teacher_kp_access(session=session, admin=admin, kp=kp)
    q.subject = kp.subject
    q.grade = kp.grade
    q.type = qtype
    q.prompt = payload.prompt
    options = payload.options if qtype == "mcq" else []
    q.options_json = json.dumps(options, ensure_ascii=False)
    q.answer = payload.answer
    q.explanation = payload.explanation
    q.difficulty = float(payload.difficulty)
    q.source = payload.source.strip()
    q.tags = payload.tags.strip()
    q.version = payload.version.strip() or "v1"
    q.cognitive_level = cognitive_level
    q.ability_subtags = ability_subtags
    session.add(q)
    session.commit()
    session.refresh(q)
    return QuestionOut(
        id=q.id,
        kp_id=q.kp_id,
        type=q.type,
        prompt=q.prompt,
        options=options,
        answer=q.answer,
        explanation=q.explanation,
        difficulty=q.difficulty,
        source=q.source,
        tags=q.tags,
        version=q.version,
        cognitive_level=q.cognitive_level or cognitive_level,
        ability_subtags=q.ability_subtags or "",
    )


@router.delete("/questions/{question_id}")
def delete_question(
    question_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    q = session.get(Question, question_id)
    _require_teacher_question_access(session=session, admin=admin, question=q)
    session.delete(q)
    session.commit()
    return {"ok": True}


@router.get("/quiz")
def get_quiz_admin(
    kp_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    quiz = session.exec(select(Quiz).where(Quiz.kp_id == kp_id)).first()
    if quiz is None:
        return {"quiz_id": None, "pass_accuracy": 0.8, "items": []}
    items = session.exec(select(QuizItem).where(QuizItem.quiz_id == quiz.id)).all()
    return {
        "quiz_id": quiz.id,
        "pass_accuracy": quiz.pass_accuracy,
        "items": [
            {
                "id": i.id,
                "type": i.type,
                "prompt": i.prompt,
                "options": json.loads(i.options_json),
                "answer": i.answer,
                "explanation": i.explanation,
                "key_item": i.key_item,
            }
            for i in items
        ],
    }


@router.put("/quiz/{kp_id}/pass_accuracy")
def update_quiz_pass_accuracy(
    kp_id: int,
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp = _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    quiz = session.exec(select(Quiz).where(Quiz.kp_id == kp_id)).first()
    if quiz is None:
        quiz = Quiz(subject=kp.subject, grade=kp.grade, kp_id=kp_id, pass_accuracy=0.8)
    pass_accuracy = float(payload.get("pass_accuracy", quiz.pass_accuracy))
    quiz.pass_accuracy = max(0.0, min(1.0, pass_accuracy))
    session.add(quiz)
    session.commit()
    return {"ok": True, "pass_accuracy": quiz.pass_accuracy}


@router.post("/quiz/item")
def create_quiz_item(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp_id = int(payload.get("kp_id"))
    kp = _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    qtype = str(payload.get("type", "")).strip()
    prompt = str(payload.get("prompt", "")).strip()
    answer = str(payload.get("answer", "")).strip()
    explanation = str(payload.get("explanation", "")).strip()
    key_item = bool(payload.get("key_item", False))
    options = payload.get("options") or []
    if qtype not in {"mcq", "blank"}:
        raise HTTPException(status_code=400, detail="Invalid quiz item type")
    if not prompt or not answer:
        raise HTTPException(status_code=400, detail="prompt/answer required")
    if qtype == "mcq" and (not isinstance(options, list) or len(options) < 2):
        raise HTTPException(status_code=400, detail="mcq options required")

    quiz = session.exec(select(Quiz).where(Quiz.kp_id == kp_id)).first()
    if quiz is None:
        quiz = Quiz(subject=kp.subject, grade=kp.grade, kp_id=kp_id, pass_accuracy=0.8)
        session.add(quiz)
        session.commit()
        session.refresh(quiz)

    item = QuizItem(
        quiz_id=quiz.id,
        type=qtype,
        prompt=prompt,
        options_json=json.dumps(options if qtype == "mcq" else [], ensure_ascii=False),
        answer=answer,
        explanation=explanation,
        key_item=key_item,
    )
    session.add(item)
    session.commit()
    session.refresh(item)
    return {"ok": True, "item_id": item.id}


@router.post("/quiz/item/from-question")
def create_quiz_item_from_question(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    kp_id = int(payload.get("kp_id"))
    kp = _require_teacher_kp_id_access(session=session, admin=admin, kp_id=kp_id)
    question_id = int(payload.get("question_id"))
    q = session.get(Question, question_id)
    if q is None or q.kp_id != kp_id:
        raise HTTPException(status_code=400, detail="Question not found for this knowledge point")
    _require_teacher_question_access(session=session, admin=admin, question=q)

    quiz = session.exec(select(Quiz).where(Quiz.kp_id == kp_id)).first()
    if quiz is None:
        quiz = Quiz(subject=kp.subject, grade=kp.grade, kp_id=kp_id, pass_accuracy=0.8)
        session.add(quiz)
        session.commit()
        session.refresh(quiz)

    exists = session.exec(
        select(QuizItem).where(QuizItem.quiz_id == quiz.id, QuizItem.prompt == q.prompt)
    ).first()
    if exists:
        return {"ok": True, "item_id": exists.id, "skipped": True}

    item = QuizItem(
        quiz_id=quiz.id,
        type=q.type,
        prompt=q.prompt,
        options_json=q.options_json,
        answer=q.answer,
        explanation=q.explanation,
        key_item=False,
    )
    session.add(item)
    session.commit()
    session.refresh(item)
    return {"ok": True, "item_id": item.id}


@router.put("/quiz/item/{item_id}")
def update_quiz_item(
    item_id: int,
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    item = session.get(QuizItem, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Quiz item not found")
    quiz = session.get(Quiz, int(item.quiz_id))
    if quiz is None:
        raise HTTPException(status_code=404, detail="Quiz not found")
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=int(quiz.kp_id))
    qtype = str(payload.get("type", item.type)).strip()
    prompt = str(payload.get("prompt", item.prompt)).strip()
    answer = str(payload.get("answer", item.answer)).strip()
    explanation = str(payload.get("explanation", item.explanation)).strip()
    key_item = bool(payload.get("key_item", item.key_item))
    options = payload.get("options") or []
    if qtype not in {"mcq", "blank"}:
        raise HTTPException(status_code=400, detail="Invalid quiz item type")
    if not prompt or not answer:
        raise HTTPException(status_code=400, detail="prompt/answer required")
    if qtype == "mcq" and (not isinstance(options, list) or len(options) < 2):
        raise HTTPException(status_code=400, detail="mcq options required")

    item.type = qtype
    item.prompt = prompt
    item.answer = answer
    item.explanation = explanation
    item.key_item = key_item
    item.options_json = json.dumps(options if qtype == "mcq" else [], ensure_ascii=False)
    session.add(item)
    session.commit()
    return {"ok": True}


@router.delete("/quiz/item/{item_id}")
def delete_quiz_item(
    item_id: int,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    item = session.get(QuizItem, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Quiz item not found")
    quiz = session.get(Quiz, int(item.quiz_id))
    if quiz is None:
        raise HTTPException(status_code=404, detail="Quiz not found")
    _require_teacher_kp_id_access(session=session, admin=admin, kp_id=int(quiz.kp_id))
    session.delete(item)
    session.commit()
    return {"ok": True}


@router.post("/seed")
def seed_derivative_demo(
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin)),
):
    subject = "\u6570\u636e\u7ed3\u6784"
    grade = "\u901a\u7528"
    kps = [
        ("DS-GEN-001", "\u7ebf\u6027\u8868\u57fa\u7840", "\u7ebf\u6027\u8868\u4e0e\u987a\u5e8f\u5b58\u50a8\u7684\u57fa\u7840\u6982\u5ff5"),
        ("DS-GEN-002", "\u6808\u4e0e\u961f\u5217", "\u987a\u5e8f\u6808\u4e0e\u961f\u5217\u7684\u57fa\u7840\u64cd\u4f5c"),
        ("DS-GEN-003", "\u4e32", "\u5b57\u7b26\u4e32\u7684\u5b9a\u4e49\u4e0e\u57fa\u672c\u64cd\u4f5c"),
    ]
    code_to_id = {}
    for code, title, desc in kps:
        kp = session.exec(select(KnowledgePoint).where(KnowledgePoint.code == code)).first()
        if kp is None:
            kp = KnowledgePoint(subject=subject, grade=grade, code=code, title=title, description=desc)
            session.add(kp)
            session.commit()
            session.refresh(kp)
        code_to_id[code] = kp.id

    def add_resource(code: str, title: str, url: str, rtype: str):
        kp_id = code_to_id[code]
        exists = session.exec(select(LearningResource).where(LearningResource.kp_id == kp_id, LearningResource.title == title)).first()
        if exists is None:
            session.add(
                LearningResource(
                    subject=subject,
                    grade=grade,
                    kp_id=kp_id,
                    title=title,
                    url=url,
                    type=ResourceType(rtype),
                )
            )
            session.commit()

    def add_question(code: str, qtype: str, prompt: str, options: list[str], answer: str, explanation: str, difficulty: float):
        kp_id = code_to_id[code]
        exists = session.exec(select(Question).where(Question.kp_id == kp_id, Question.prompt == prompt)).first()
        if exists is None:
            session.add(
                Question(
                    subject=subject,
                    grade=grade,
                    kp_id=kp_id,
                    type=qtype,
                    prompt=prompt,
                    options_json=json.dumps(options, ensure_ascii=False),
                    answer=answer,
                    explanation=explanation,
                    difficulty=float(difficulty),
                )
            )
            session.commit()

    add_resource(
        "DS-GEN-001",
        "\u793a\u4f8bMP4\u89c6\u9891",
        "https://interactive-examples.mdn.mozilla.net/media/cc0-videos/flower.mp4",
        "video",
    )

    add_question(
        "DS-GEN-001",
        "mcq",
        "\u5173\u4e8e\u7ebf\u6027\u8868\u7684\u8bf4\u6cd5\u6b63\u786e\u7684\u662f\uff1f",
        ["\u53ea\u80fd\u7528\u6570\u7ec4\u5b58\u50a8", "\u53ea\u80fd\u7528\u94fe\u8868\u5b58\u50a8", "\u53ef\u4ee5\u7528\u6570\u7ec4\u6216\u94fe\u8868\u5b58\u50a8", "\u4e0d\u80fd\u904d\u5386"],
        "A",
        "\u7ebf\u6027\u8868\u53ef\u4ee5\u987a\u5e8f\u6216\u94fe\u5f0f\u5b58\u50a8\u3002",
        0.3,
    )
    add_question(
        "DS-GEN-002",
        "blank",
        "\u6808\u7684\u7279\u70b9\u662f____\uff08\u5148\u5165\u540e\u51fa/\u540e\u5165\u5148\u51fa\uff09",
        [],
        "\u5148\u5165\u540e\u51fa",
        "\u6808\u662fLIFO\u7ed3\u6784\u3002",
        0.3,
    )

    _log_action(session, _admin, "seed_demo", f"subject={subject} kps={len(kps)}")
    return {"ok": True, "kps": len(kps)}


@router.post("/seed/full")
def seed_full_system(
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin)),
):
    """
    Seed a CS-only dataset (4 subjects, no grade) with full question banks.
    Idempotent: safe to run multiple times.
    """

    subjects: list[tuple[str, str, list[tuple[str, str]]]] = [
        (
            "\u6570\u636e\u7ed3\u6784",
            "DS",
            [
                ("DS-GEN-001", "\u7ebf\u6027\u8868\u57fa\u7840"),
                ("DS-GEN-002", "\u6808\u4e0e\u961f\u5217"),
                ("DS-GEN-003", "\u4e32"),
                ("DS-GEN-004", "\u6570\u7ec4\u4e0e\u77e9\u9635"),
                ("DS-GEN-005", "\u6811\u4e0e\u4e8c\u53c9\u6811"),
                ("DS-GEN-006", "\u56fe\u57fa\u7840"),
                ("DS-GEN-007", "\u67e5\u627e"),
                ("DS-GEN-008", "\u6392\u5e8f"),
                ("DS-GEN-009", "\u54c8\u5e0c\u8868"),
                ("DS-GEN-010", "\u5806\u4e0e\u4f18\u5148\u961f\u5217"),
            ],
        ),
        (
            "\u8ba1\u7b97\u673a\u7ec4\u6210\u539f\u7406",
            "CO",
            [
                ("CO-GEN-001", "\u6570\u5236\u4e0e\u7f16\u7801"),
                ("CO-GEN-002", "\u903b\u8f91\u7535\u8def\u57fa\u7840"),
                ("CO-GEN-003", "\u6307\u4ee4\u7cfb\u7edf"),
                ("CO-GEN-004", "CPU\u7ed3\u6784\u4e0e\u63a7\u5236"),
                ("CO-GEN-005", "\u6d41\u6c34\u7ebf"),
                ("CO-GEN-006", "\u5b58\u50a8\u5c42\u6b21"),
                ("CO-GEN-007", "\u8f93\u5165\u8f93\u51fa"),
                ("CO-GEN-008", "\u603b\u7ebf\u4e0e\u63a5\u53e3"),
                ("CO-GEN-009", "\u4e2d\u65ad\u4e0e\u5f02\u5e38"),
                ("CO-GEN-010", "\u6027\u80fd\u4e0e\u5e76\u884c"),
            ],
        ),
        (
            "\u64cd\u4f5c\u7cfb\u7edf",
            "OS",
            [
                ("OS-GEN-001", "\u64cd\u4f5c\u7cfb\u7edf\u6982\u8ff0"),
                ("OS-GEN-002", "\u8fdb\u7a0b\u4e0e\u7ebf\u7a0b"),
                ("OS-GEN-003", "CPU\u8c03\u5ea6"),
                ("OS-GEN-004", "\u540c\u6b65\u4e0e\u4e92\u65a5"),
                ("OS-GEN-005", "\u6b7b\u9501"),
                ("OS-GEN-006", "\u5185\u5b58\u7ba1\u7406"),
                ("OS-GEN-007", "\u865a\u62df\u5185\u5b58"),
                ("OS-GEN-008", "\u6587\u4ef6\u7cfb\u7edf"),
                ("OS-GEN-009", "I/O\u4e0e\u8bbe\u5907"),
                ("OS-GEN-010", "\u5b89\u5168\u4e0e\u4fdd\u62a4"),
            ],
        ),
        (
            "\u8ba1\u7b97\u673a\u7f51\u7edc",
            "CN",
            [
                ("CN-GEN-001", "\u7f51\u7edc\u4f53\u7cfb\u7ed3\u6784"),
                ("CN-GEN-002", "\u7269\u7406\u5c42"),
                ("CN-GEN-003", "\u6570\u636e\u94fe\u8def\u5c42"),
                ("CN-GEN-004", "\u4ecb\u8d28\u8bbf\u95ee\u63a7\u5236"),
                ("CN-GEN-005", "\u7f51\u7edc\u5c42"),
                ("CN-GEN-006", "\u8def\u7531\u4e0e\u8f6c\u53d1"),
                ("CN-GEN-007", "\u4f20\u8f93\u5c42"),
                ("CN-GEN-008", "\u5e94\u7528\u5c42"),
                ("CN-GEN-009", "\u7f51\u7edc\u5b89\u5168"),
                ("CN-GEN-010", "\u65e0\u7ebf\u4e0e\u79fb\u52a8\u7f51\u7edc"),
            ],
        ),
    ]

    grade_name = "\u901a\u7528"
    demo_course_codes = {"DS", "CO", "OS", "CN"}

    for username, password, role in [
        ("admin", "admin123", UserRole.admin),
        ("teacher1", "teacher123", UserRole.teacher),
        ("student1", "student123", UserRole.student),
        ("student2", "student123", UserRole.student),
        ("student3", "student123", UserRole.student),
    ]:
        exists = session.exec(select(User).where(User.username == username)).first()
        if exists is None:
            session.add(User(username=username, password_hash=hash_password(password), role=role))
    session.commit()
    teacher_user = session.exec(select(User).where(User.username == "teacher1")).first()
    admin_user = session.exec(select(User).where(User.username == "admin")).first()
    if teacher_user is not None and teacher_user.id is not None:
        if not str(teacher_user.full_name or "").strip():
            teacher_user.full_name = "\u5f20\u660e\uff08\u793a\u4f8b\u6559\u5e08\uff09"
            session.add(teacher_user)
            session.commit()
    if admin_user is not None and admin_user.id is not None:
        if not str(admin_user.full_name or "").strip():
            admin_user.full_name = "\u7cfb\u7edf\u7ba1\u7406\u5458"
            session.add(admin_user)
            session.commit()
    student_profiles: list[tuple[str, str, str, str]] = [
        ("student1", "\u674e\u6668", "2026001", "\u8ba1\u79d12301"),
        ("student2", "\u5468\u60a6", "2026002", "\u8ba1\u79d12301"),
        ("student3", "\u9648\u660a", "2026003", "\u8ba1\u79d12302"),
    ]
    for username, full_name, student_no, class_name in student_profiles:
        row = session.exec(select(User).where(User.username == username)).first()
        if row is None or row.id is None:
            continue
        row.full_name = full_name
        row.student_no = student_no
        row.class_name = class_name
        session.add(row)
    session.commit()

    def ensure_eval_config(subj: str) -> None:
        cfg = session.exec(select(EvalConfig).where(EvalConfig.subject == subj, EvalConfig.grade == grade_name)).first()
        if cfg is None:
            session.add(EvalConfig(subject=subj, grade=grade_name))
            session.commit()

    def ensure_course(subj: str, code: str) -> Course:
        now = datetime.utcnow()
        exists = session.exec(select(Course).where(Course.code == code)).first()
        if exists is None:
            exists = Course(
                code=code,
                title=subj,
                description=f"{subj}课程",
                teacher_id=int(teacher_user.id) if teacher_user and teacher_user.id is not None else None,
            )
            session.add(exists)
            session.commit()
            session.refresh(exists)
        elif teacher_user and teacher_user.id is not None and exists.teacher_id is None:
            exists.teacher_id = int(teacher_user.id)
            session.add(exists)
            session.commit()
            session.refresh(exists)
        if exists.id is not None and code in demo_course_codes:
            exists.lifecycle_status = CourseLifecycleStatus.active
            exists.active = True
            exists.start_at = now - timedelta(days=7)
            exists.end_at = now + timedelta(days=180)
            exists.target_class = "\u8ba1\u79d12301"
            exists.enroll_status = CourseEnrollStatus.open
            session.add(exists)
            session.commit()
        return exists

    def ensure_kp(subj: str, code: str, title: str) -> KnowledgePoint:
        kp = session.exec(select(KnowledgePoint).where(KnowledgePoint.code == code)).first()
        if kp is None:
            kp = KnowledgePoint(
                subject=subj,
                grade=grade_name,
                code=code,
                title=title,
                description=f"{title}\uff08\u793a\u4f8b\uff09",
            )
            session.add(kp)
            session.commit()
            session.refresh(kp)
        return kp

    def ensure_edge(subj: str, prereq_id: int, next_id: int) -> None:
        e = session.exec(
            select(KnowledgeEdge).where(KnowledgeEdge.prereq_id == prereq_id, KnowledgeEdge.next_id == next_id)
        ).first()
        if e is None:
            session.add(
                KnowledgeEdge(
                    subject=subj,
                    grade=grade_name,
                    prereq_id=prereq_id,
                    next_id=next_id,
                    relation_type=RelationType.prerequisite,
                )
            )
            session.commit()

    def ensure_question(
        subj: str,
        kp_id: int,
        qtype: str,
        prompt: str,
        options: list[str],
        answer: str,
        explanation: str,
        difficulty: float,
    ) -> None:
        q = session.exec(select(Question).where(Question.kp_id == kp_id, Question.prompt == prompt)).first()
        if q is None:
            session.add(
                Question(
                    subject=subj,
                    grade=grade_name,
                    kp_id=kp_id,
                    type=qtype,
                    prompt=prompt,
                    options_json=json.dumps(options, ensure_ascii=False),
                    answer=answer,
                    explanation=explanation,
                    difficulty=float(difficulty),
                )
            )
            session.commit()

    def ensure_quiz(kp: KnowledgePoint, pass_accuracy: float = 0.8) -> Quiz:
        quiz = session.exec(select(Quiz).where(Quiz.kp_id == kp.id)).first()
        if quiz is None:
            quiz = Quiz(subject=kp.subject, grade=kp.grade, kp_id=kp.id, pass_accuracy=pass_accuracy)
            session.add(quiz)
            session.commit()
            session.refresh(quiz)
        return quiz

    def ensure_quiz_item(quiz: Quiz, it: dict) -> None:
        exists = session.exec(select(QuizItem).where(QuizItem.quiz_id == quiz.id, QuizItem.prompt == it["prompt"])).first()
        if exists is None:
            session.add(
                QuizItem(
                    quiz_id=quiz.id,
                    type=it["type"],
                    prompt=it["prompt"],
                    options_json=json.dumps(it.get("options", []), ensure_ascii=False),
                    answer=it["answer"],
                    explanation=it.get("explanation", ""),
                    key_item=bool(it.get("key_item", False)),
                )
            )
            session.commit()

    def ensure_resource(kp: KnowledgePoint) -> None:
        title = f"{kp.title} \u793a\u4f8b\u89c6\u9891"
        exists = session.exec(select(LearningResource).where(LearningResource.kp_id == kp.id, LearningResource.title == title)).first()
        if exists is None:
            session.add(
                LearningResource(
                    subject=kp.subject,
                    grade=kp.grade,
                    kp_id=kp.id,
                    title=title,
                    url="https://interactive-examples.mdn.mozilla.net/media/cc0-videos/flower.mp4",
                    type=ResourceType.video,
                )
            )
            session.commit()

    created_kp = 0
    created_questions = 0

    difficulties = [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]

    for subj_name, _subj_code, kp_list in subjects:
        ensure_eval_config(subj_name)
        ensure_course(subj_name, _subj_code)
        kp_objs: list[KnowledgePoint] = []
        for code, title in kp_list:
            kp = ensure_kp(subj_name, code, title)
            kp_objs.append(kp)
            created_kp += 1
            ensure_resource(kp)

        for idx in range(len(kp_objs) - 1):
            ensure_edge(subj_name, kp_objs[idx].id, kp_objs[idx + 1].id)

        for kp in kp_objs:
            base = kp.code
            for i in range(20):
                diff = difficulties[i % len(difficulties)]
                options = [
                    f"{kp.title} \u76f8\u5173\u6982\u5ff5 {i + 1}",
                    f"{kp.title} \u7ed3\u8bba {i + 1}",
                    f"{kp.title} \u65b9\u6cd5 {i + 2}",
                    f"{kp.title} \u5e94\u7528 {i + 3}",
                ]
                answer = "A" if i % 2 == 0 else "B"
                prompt = f"{base}\uff1a\u5173\u4e8e{kp.title}\u7684\u9009\u62e9\u9898{i + 1}"
                ensure_question(subj_name, kp.id, "mcq", prompt, options, answer, "\u793a\u4f8b\u89e3\u6790", diff)
                created_questions += 1

            for i in range(20):
                diff = difficulties[i % len(difficulties)]
                prompt = f"{base}\uff1a{kp.title}\u586b\u7a7a____\uff08{i + 1}\uff09"
                answer = f"\u7b54\u6848{i + 1}"
                ensure_question(subj_name, kp.id, "blank", prompt, [], answer, "\u793a\u4f8b\u89e3\u6790", diff)
                created_questions += 1

            quiz = ensure_quiz(kp, pass_accuracy=0.8)
            ensure_quiz_item(
                quiz,
                {
                    "type": "mcq",
                    "prompt": f"{base}\uff1a\u5c0f\u6d4b\u9898\u76ee1",
                    "options": ["A. \u9009\u98791", "B. \u9009\u98792", "C. \u9009\u98793", "D. \u9009\u98794"],
                    "answer": "B",
                    "explanation": "\u793a\u4f8b\u89e3\u6790",
                    "key_item": True,
                },
            )
            ensure_quiz_item(
                quiz,
                {
                    "type": "mcq",
                    "prompt": f"{base}\uff1a\u5c0f\u6d4b\u9898\u76ee2",
                    "options": ["A. \u9009\u98791", "B. \u9009\u98792", "C. \u9009\u98793", "D. \u9009\u98794"],
                    "answer": "A",
                    "explanation": "\u793a\u4f8b\u89e3\u6790",
                },
            )
            ensure_quiz_item(
                quiz,
                {
                    "type": "blank",
                    "prompt": f"{base}\uff1a\u5c0f\u6d4b\u586b\u7a7a1____",
                    "options": [],
                    "answer": "\u793a\u4f8b\u7b54\u6848",
                    "explanation": "\u793a\u4f8b\u89e3\u6790",
                },
            )
            ensure_quiz_item(
                quiz,
                {
                    "type": "blank",
                    "prompt": f"{base}\uff1a\u5c0f\u6d4b\u586b\u7a7a2____",
                    "options": [],
                    "answer": "\u793a\u4f8b\u7b54\u6848",
                    "explanation": "\u793a\u4f8b\u89e3\u6790",
                },
            )

        auto_tag_knowledge_points(session, subject=subj_name, grade=grade_name, overwrite=False)

    created_enrollments = 0
    for uname in ("student1", "student2", "student3"):
        st = session.exec(select(User).where(User.username == uname)).first()
        if st is None or st.id is None:
            continue
        for code in demo_course_codes:
            c = session.exec(select(Course).where(Course.code == code)).first()
            if c is None or c.id is None:
                continue
            ex = session.exec(
                select(Enrollment).where(
                    Enrollment.course_id == int(c.id),
                    Enrollment.student_id == int(st.id),
                )
            ).first()
            if ex is not None:
                if ex.status != EnrollmentStatus.active:
                    ex.status = EnrollmentStatus.active
                    session.add(ex)
                    session.commit()
                continue
            session.add(
                Enrollment(
                    course_id=int(c.id),
                    student_id=int(st.id),
                    status=EnrollmentStatus.active,
                )
            )
            session.commit()
            created_enrollments += 1

    _log_action(
        session,
        _admin,
        "seed_full",
        f"created_kp={created_kp} created_questions={created_questions} created_enrollments={created_enrollments}",
    )
    return {
        "ok": True,
        "created_kp": created_kp,
        "created_questions": created_questions,
        "created_enrollments": created_enrollments,
    }


@router.get("/audit", response_model=PageOut)
def audit_logs(
    page: int = 1,
    page_size: int = 20,
    keyword: str | None = None,
    actor: str | None = None,
    action: str | None = None,
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    stmt = select(AuditLog)
    if keyword:
        like = f"%{keyword.strip()}%"
        stmt = stmt.where(
            or_(
                AuditLog.actor.like(like),
                AuditLog.role.like(like),
                AuditLog.action.like(like),
                AuditLog.detail.like(like),
            )
        )
    if actor:
        stmt = stmt.where(AuditLog.actor == actor)
    if action:
        stmt = stmt.where(AuditLog.action.contains(action))
    total = session.exec(select(func.count()).select_from(stmt.subquery())).first() or 0
    rows = session.exec(
        stmt.order_by(AuditLog.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    items = [
        AuditLogOut(
            id=r.id,
            actor=r.actor,
            role=r.role,
            action=r.action,
            detail=r.detail,
            created_at=r.created_at.isoformat(),
        )
        for r in rows
    ]
    return PageOut(items=items, total=int(total), page=page, page_size=page_size)


@router.get("/config")
def get_config(
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
    cfg = session.exec(select(EvalConfig).where(EvalConfig.subject == subject, EvalConfig.grade == grade)).first()
    if cfg is None:
        cfg = EvalConfig(subject=subject, grade=grade)
        session.add(cfg)
        session.commit()
        session.refresh(cfg)
    persona_rule = get_or_create_persona_rule(session, subject=subject, grade=grade)
    return {
        "weights": json.loads(cfg.weights_json),
        "thresholds": json.loads(cfg.thresholds_json),
        "window": json.loads(cfg.window_json),
        "persona": {
            "thresholds": resolve_persona_thresholds(persona_rule),
            "weights": resolve_persona_weights(persona_rule),
            "strategies": json.loads(persona_rule.strategy_json),
        },
    }


@router.put("/config")
def update_config(
    subject: str,
    grade: str,
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
    cfg = session.exec(select(EvalConfig).where(EvalConfig.subject == subject, EvalConfig.grade == grade)).first()
    if cfg is None:
        cfg = EvalConfig(subject=subject, grade=grade)
    cfg.weights_json = json.dumps(payload.get("weights", {}), ensure_ascii=False)
    cfg.thresholds_json = json.dumps(payload.get("thresholds", {}), ensure_ascii=False)
    cfg.window_json = json.dumps(payload.get("window", {}), ensure_ascii=False)
    session.add(cfg)
    persona_payload = payload.get("persona") or {}
    if persona_payload:
        rule = get_or_create_persona_rule(session, subject=subject, grade=grade)
        if "thresholds" in persona_payload:
            rule.thresholds_json = json.dumps(persona_payload.get("thresholds", {}), ensure_ascii=False)
        if "weights" in persona_payload:
            rule.weights_json = json.dumps(persona_payload.get("weights", {}), ensure_ascii=False)
        if "strategies" in persona_payload:
            rule.strategy_json = json.dumps(persona_payload.get("strategies", {}), ensure_ascii=False)
        session.add(rule)
    session.commit()
    return {"ok": True}


@router.get("/persona/rules", response_model=PersonaRuleOut)
def get_persona_rules(
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _require_teacher_subject_access(session=session, admin=admin, subject=subject, grade=grade)
    rule = get_or_create_persona_rule(session, subject=subject, grade=grade)
    return PersonaRuleOut(
        subject=subject,
        grade=grade,
        thresholds=resolve_persona_thresholds(rule),
        weights=resolve_persona_weights(rule),
        strategies=json.loads(rule.strategy_json),
    )


@router.put("/persona/rules", response_model=PersonaRuleOut)
def update_persona_rules(
    subject: str,
    grade: str,
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    rule = get_or_create_persona_rule(session, subject=subject, grade=grade)
    if "thresholds" in payload:
        rule.thresholds_json = json.dumps(payload.get("thresholds", {}), ensure_ascii=False)
    if "weights" in payload:
        rule.weights_json = json.dumps(payload.get("weights", {}), ensure_ascii=False)
    if "strategies" in payload:
        rule.strategy_json = json.dumps(payload.get("strategies", {}), ensure_ascii=False)
    session.add(rule)
    session.commit()
    session.refresh(rule)
    _log_action(session, admin, "persona_rule_update", f"subject={subject} grade={grade}")
    return PersonaRuleOut(
        subject=subject,
        grade=grade,
        thresholds=resolve_persona_thresholds(rule),
        weights=resolve_persona_weights(rule),
        strategies=json.loads(rule.strategy_json),
    )


@router.post("/persona/recalculate")
def recalculate_persona(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    subject = str(payload.get("subject", "")).strip()
    grade = str(payload.get("grade", "")).strip()
    refresh_mastery = bool(payload.get("refresh_mastery", False))
    if not subject or not grade:
        raise HTTPException(status_code=400, detail="subject/grade required")
    course = session.exec(select(Course).where(Course.title == subject)).first()
    if admin.role == UserRole.teacher and course is not None and not teacher_has_course_access(session, int(admin.id), course):
        raise HTTPException(status_code=403, detail="No permission for this subject")
    snapshots = recalculate_profiles_for_subject(
        session,
        subject=subject,
        grade=grade,
        refresh_mastery=refresh_mastery,
    )
    _log_action(session, admin, "persona_recalculate", f"subject={subject} grade={grade} count={len(snapshots)}")
    return {"ok": True, "count": len(snapshots)}


@router.get("/persona/overrides")
def list_persona_overrides(
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    _admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    rows = session.exec(
        select(LearnerPersonaOverride)
        .where(LearnerPersonaOverride.subject == subject, LearnerPersonaOverride.grade == grade)
        .order_by(LearnerPersonaOverride.updated_at.desc())
    ).all()
    return {
        "items": [
            PersonaOverrideOut(
                user_id=row.user_id,
                subject=row.subject,
                grade=row.grade,
                persona_type=row.persona_type.value,
                note=row.note,
                updated_by=row.updated_by,
                updated_at=row.updated_at.isoformat(),
            ).model_dump()
            for row in rows
        ]
    }


@router.get("/persona/students")
def list_persona_students(
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    student_stmt = select(User).where(User.role == UserRole.student)
    if course is not None and course.id is not None:
        active_student_ids = _active_course_student_ids(session, course_id=int(course.id))
        if not active_student_ids:
            return {"items": []}
        student_stmt = student_stmt.where(User.id.in_(active_student_ids))
    students = session.exec(student_stmt.order_by(User.id)).all()
    items = []
    for student in students:
        if student.id is None:
            continue
        snapshot = _snapshot_for_user(session, user_id=int(student.id), subject=subject, grade=grade)
        stage_snapshot = get_latest_stage_snapshot(session, user_id=int(student.id), subject=subject, grade=grade)
        items.append(
            {
                "user_id": int(student.id),
                "username": student.username,
                "full_name": student.full_name,
                "student_no": student.student_no,
                "class_name": student.class_name,
                "persona_type": snapshot.persona_type.value,
                "persona_label": persona_label(snapshot.persona_type),
                "dynamic_score": float(snapshot.dynamic_score),
                "course_mastery": float(snapshot.course_mastery),
                "risk_level": snapshot.risk_level,
                "override_source": snapshot.override_source,
                "reason_summary": snapshot.reason_summary,
                "updated_at": snapshot.updated_at.isoformat(),
                "latest_stage_title": stage_snapshot.stage_title if stage_snapshot is not None else "",
                "stage_trend": stage_snapshot.trend_label if stage_snapshot is not None else "",
            }
        )
    return {"items": items}


@router.put("/persona/override", response_model=PersonaOverrideOut)
def set_persona_override(
    payload: PersonaOverrideIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin)),
):
    try:
        persona_type = PersonaType(payload.persona_type)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid persona_type")
    override = upsert_persona_override(
        session,
        user_id=payload.user_id,
        subject=payload.subject,
        grade=payload.grade,
        persona_type=persona_type,
        note=payload.note,
        updated_by=admin.username,
    )
    sync_profile_snapshot_from_stage(
        session,
        user_id=payload.user_id,
        subject=payload.subject,
        grade=payload.grade,
        persist=True,
    ) or recalculate_profile_snapshot(
        session,
        user_id=payload.user_id,
        subject=payload.subject,
        grade=payload.grade,
        refresh_mastery=False,
        persist=True,
    )
    _log_action(session, admin, "persona_override_set", f"user_id={payload.user_id} persona={payload.persona_type}")
    return PersonaOverrideOut(
        user_id=override.user_id,
        subject=override.subject,
        grade=override.grade,
        persona_type=override.persona_type.value,
        note=override.note,
        updated_by=override.updated_by,
        updated_at=override.updated_at.isoformat(),
    )


@router.delete("/persona/override")
def delete_persona_override(
    user_id: int,
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin)),
):
    ok = clear_persona_override(session, user_id=user_id, subject=subject, grade=grade)
    sync_profile_snapshot_from_stage(
        session,
        user_id=user_id,
        subject=subject,
        grade=grade,
        persist=True,
    ) or recalculate_profile_snapshot(
        session,
        user_id=user_id,
        subject=subject,
        grade=grade,
        refresh_mastery=False,
        persist=True,
    )
    _log_action(session, admin, "persona_override_delete", f"user_id={user_id} subject={subject} grade={grade}")
    return {"ok": ok}


@router.get("/stage-feedback")
def get_stage_feedback(
    user_id: int,
    subject: str,
    grade: str,
    stage_id: int | None = None,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    row = get_stage_teacher_feedback(session, user_id=user_id, subject=subject, grade=grade, stage_id=stage_id)
    if row is None:
        return None
    return {
        "id": int(row.id),
        "user_id": int(row.user_id),
        "stage_id": int(row.stage_id),
        "feedback_tag": row.feedback_tag,
        "comment": row.comment,
        "updated_by": row.updated_by,
        "updated_at": row.updated_at.isoformat(),
    }


@router.get("/stage-feedback/history")
def get_stage_feedback_history(
    user_id: int,
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)

    rows = session.exec(
        select(StageTeacherFeedback)
        .where(
            StageTeacherFeedback.user_id == user_id,
            StageTeacherFeedback.subject == subject,
            StageTeacherFeedback.grade == grade,
        )
        .order_by(StageTeacherFeedback.updated_at.desc())
    ).all()

    stage_ids = sorted({int(row.stage_id) for row in rows})
    stage_map: dict[int, CourseStage] = {}
    if stage_ids:
        stages = session.exec(select(CourseStage).where(CourseStage.id.in_(stage_ids))).all()
        stage_map = {int(stage.id): stage for stage in stages if stage.id is not None}

    return {
        "items": [
            {
                "id": int(row.id),
                "user_id": int(row.user_id),
                "stage_id": int(row.stage_id),
                "stage_title": stage_map.get(int(row.stage_id)).title if int(row.stage_id) in stage_map else "",
                "stage_order": stage_map.get(int(row.stage_id)).stage_order if int(row.stage_id) in stage_map else None,
                "feedback_tag": row.feedback_tag,
                "comment": row.comment,
                "updated_by": row.updated_by,
                "updated_at": row.updated_at.isoformat(),
            }
            for row in rows
            if row.id is not None
        ]
    }


@router.put("/stage-feedback")
def save_stage_feedback(
    payload: dict,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    user_id = int(payload.get("user_id"))
    stage_id = int(payload.get("stage_id"))
    feedback_tag = str(payload.get("feedback_tag", "")).strip()
    comment = str(payload.get("comment", "")).strip()
    stage = session.get(CourseStage, stage_id)
    if stage is None:
        raise HTTPException(status_code=404, detail="Stage not found")
    course = session.get(Course, stage.course_id)
    if admin.role == UserRole.teacher and course is not None and not teacher_has_course_access(session, int(admin.id), course):
        raise HTTPException(status_code=403, detail="No permission for this subject")
    row = upsert_stage_teacher_feedback(
        session,
        user_id=user_id,
        stage_id=stage_id,
        subject=stage.subject,
        grade=stage.grade,
        course_id=stage.course_id,
        feedback_tag=feedback_tag,
        comment=comment,
        updated_by=admin.username,
    )
    recalculate_stage_snapshots_for_stage(
        session,
        stage_id=stage_id,
        user_ids=[user_id],
        persist=True,
    )
    _log_action(session, admin, "stage_feedback_save", f"user_id={user_id} stage_id={stage_id}")
    return {
        "id": int(row.id),
        "user_id": int(row.user_id),
        "stage_id": int(row.stage_id),
        "feedback_tag": row.feedback_tag,
        "comment": row.comment,
        "updated_by": row.updated_by,
        "updated_at": row.updated_at.isoformat(),
        "recalculated": True,
    }


@router.get("/analytics/overview", response_model=AdminAnalyticsOut)
def analytics_overview(
    subject: str | None = None,
    grade: str | None = None,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    if not subject or not grade:
        first_kp = session.exec(select(KnowledgePoint).order_by(KnowledgePoint.id)).first()
        if first_kp is None:
            return AdminAnalyticsOut(
                subject=subject or "",
                grade=grade or "",
                total_students=0,
                persona_distribution=[],
                stage_summary=[],
                latest_stage=None,
                risk_students=[],
                weak_kps=[],
                progress_ranking=[],
                ability_practice_cohort={},
            )
        subject = subject or first_kp.subject
        grade = grade or first_kp.grade

    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)

    student_stmt = select(User).where(User.role == UserRole.student)
    if course is not None and course.id is not None:
        active_student_ids = _active_course_student_ids(session, course_id=int(course.id))
        if not active_student_ids:
            return AdminAnalyticsOut(
                subject=subject,
                grade=grade,
                total_students=0,
                persona_distribution=[],
                stage_summary=[],
                latest_stage=None,
                risk_students=[],
                weak_kps=[],
                progress_ranking=[],
                ability_practice_cohort={},
            )
        student_stmt = student_stmt.where(User.id.in_(active_student_ids))
    students = session.exec(student_stmt.order_by(User.id)).all()
    student_ids = [int(item.id) for item in students if item.id is not None]
    snapshots = [
        _snapshot_for_user(session, user_id=int(student.id), subject=subject, grade=grade)
        for student in students
        if student.id is not None
    ]
    snapshots = [snapshot for snapshot in snapshots if snapshot is not None]
    latest_stage_map: dict[int, StageEvaluationSnapshot] = {}
    stage_rows = session.exec(
        select(StageEvaluationSnapshot)
        .where(StageEvaluationSnapshot.user_id.in_(student_ids) if student_ids else False)
        .order_by(StageEvaluationSnapshot.stage_order, StageEvaluationSnapshot.updated_at)
    ).all()
    subject_text = str(subject or "").strip()
    grade_text = str(grade or "").strip()
    all_stage_rows = [
        row
        for row in stage_rows
        if str(row.subject or "").strip() == subject_text and str(row.grade or "").strip() == grade_text
    ]
    stage_bucket: dict[int, dict] = {}
    for row in all_stage_rows:
        latest_stage_map[int(row.user_id)] = row
        bucket = stage_bucket.setdefault(
            int(row.stage_id),
            {
                "stage_id": int(row.stage_id),
                "stage_title": row.stage_title,
                "stage_order": int(row.stage_order),
                "student_count": 0,
                "avg_dynamic_score": 0.0,
                "avg_course_mastery": 0.0,
                "risk_count": 0,
                "progress_count": 0,
                "steady_count": 0,
                "regress_count": 0,
            },
        )
        bucket["student_count"] += 1
        bucket["avg_dynamic_score"] += float(row.dynamic_score)
        bucket["avg_course_mastery"] += float(row.course_mastery)
        if row.risk_level == "风险":
            bucket["risk_count"] += 1
        if row.trend_label == "进步":
            bucket["progress_count"] += 1
        elif row.trend_label == "退步":
            bucket["regress_count"] += 1
        else:
            bucket["steady_count"] += 1

    stage_summary = []
    for item in sorted(stage_bucket.values(), key=lambda row: int(row["stage_order"])):
        total = max(1, int(item["student_count"]))
        stage_summary.append(
            {
                **item,
                "avg_dynamic_score": float(item["avg_dynamic_score"]) / total,
                "avg_course_mastery": float(item["avg_course_mastery"]) / total,
            }
        )
    latest_stage = stage_summary[-1] if stage_summary else None

    persona_distribution_map: dict[str, int] = {}
    for snapshot in snapshots:
        label = persona_label(snapshot.persona_type)
        persona_distribution_map[label] = persona_distribution_map.get(label, 0) + 1
    persona_distribution = [
        {"persona_label": key, "count": value}
        for key, value in sorted(persona_distribution_map.items(), key=lambda item: item[1], reverse=True)
    ]

    student_map = {int(item.id): item for item in students if item.id is not None}
    risk_students = [
        {
            "user_id": int(snapshot.user_id),
            "username": student_map[int(snapshot.user_id)].username if int(snapshot.user_id) in student_map else "",
            "full_name": student_map[int(snapshot.user_id)].full_name if int(snapshot.user_id) in student_map else "",
            "persona_label": persona_label(snapshot.persona_type),
            "latest_stage_title": latest_stage_map[int(snapshot.user_id)].stage_title if int(snapshot.user_id) in latest_stage_map else "",
            "stage_trend": latest_stage_map[int(snapshot.user_id)].trend_label if int(snapshot.user_id) in latest_stage_map else "",
            "dynamic_score": float(snapshot.dynamic_score),
            "risk_level": snapshot.risk_level,
            "reason_summary": snapshot.reason_summary,
        }
        for snapshot in sorted(snapshots, key=lambda item: (float(item.dynamic_score), float(item.course_mastery)))[:10]
    ]

    kps = session.exec(
        select(KnowledgePoint)
        .where(KnowledgePoint.subject == subject, KnowledgePoint.grade == grade)
        .order_by(KnowledgePoint.chapter, KnowledgePoint.id)
    ).all()
    kp_ids = [int(kp.id) for kp in kps if kp.id is not None]
    mastery_rows = []
    if kp_ids and student_ids:
        mastery_rows = session.exec(
            select(Mastery).where(Mastery.kp_id.in_(kp_ids), Mastery.user_id.in_(student_ids))
        ).all()
    mastery_bucket: dict[int, list[float]] = {int(kp.id): [] for kp in kps if kp.id is not None}
    for row in mastery_rows:
        mastery_bucket.setdefault(int(row.kp_id), []).append(float(row.value))
    weak_kps = []
    for kp in kps:
        if kp.id is None:
            continue
        values = mastery_bucket.get(int(kp.id), [])
        avg_value = (sum(values) / len(student_ids)) if student_ids else 0.0
        weak_kps.append(
            {
                "kp_id": int(kp.id),
                "code": kp.code,
                "title": kp.title,
                "chapter": kp.chapter,
                "avg_mastery": avg_value,
            }
        )
    weak_kps.sort(key=lambda item: item["avg_mastery"])

    cohort_user_ids = student_ids
    if course is not None and course.id is not None:
        enr_rows = session.exec(
            select(Enrollment).where(
                Enrollment.course_id == int(course.id),
                Enrollment.status == EnrollmentStatus.active,
            )
        ).all()
        cohort_user_ids = [int(r.student_id) for r in enr_rows if r.student_id is not None]
    ability_practice_cohort = build_cohort_ability_practice_summary(
        session,
        kp_ids=kp_ids,
        user_ids=cohort_user_ids,
    )

    progress_ranking = [
        {
            "user_id": int(snapshot.user_id),
            "username": student_map[int(snapshot.user_id)].username if int(snapshot.user_id) in student_map else "",
            "full_name": student_map[int(snapshot.user_id)].full_name if int(snapshot.user_id) in student_map else "",
            "persona_label": persona_label(snapshot.persona_type),
            "latest_stage_title": latest_stage_map[int(snapshot.user_id)].stage_title if int(snapshot.user_id) in latest_stage_map else "",
            "stage_trend": latest_stage_map[int(snapshot.user_id)].trend_label if int(snapshot.user_id) in latest_stage_map else "",
            "course_mastery": float(snapshot.course_mastery),
            "dynamic_score": float(snapshot.dynamic_score),
        }
        for snapshot in sorted(snapshots, key=lambda item: float(item.dynamic_score), reverse=True)[:20]
    ]

    return AdminAnalyticsOut(
        subject=subject,
        grade=grade,
        total_students=len(students),
        persona_distribution=persona_distribution,
        stage_summary=stage_summary,
        latest_stage=latest_stage,
        risk_students=risk_students,
        weak_kps=weak_kps[:10],
        progress_ranking=progress_ranking,
        ability_practice_cohort=ability_practice_cohort,
    )


@router.get("/analytics/student-detail")
def analytics_student_detail(
    user_id: int,
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    if course is not None and course.id is not None:
        active_student_ids = set(_active_course_student_ids(session, course_id=int(course.id)))
        if user_id not in active_student_ids:
            raise HTTPException(status_code=404, detail="Student not enrolled in this course")
    return _build_student_detail_payload(
        session,
        user_id=user_id,
        subject=subject,
        grade=grade,
        course=course,
    )


@router.get("/final-score/students")
def list_final_score_students(
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    student_stmt = select(User).where(User.role == UserRole.student)
    if course is not None and course.id is not None:
        active_student_ids = _active_course_student_ids(session, course_id=int(course.id))
        if not active_student_ids:
            return {"items": []}
        student_stmt = student_stmt.where(User.id.in_(active_student_ids))
    students = session.exec(student_stmt.order_by(User.id)).all()
    confirmation_map: dict[int, TeacherFinalScoreConfirmation] = {}
    if course is not None and course.id is not None:
        rows = session.exec(
            select(TeacherFinalScoreConfirmation).where(TeacherFinalScoreConfirmation.course_id == int(course.id))
        ).all()
        confirmation_map = {int(row.user_id): row for row in rows}
    items = []
    for student in students:
        if student.id is None:
            continue
        snapshot = _snapshot_for_user(session, user_id=int(student.id), subject=subject, grade=grade)
        term_summary = _json_load(snapshot.portrait_summary_json, {}).get("term_summary", {})
        confirmation = confirmation_map.get(int(student.id))
        items.append(
            {
                "user_id": int(student.id),
                "username": student.username,
                "full_name": student.full_name,
                "student_no": student.student_no,
                "class_name": student.class_name,
                "persona_label": persona_label(snapshot.persona_type),
                "dynamic_score": float(snapshot.dynamic_score),
                "course_mastery": float(snapshot.course_mastery),
                "risk_level": snapshot.risk_level,
                "suggested_score": float(term_summary.get("final_score_reference", 0.0)),
                "confirmed_score": float(confirmation.confirmed_score) if confirmation is not None else None,
                "confirmed_level": confirmation.confirmed_level if confirmation is not None else "",
                "confirmed_at": confirmation.confirmed_at.isoformat() if confirmation is not None else None,
            }
        )
    return {"items": items}


@router.get("/final-score/detail")
def final_score_detail(
    user_id: int,
    subject: str,
    grade: str,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _resolve_course_for_subject(session, subject=subject, grade=grade, admin=admin)
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    if course is not None and course.id is not None:
        active_student_ids = set(_active_course_student_ids(session, course_id=int(course.id)))
        if user_id not in active_student_ids:
            raise HTTPException(status_code=404, detail="Student not enrolled in this course")
    return _build_student_detail_payload(
        session,
        user_id=user_id,
        subject=subject,
        grade=grade,
        course=course,
    )


@router.put("/final-score/confirm")
def confirm_final_score(
    payload: TeacherFinalScoreConfirmIn,
    session: Session = Depends(get_session),
    admin=Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _resolve_course_for_subject(session, subject=payload.subject, grade=payload.grade, admin=admin)
    if course is None or course.id is None:
        raise HTTPException(status_code=404, detail="课程不存在")
    _check_teacher_subject_access(session=session, admin=admin, course=course)
    active_student_ids = set(_active_course_student_ids(session, course_id=int(course.id)))
    if payload.user_id not in active_student_ids:
        raise HTTPException(status_code=404, detail="Student not enrolled in this course")
    student = session.get(User, payload.user_id)
    if student is None or student.role != UserRole.student:
        raise HTTPException(status_code=404, detail="Student not found")
    snapshot = _snapshot_for_user(session, user_id=payload.user_id, subject=payload.subject, grade=payload.grade)
    term_summary = _json_load(snapshot.portrait_summary_json, {}).get("term_summary", {})
    if int(term_summary.get("stage_count") or 0) <= 0:
        raise HTTPException(status_code=400, detail="No stage evaluation data for this student")
    confirmed_score = max(0.0, min(1.0, float(payload.confirmed_score)))
    row = session.exec(
        select(TeacherFinalScoreConfirmation).where(
            TeacherFinalScoreConfirmation.user_id == payload.user_id,
            TeacherFinalScoreConfirmation.course_id == int(course.id),
        )
    ).first()
    if row is None:
        row = TeacherFinalScoreConfirmation(
            user_id=payload.user_id,
            course_id=int(course.id),
            subject=payload.subject,
            grade=payload.grade,
        )
    row.subject = payload.subject
    row.grade = payload.grade
    row.suggested_score = float(term_summary.get("final_score_reference", 0.0))
    row.confirmed_score = confirmed_score
    row.confirmed_level = payload.confirmed_level.strip()
    row.comment = payload.comment.strip()
    row.recommendation_summary = payload.recommendation_summary.strip()
    row.confirmed_by = admin.username
    row.confirmed_at = datetime.utcnow()
    row.updated_at = datetime.utcnow()
    session.add(row)
    session.commit()
    session.refresh(row)
    _log_action(
        session,
        admin,
        "final_score_confirm",
        f"user_id={payload.user_id} subject={payload.subject} confirmed_score={round(confirmed_score * 100)}",
    )
    return {
        "ok": True,
        "id": int(row.id),
        "suggested_score": float(row.suggested_score),
        "confirmed_score": float(row.confirmed_score),
        "confirmed_level": row.confirmed_level,
        "comment": row.comment,
        "recommendation_summary": row.recommendation_summary,
        "confirmed_by": row.confirmed_by,
        "confirmed_at": row.confirmed_at.isoformat(),
        "updated_at": row.updated_at.isoformat(),
    }
