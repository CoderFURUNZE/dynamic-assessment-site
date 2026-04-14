import csv
import json
import logging
from datetime import datetime
from io import BytesIO, StringIO

from fastapi import HTTPException, UploadFile
from sqlmodel import Session, select

from app.api.deps import teacher_has_course_access
from app.db.models import AuditLog, Course, CourseLifecycleStatus, CourseStage, KnowledgePoint, StageMetricType, User, UserRole
from app.schemas.stage import CourseStageOut, StageMetricGuideOut

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


logger = logging.getLogger("app.audit")

STAGE_IMPORT_GUIDES: dict[StageMetricType, dict[str, object]] = {
    StageMetricType.video: {
        "label": "视频学习记录",
        "summary": "用于反映学生在本阶段的视频学习投入、学习连续性和资源使用情况。",
        "template_fields": ["username", "student_no", "kp_code", "watched_minutes", "completion_ratio", "happened_at", "note"],
        "affected_dimensions": ["学习行为与过程", "情感与社会性发展"],
        "affected_indicators": ["学习动机与态度", "辅助学习策略", "资源偏好：图像/空间型"],
        "next_action": "导入完成后，系统会重算本阶段的视频学习结果，教师可在学生详情中查看变化。",
    },
    StageMetricType.assignment: {
        "label": "作业完成记录",
        "summary": "用于反映学生在本阶段的作业完成质量、按时提交情况和知识理解表现。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "submitted_on_time", "happened_at", "note"],
        "affected_dimensions": ["知识与认知状态", "情感与社会性发展"],
        "affected_indicators": ["语言类知识掌握度", "逻辑类知识掌握度", "自我调节与元认知"],
        "next_action": "导入完成后，系统会重算作业相关指标，教师可结合补充评价继续修正画像。",
    },
    StageMetricType.quiz: {
        "label": "小测成绩记录",
        "summary": "用于反映学生在本阶段的小测表现、知识掌握情况和认知层级变化。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "duration_minutes", "happened_at", "note"],
        "affected_dimensions": ["知识与认知状态", "学习行为与过程"],
        "affected_indicators": ["学科能力层级与认知路径", "逻辑类知识掌握度", "跨学科知识关联能力"],
        "next_action": "导入完成后，系统会重算小测相关指标，教师可在阶段结果里查看变化趋势。",
    },
    StageMetricType.attendance: {
        "label": "考勤记录",
        "summary": "用于反映学生在本阶段的到课情况、稳定性和学习投入基础表现。",
        "template_fields": ["username", "student_no", "attendance_value", "status", "happened_at", "note"],
        "affected_dimensions": ["情感与社会性发展", "学习行为与过程"],
        "affected_indicators": ["学习动机与态度", "自我调节与元认知", "辅助学习策略"],
        "next_action": "导入完成后，系统会重算考勤与投入稳定性，教师可在学生详情中对比阶段变化。",
    },
    StageMetricType.task: {
        "label": "任务完成记录",
        "summary": "用于反映学生在本阶段的任务完成度、实践执行和迁移表现。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "status", "happened_at", "note"],
        "affected_dimensions": ["潜能与特质倾向", "知识与认知状态", "学习行为与过程"],
        "affected_indicators": ["跨情境迁移能力", "实践/体验型交互偏好", "学科能力层级与认知路径"],
        "next_action": "导入完成后，系统会重算任务表现与迁移相关指标。",
    },
    StageMetricType.participation: {
        "label": "课堂参与记录",
        "summary": "用于反映学生在本阶段的课堂活跃度、协作表现和交流偏好。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "status", "happened_at", "note"],
        "affected_dimensions": ["情感与社会性发展", "学习行为与过程"],
        "affected_indicators": ["协作能力与社交网络", "文本/讨论型交互偏好", "学习动机与态度"],
        "next_action": "导入完成后，系统会重算参与度与协作相关结果。",
    },
}

IMPORT_FIELD_ALIASES: dict[str, list[str]] = {
    "username": ["账号", "用户名"],
    "student_no": ["学号"],
    "kp_id": ["知识点ID"],
    "kp_code": ["知识点编码", "知识点代码"],
    "watched_minutes": ["视频学习分钟", "观看分钟"],
    "completion_ratio": ["完成率"],
    "completion_value": ["完成度"],
    "happened_at": ["发生时间", "记录时间"],
    "note": ["备注"],
    "score": ["分数", "得分"],
    "submitted_on_time": ["是否按时提交", "按时提交"],
    "duration_minutes": ["时长分钟", "用时分钟"],
    "attendance_value": ["考勤值", "出勤值"],
    "status": ["状态"],
}


def log_action(session: Session, user: User, action: str, detail: str = "") -> None:
    try:
        logger.info("actor=%s role=%s action=%s detail=%s", user.username, user.role.value, action, detail)
        session.add(AuditLog(actor=user.username, role=user.role.value, action=action, detail=detail))
        session.commit()
    except Exception:
        logger.info("action=%s detail=%s", action, detail)


def parse_dt(value: str | None) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw)
    except ValueError as exc:
        raise ValueError(f"invalid datetime: {raw}") from exc


def get_course_or_403(session: Session, user: User, course_id: int) -> Course:
    course = session.get(Course, course_id)
    if course is None:
        raise HTTPException(status_code=404, detail="course not found")
    if user.role == UserRole.teacher and not teacher_has_course_access(session, int(user.id), course):
        raise HTTPException(status_code=403, detail="forbidden")
    return course


def assert_course_stage_editable(course: Course) -> None:
    lifecycle = course.lifecycle_status.value if isinstance(course.lifecycle_status, CourseLifecycleStatus) else str(course.lifecycle_status or "")
    if not bool(course.active) or lifecycle != CourseLifecycleStatus.active.value:
        raise HTTPException(status_code=400, detail="当前课程已归档或未处于开课状态，不能修改阶段数据")


def _looks_broken_text(value: str | None) -> bool:
    text = str(value or "").strip()
    return not text or "?" in text or "�" in text


def _normalized_stage_identity(stage: CourseStage, course: Course | None = None) -> tuple[str, str]:
    subject = str(stage.subject or "").strip()
    grade = str(stage.grade or "").strip()
    if _looks_broken_text(subject):
        subject = str((course.title if course is not None else "") or "").strip() or subject or "未命名课程"
    if _looks_broken_text(grade):
        grade = "通用"
    return subject, grade


def get_stage_or_403(session: Session, user: User, stage_id: int) -> CourseStage:
    stage = session.get(CourseStage, stage_id)
    if stage is None:
        raise HTTPException(status_code=404, detail="stage not found")
    course = get_course_or_403(session, user, int(stage.course_id))
    subject, grade = _normalized_stage_identity(stage, course)
    if subject != stage.subject or grade != stage.grade:
        stage.subject = subject
        stage.grade = grade
        session.add(stage)
        session.commit()
        session.refresh(stage)
    return stage


def course_out(stage: CourseStage, course: Course | None = None) -> CourseStageOut:
    subject, grade = _normalized_stage_identity(stage, course)
    return CourseStageOut(
        id=int(stage.id),
        course_id=stage.course_id,
        subject=subject,
        grade=grade,
        title=stage.title,
        stage_order=stage.stage_order,
        starts_at=stage.starts_at.isoformat() if stage.starts_at else None,
        ends_at=stage.ends_at.isoformat() if stage.ends_at else None,
        description=stage.description,
        created_at=stage.created_at.isoformat(),
    )


def metric_guide(metric_type: StageMetricType) -> StageMetricGuideOut:
    payload = STAGE_IMPORT_GUIDES[metric_type]
    return StageMetricGuideOut(
        metric_type=metric_type.value,
        label=str(payload["label"]),
        summary=str(payload["summary"]),
        template_fields=list(payload["template_fields"]),
        affected_dimensions=list(payload["affected_dimensions"]),
        affected_indicators=list(payload["affected_indicators"]),
        next_action=str(payload["next_action"]),
    )


def template_csv(metric_type: StageMetricType) -> str:
    rows = {
        StageMetricType.video: ["账号,学号,知识点编码,视频学习分钟,完成率,发生时间,备注", "student1,2026001,CN-GEN-001,45,80%,2026-03-10,视频学习样例"],
        StageMetricType.assignment: ["账号,学号,知识点编码,分数,完成率,是否按时提交,发生时间,备注", "student1,2026001,CN-GEN-001,88,100%,是,2026-03-10,作业完成样例"],
        StageMetricType.quiz: ["账号,学号,知识点编码,分数,完成率,时长分钟,发生时间,备注", "student1,2026001,CN-GEN-001,76,76%,18,2026-03-10,小测样例"],
        StageMetricType.attendance: ["账号,学号,考勤值,状态,发生时间,备注", "student1,2026001,1,出勤,2026-03-10,考勤样例"],
        StageMetricType.task: ["账号,学号,知识点编码,分数,完成率,状态,发生时间,备注", "student1,2026001,CN-GEN-001,90,100%,完成,2026-03-10,任务完成样例"],
        StageMetricType.participation: ["账号,学号,知识点编码,分数,完成率,状态,发生时间,备注", "student1,2026001,CN-GEN-001,1,100%,积极,2026-03-10,课堂参与样例"],
    }
    return "\n".join(rows[metric_type])


def build_import_summary(*, course: Course, stage: CourseStage, metric_type: str, total_rows: int, success_rows: int, failed_rows: int, recalculated_users: int, source_mode: str) -> dict[str, object]:
    success_rate = (float(success_rows) / float(total_rows)) if total_rows else 0.0
    quality_status = "excellent" if failed_rows == 0 else "warning" if success_rate >= 0.8 else "risk"
    return {
        "course_id": int(course.id) if course.id is not None else None,
        "course_title": course.title,
        "stage_id": int(stage.id) if stage.id is not None else None,
        "stage_title": stage.title,
        "stage_order": int(stage.stage_order),
        "metric_type": metric_type,
        "source_mode": source_mode,
        "success_rate": round(success_rate, 4),
        "quality_status": quality_status,
        "recalculated_at": datetime.utcnow().isoformat(),
        "recalculation_scope": f"已按当前阶段重算 {recalculated_users} 名学生的阶段画像与动态评价",
        "quality_hint": "本次数据质量良好，可直接进入学生分析或学习报告查看结果。" if failed_rows == 0 else "本次仍有失败记录，建议先根据错误提示修正后再补导一次。",
    }


def normalize_import_row(row: dict[str, str]) -> dict[str, str]:
    normalized = {str(key).strip(): "" if value is None else str(value).strip() for key, value in row.items()}
    for canonical, aliases in IMPORT_FIELD_ALIASES.items():
        if canonical in normalized and normalized[canonical]:
            continue
        for alias in aliases:
            if alias in normalized and normalized[alias]:
                normalized[canonical] = normalized[alias]
                break
    return normalized


def rows_from_upload(file: UploadFile, payload: bytes) -> list[dict[str, str]]:
    name = (file.filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [normalize_import_row(row) for row in reader]
    if name.endswith(".xlsx"):
        if load_workbook is None:
            raise HTTPException(status_code=400, detail="xlsx import requires openpyxl")
        wb = load_workbook(BytesIO(payload), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(item).strip() if item is not None else "" for item in values[0]]
        rows: list[dict[str, str]] = []
        for value_row in values[1:]:
            item: dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                cell = value_row[idx] if idx < len(value_row) else ""
                item[header] = "" if cell is None else str(cell).strip()
            rows.append(normalize_import_row(item))
        return rows
    raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")


def find_user(session: Session, row: dict[str, str]) -> User:
    username = (row.get("username") or "").strip()
    student_no = (row.get("student_no") or "").strip()
    user = None
    if username:
        user = session.exec(select(User).where(User.username == username)).first()
    elif student_no:
        user = session.exec(select(User).where(User.student_no == student_no)).first()
    if user is None:
        raise ValueError("student not found by username/student_no")
    return user


def find_kp(session: Session, *, subject: str, grade: str, row: dict[str, str]) -> KnowledgePoint | None:
    kp_id = (row.get("kp_id") or "").strip()
    kp_code = (row.get("kp_code") or "").strip()
    if kp_id:
        kp = session.get(KnowledgePoint, int(kp_id))
        if kp is None:
            raise ValueError(f"kp_id not found: {kp_id}")
        return kp
    if kp_code:
        kp = session.exec(select(KnowledgePoint).where(KnowledgePoint.code == kp_code, KnowledgePoint.subject == subject, KnowledgePoint.grade == grade)).first()
        if kp is None:
            raise ValueError(f"kp_code not found: {kp_code}")
        return kp
    return None


def to_float(value: str | None, *, default: float = 0.0) -> float:
    raw = (value or "").strip()
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError as exc:
        raise ValueError(f"invalid numeric value: {raw}") from exc


def to_ratio(value: str | None) -> float:
    raw = (value or "").strip()
    if not raw:
        return 0.0
    if raw.endswith("%"):
        raw = raw[:-1].strip()
    try:
        ratio = float(raw)
    except ValueError as exc:
        raise ValueError(f"invalid ratio value: {raw}") from exc
    if ratio > 1.0 and ratio <= 100.0:
        ratio = ratio / 100.0
    return max(0.0, min(1.0, ratio))


def to_bool(value: str | None) -> bool:
    return (value or "").strip().lower() in {"1", "true", "yes", "y", "是", "已提交", "on_time"}
