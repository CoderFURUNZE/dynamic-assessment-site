import csv
import json
import logging
from collections import Counter
from datetime import datetime, timedelta
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, Response, UploadFile
from sqlmodel import Session, select
from sqlalchemy import false, or_

from app.api.deps import require_role
from app.api.routers import stage_support
from app.db.models import (
    AuditLog,
    Course,
    CourseLifecycleStatus,
    CourseStage,
    Enrollment,
    EnrollmentStatus,
    ExpressionEvent,
    KnowledgePoint,
    LearnerProfileSnapshot,
    LearningBehaviorEvent,
    Mastery,
    PracticeAttempt,
    QuestionnairePortraitIndicatorInput,
    RecommendationLog,
    StageImportBatch,
    StageImportRecord,
    StageEvaluationSnapshot,
    StageMetricType,
    User,
    UserRole,
    VideoProgress,
)
from app.db.session import get_session
from app.schemas.stage import (
    CourseStageIn,
    CourseStageOut,
    CourseStageUpdateIn,
    StageImportBatchOut,
    StageMetricGuideOut,
    StageImportResultOut,
)
from app.services.learner_profile import recalculate_stage_snapshots_for_stage

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependency guard
    load_workbook = None


router = APIRouter(prefix="/stages", tags=["stages"])
logger = logging.getLogger("app.audit")


STAGE_IMPORT_GUIDES: dict[StageMetricType, dict[str, object]] = {
    StageMetricType.video: {
        "label": "视频学习记录",
        "summary": "用于反映学生在本阶段的视频学习投入、学习连续性和资源使用情况。",
        "template_fields": ["username", "student_no", "kp_code", "watched_minutes", "completion_ratio", "happened_at", "note"],
        "affected_dimensions": ["学习行为与过程", "情感与社会性发展"],
        "affected_indicators": ["学习动机与态度", "辅助学习策略", "资源偏好：图像/空间型"],
        "next_action": "导入完成后，系统会重算本阶段的投入类与资源偏好类结果，老师可去“单学生详情”查看阶段画像。",
    },
    StageMetricType.assignment: {
        "label": "作业完成记录",
        "summary": "用于反映学生在本阶段的作业完成质量、按时提交情况和知识理解表现。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "submitted_on_time", "happened_at", "note"],
        "affected_dimensions": ["知识与认知状态", "情感与社会性发展"],
        "affected_indicators": ["语言类知识掌握度", "逻辑类知识掌握度", "自我调节与元认知"],
        "next_action": "导入完成后，系统会重算知识掌握与学习习惯结果，老师可结合补充评价继续修正画像。",
    },
    StageMetricType.quiz: {
        "label": "小测成绩记录",
        "summary": "用于反映学生在本阶段的小测表现、知识掌握情况和认知层级变化。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "duration_minutes", "happened_at", "note"],
        "affected_dimensions": ["知识与认知状态", "学习行为与过程"],
        "affected_indicators": ["学科能力层级与认知路径", "逻辑类知识掌握度", "跨学科知识关联能力"],
        "next_action": "导入完成后，系统会重算本阶段的认知状态相关指标，老师可在阶段结果里查看变化趋势。",
    },
    StageMetricType.attendance: {
        "label": "考勤记录",
        "summary": "用于反映学生在本阶段的到课情况、稳定性和学习投入基础表现。",
        "template_fields": ["username", "student_no", "attendance_value", "status", "happened_at", "note"],
        "affected_dimensions": ["情感与社会性发展", "学习行为与过程"],
        "affected_indicators": ["学习动机与态度", "自我调节与元认知", "辅助学习策略"],
        "next_action": "导入完成后，系统会重算本阶段的出勤与投入稳定性，老师可在学生详情中对比阶段变化。",
    },
    StageMetricType.task: {
        "label": "任务完成记录",
        "summary": "用于反映学生在本阶段的任务完成度、实践执行和迁移表现。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "status", "happened_at", "note"],
        "affected_dimensions": ["潜能与特质倾向", "知识与认知状态", "学习行为与过程"],
        "affected_indicators": ["跨情境迁移能力", "实践/体验型交互偏好", "学科能力层级与认知路径"],
        "next_action": "导入完成后，系统会重算任务表现与迁移相关指标，老师可再补充高阶能力评分。",
    },
    StageMetricType.participation: {
        "label": "课堂参与记录",
        "summary": "用于反映学生在本阶段的课堂活跃度、协作表现和交流偏好。",
        "template_fields": ["username", "student_no", "kp_code", "score", "completion_ratio", "status", "happened_at", "note"],
        "affected_dimensions": ["情感与社会性发展", "学习行为与过程"],
        "affected_indicators": ["协作能力与社交网络", "文本/讨论型交互偏好", "学习动机与态度"],
        "next_action": "导入完成后，系统会重算参与度与协作相关结果，老师可结合课堂观察补充评分。",
    },
}

IMPORT_FIELD_ALIASES: dict[str, list[str]] = {
    "username": ["账号", "用户名"],
    "student_no": ["学号"],
    "kp_id": ["知识点ID"],
    "kp_code": ["知识点编码", "知识点代码"],
    "watched_minutes": ["视频学习分钟", "观看分钟"],
    "completion_ratio": ["完成率"],
    "completion_value": ["完成度"],
    "happened_at": ["发生时间", "记录时间"],
    "note": ["备注"],
    "score": ["分数", "得分"],
    "submitted_on_time": ["是否按时提交", "按时提交"],
    "duration_minutes": ["时长分钟", "用时分钟"],
    "attendance_value": ["考勤值", "出勤值"],
    "status": ["状态"],
}


def _log_action(session: Session, user: User, action: str, detail: str = "") -> None:
    stage_support.log_action(session, user, action, detail)


def _parse_dt(value: str | None) -> datetime | None:
    return stage_support.parse_dt(value)


def _course_out(stage: CourseStage, course: Course | None = None) -> CourseStageOut:
    return stage_support.course_out(stage, course)


def _metric_type_value(value: StageMetricType | str) -> str:
    return value.value if isinstance(value, StageMetricType) else str(value)


def _get_course_or_403(session: Session, user: User, course_id: int) -> Course:
    return stage_support.get_course_or_403(session, user, course_id)


def _assert_course_stage_editable(course: Course) -> None:
    lifecycle = course.lifecycle_status.value if isinstance(course.lifecycle_status, CourseLifecycleStatus) else str(course.lifecycle_status or "")
    if not bool(course.active) or lifecycle != CourseLifecycleStatus.active.value:
        raise HTTPException(status_code=400, detail="当前课程已归档或未处于开课状态，不能修改阶段数据")


def _get_stage_or_403(session: Session, user: User, stage_id: int) -> CourseStage:
    return stage_support.get_stage_or_403(session, user, stage_id)


def _find_user(session: Session, row: dict[str, str]) -> User:
    return stage_support.find_user(session, row)


def _find_kp(session: Session, *, subject: str, grade: str, row: dict[str, str]) -> KnowledgePoint | None:
    return stage_support.find_kp(session, subject=subject, grade=grade, row=row)


def _to_float(value: str | None, *, default: float = 0.0) -> float:
    return stage_support.to_float(value, default=default)


def _to_ratio(value: str | None) -> float:
    return stage_support.to_ratio(value)


def _to_bool(value: str | None) -> bool:
    raw = (value or "").strip().lower()
    return raw in {"1", "true", "yes", "y", "是", "已提交", "on_time"}


def _clamp01(value: float) -> float:
    return max(0.0, min(1.0, float(value)))


def _build_import_summary(
    *,
    course: Course,
    stage: CourseStage,
    metric_type: str,
    total_rows: int,
    success_rows: int,
    failed_rows: int,
    recalculated_users: int,
    source_mode: str,
) -> dict[str, object]:
    success_rate = (float(success_rows) / float(total_rows)) if total_rows else 0.0
    quality_status = "excellent" if failed_rows == 0 else "warning" if success_rate >= 0.8 else "risk"
    return {
        "course_id": int(course.id) if course.id is not None else None,
        "course_title": course.title,
        "stage_id": int(stage.id) if stage.id is not None else None,
        "stage_title": stage.title,
        "stage_order": int(stage.stage_order),
        "metric_type": metric_type,
        "source_mode": source_mode,
        "success_rate": round(success_rate, 4),
        "quality_status": quality_status,
        "recalculated_at": datetime.utcnow().isoformat(),
        "recalculation_scope": f"已按当前阶段重算 {recalculated_users} 名学生的阶段画像与动态评价",
        "quality_hint": (
            "本次数据质量良好，可直接进入学生分析或学习报告查看结果。"
            if failed_rows == 0
            else "本次仍有失败记录，建议先根据错误提示修正后再补导一次。"
        ),
    }


def _risk_level(dynamic_score: float) -> str:
    value = float(dynamic_score or 0.0)
    if value >= 0.85:
        return "优秀"
    if value >= 0.70:
        return "良好"
    if value >= 0.50:
        return "预警"
    return "风险"


def _normalize_import_row(row: dict[str, str]) -> dict[str, str]:
    normalized = {str(key).strip(): "" if value is None else str(value).strip() for key, value in row.items()}
    for canonical, aliases in IMPORT_FIELD_ALIASES.items():
        if canonical in normalized and normalized[canonical]:
            continue
        for alias in aliases:
            if alias in normalized and normalized[alias]:
                normalized[canonical] = normalized[alias]
                break
    return normalized


def _rows_from_upload(file: UploadFile, payload: bytes) -> list[dict[str, str]]:
    name = (file.filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [_normalize_import_row(row) for row in reader]
    if name.endswith(".xlsx"):
        if load_workbook is None:
            raise HTTPException(status_code=400, detail="xlsx import requires openpyxl")
        wb = load_workbook(BytesIO(payload), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(item).strip() if item is not None else "" for item in values[0]]
        rows: list[dict[str, str]] = []
        for value_row in values[1:]:
            item: dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                cell = value_row[idx] if idx < len(value_row) else ""
                item[header] = "" if cell is None else str(cell).strip()
            rows.append(_normalize_import_row(item))
        return rows
    raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")


def _template_csv(metric_type: StageMetricType) -> str:
    rows = {
        StageMetricType.video: [
            "账号,学号,知识点编码,视频学习分钟,完成率,发生时间,备注",
            "student1,2026001,CN-GEN-001,45,80%,2026-03-10,视频学习样例",
        ],
        StageMetricType.assignment: [
            "账号,学号,知识点编码,分数,完成率,是否按时提交,发生时间,备注",
            "student1,2026001,CN-GEN-001,88,100%,是,2026-03-10,作业完成样例",
        ],
        StageMetricType.quiz: [
            "账号,学号,知识点编码,分数,完成率,时长分钟,发生时间,备注",
            "student1,2026001,CN-GEN-001,76,76%,18,2026-03-10,小测样例",
        ],
        StageMetricType.attendance: [
            "账号,学号,考勤值,状态,发生时间,备注",
            "student1,2026001,1,出勤,2026-03-10,考勤样例",
        ],
        StageMetricType.task: [
            "账号,学号,知识点编码,分数,完成率,状态,发生时间,备注",
            "student1,2026001,CN-GEN-001,90,100%,完成,2026-03-10,任务完成样例",
        ],
        StageMetricType.participation: [
            "账号,学号,知识点编码,分数,完成率,状态,发生时间,备注",
            "student1,2026001,CN-GEN-001,1,100%,积极,2026-03-10,课堂参与样例",
        ],
    }
    return "\n".join(rows[metric_type])


def _metric_guide(metric_type: StageMetricType) -> StageMetricGuideOut:
    payload = STAGE_IMPORT_GUIDES[metric_type]
    return StageMetricGuideOut(
        metric_type=metric_type.value,
        label=str(payload["label"]),
        summary=str(payload["summary"]),
        template_fields=list(payload["template_fields"]),
        affected_dimensions=list(payload["affected_dimensions"]),
        affected_indicators=list(payload["affected_indicators"]),
        next_action=str(payload["next_action"]),
    )


def _stage_window(stage: CourseStage) -> tuple[datetime, datetime]:
    start = stage.starts_at or (stage.created_at - timedelta(days=14))
    end = stage.ends_at or datetime.utcnow()
    if end < start:
        end = start
    return start, end


def _json_payload(raw: str | None) -> dict[str, object]:
    if not raw:
        return {}
    try:
        value = json.loads(raw)
        return value if isinstance(value, dict) else {}
    except Exception:
        return {}


def _internal_stage_rows(session: Session, *, course: Course, stage: CourseStage) -> tuple[dict[str, object], list[dict[str, object]]]:
    enrollments = session.exec(
        select(Enrollment).where(
            Enrollment.course_id == int(course.id),
            Enrollment.status == EnrollmentStatus.active,
        )
    ).all()
    student_ids = [int(item.student_id) for item in enrollments if item.student_id is not None]
    if not student_ids:
        return {
            "course_id": int(course.id),
            "stage_id": int(stage.id),
            "stage_title": stage.title,
            "student_count": 0,
            "video_students": 0,
            "practice_students": 0,
            "questionnaire_students": 0,
            "recommendation_students": 0,
        }, []

    students = session.exec(select(User).where(User.id.in_(student_ids))).all()
    student_map = {int(item.id): item for item in students if item.id is not None}
    kp_ids = [
        int(item)
        for item in session.exec(
            select(KnowledgePoint.id).where(
                KnowledgePoint.subject == stage.subject,
                KnowledgePoint.grade == stage.grade,
            )
        ).all()
        if item is not None
    ]
    start, end = _stage_window(stage)

    video_rows = []
    practice_rows = []
    if kp_ids:
        video_rows = session.exec(
            select(VideoProgress).where(
                VideoProgress.user_id.in_(student_ids),
                VideoProgress.kp_id.in_(kp_ids),
                VideoProgress.updated_at >= start,
                VideoProgress.updated_at <= end,
            )
        ).all()
        practice_rows = session.exec(
            select(PracticeAttempt).where(
                PracticeAttempt.user_id.in_(student_ids),
                PracticeAttempt.kp_id.in_(kp_ids),
                PracticeAttempt.created_at >= start,
                PracticeAttempt.created_at <= end,
            )
        ).all()
    recommendation_rows = session.exec(
        select(RecommendationLog).where(
            RecommendationLog.user_id.in_(student_ids),
            RecommendationLog.subject == stage.subject,
            RecommendationLog.grade == stage.grade,
            RecommendationLog.created_at >= start,
            RecommendationLog.created_at <= end,
        )
    ).all()
    questionnaire_rows = session.exec(
        select(QuestionnairePortraitIndicatorInput).where(
            QuestionnairePortraitIndicatorInput.user_id.in_(student_ids),
            QuestionnairePortraitIndicatorInput.course_id == int(course.id),
            QuestionnairePortraitIndicatorInput.updated_at >= start,
            QuestionnairePortraitIndicatorInput.updated_at <= end,
        )
    ).all()

    mastery_rows = []
    if kp_ids:
        mastery_rows = session.exec(
            select(Mastery).where(
                Mastery.user_id.in_(student_ids),
                Mastery.kp_id.in_(kp_ids),
            )
        ).all()
    profile_rows = session.exec(
        select(LearnerProfileSnapshot).where(
            LearnerProfileSnapshot.user_id.in_(student_ids),
            LearnerProfileSnapshot.subject == stage.subject,
            LearnerProfileSnapshot.grade == stage.grade,
        )
    ).all()

    video_map: dict[int, list[VideoProgress]] = {}
    for row in video_rows:
        video_map.setdefault(int(row.user_id), []).append(row)
    practice_map: dict[int, list[PracticeAttempt]] = {}
    for row in practice_rows:
        practice_map.setdefault(int(row.user_id), []).append(row)
    reco_map: dict[int, list[RecommendationLog]] = {}
    for row in recommendation_rows:
        reco_map.setdefault(int(row.user_id), []).append(row)
    questionnaire_map: dict[int, list[QuestionnairePortraitIndicatorInput]] = {}
    for row in questionnaire_rows:
        questionnaire_map.setdefault(int(row.user_id), []).append(row)
    mastery_map: dict[int, list[Mastery]] = {}
    for row in mastery_rows:
        mastery_map.setdefault(int(row.user_id), []).append(row)
    profile_map = {int(row.user_id): row for row in profile_rows if row.user_id is not None}

    rows: list[dict[str, object]] = []
    for student_id in student_ids:
        student = student_map.get(student_id)
        if student is None:
            continue
        videos = video_map.get(student_id, [])
        practices = practice_map.get(student_id, [])
        recos = reco_map.get(student_id, [])
        questionnaires = questionnaire_map.get(student_id, [])
        mastery_items = mastery_map.get(student_id, [])
        snapshot = profile_map.get(student_id)

        watched_minutes = sum(max(0.0, float(item.watched_seconds or 0.0)) for item in videos) / 60.0
        completion_values = [
            min(1.0, max(0.0, float(item.watched_seconds or 0.0) / float(item.duration_seconds)))
            for item in videos
            if float(item.duration_seconds or 0.0) > 0
        ]
        practice_total = len(practices)
        practice_correct = len([item for item in practices if bool(item.correct)])
        avg_mastery = (
            sum(float(item.value or 0.0) for item in mastery_items) / len(mastery_items)
            if mastery_items
            else 0.0
        )

        rows.append(
            {
                "user_id": student_id,
                "username": student.username,
                "student_no": student.student_no or "",
                "full_name": student.full_name or "",
                "class_name": student.class_name or "",
                "video_records": len(videos),
                "watched_minutes": round(watched_minutes, 1),
                "avg_video_completion": round((sum(completion_values) / len(completion_values)) if completion_values else 0.0, 4),
                "practice_attempts": practice_total,
                "practice_accuracy": round((practice_correct / practice_total) if practice_total else 0.0, 4),
                "recommendation_count": len(recos),
                "questionnaire_updates": len(questionnaires),
                "course_mastery": round(avg_mastery, 4),
                "dynamic_score": round(float(snapshot.dynamic_score or 0.0), 4) if snapshot is not None else 0.0,
                "risk_level": snapshot.risk_level if snapshot is not None else "暂无",
            }
        )

    summary = {
        "course_id": int(course.id),
        "stage_id": int(stage.id),
        "stage_title": stage.title,
        "student_count": len(rows),
        "video_students": len([row for row in rows if float(row["watched_minutes"]) > 0]),
        "practice_students": len([row for row in rows if int(row["practice_attempts"]) > 0]),
        "questionnaire_students": len([row for row in rows if int(row["questionnaire_updates"]) > 0]),
        "recommendation_students": len([row for row in rows if int(row["recommendation_count"]) > 0]),
    }
    return summary, rows


_BEHAVIOR_POSITIVE_WEIGHTS = {
    "resource_visit": 1.0,
    "resource_download": 1.2,
    "graph_view": 0.9,
    "recommendation_open": 0.7,
    "recommendation_click": 0.7,
    "recommend_click": 0.7,
    "video_progress": 1.0,
    "video_play": 0.8,
    "video_complete": 1.2,
    "practice_submit": 1.1,
    "quiz_submit": 1.15,
    "note_create": 0.9,
    "questionnaire_submit": 0.8,
    "stage_view": 0.5,
    "course_view": 0.5,
}

_BEHAVIOR_NEGATIVE_WEIGHTS = {
    "distracted": 1.2,
    "fidgeting": 1.0,
    "no_face": 0.8,
    "eat": 1.1,
    "eating": 1.1,
    "offtask": 1.0,
    "sleepy": 1.0,
    "confused": 0.6,
    "strained": 0.8,
}

_BEHAVIOR_POSITIVE_LABELS = {"focused", "relaxed", "neutral"}
_BEHAVIOR_NEGATIVE_LABELS = {"distracted", "fidgeting", "no_face", "confused", "strained", "eat", "eating", "offtask", "sleepy"}


def _behavior_signal_score(*, event_type: str, payload: dict[str, object], expression_label: str | None = None, confidence: float = 0.0) -> tuple[float, float]:
    positive = 0.0
    negative = 0.0
    normalized_type = (event_type or "").strip().lower()
    if normalized_type in _BEHAVIOR_POSITIVE_WEIGHTS:
        positive += float(_BEHAVIOR_POSITIVE_WEIGHTS[normalized_type])
    elif normalized_type in _BEHAVIOR_NEGATIVE_WEIGHTS:
        negative += float(_BEHAVIOR_NEGATIVE_WEIGHTS[normalized_type])
    else:
        positive += 0.2

    label = (expression_label or str(payload.get("label") or payload.get("emotion") or "")).strip().lower()
    if label in _BEHAVIOR_POSITIVE_LABELS:
        positive += 0.9 * max(0.3, float(confidence) or 0.0)
    elif label in _BEHAVIOR_NEGATIVE_LABELS:
        negative += 0.9 * max(0.3, float(confidence) or 0.0)

    hint = str(payload.get("signal") or payload.get("type") or payload.get("kind") or "").strip().lower()
    if hint in _BEHAVIOR_POSITIVE_WEIGHTS:
        positive += 0.4
    elif hint in _BEHAVIOR_NEGATIVE_WEIGHTS:
        negative += 0.4

    return positive, negative


def _behavior_stage_rows(session: Session, *, course: Course, stage: CourseStage) -> tuple[dict[str, object], list[dict[str, object]]]:
    enrollments = session.exec(
        select(Enrollment).where(
            Enrollment.course_id == int(course.id),
            Enrollment.status == EnrollmentStatus.active,
        )
    ).all()
    student_ids = [int(item.student_id) for item in enrollments if item.student_id is not None]
    if not student_ids:
        return {
            "course_id": int(course.id),
            "stage_id": int(stage.id),
            "stage_title": stage.title,
            "student_count": 0,
            "behavior_students": 0,
            "expression_students": 0,
            "positive_events": 0,
            "negative_events": 0,
        }, []

    students = session.exec(select(User).where(User.id.in_(student_ids))).all()
    student_map = {int(item.id): item for item in students if item.id is not None}
    kp_ids = [
        int(item)
        for item in session.exec(
            select(KnowledgePoint.id).where(
                KnowledgePoint.subject == stage.subject,
                KnowledgePoint.grade == stage.grade,
            )
        ).all()
        if item is not None
    ]
    start, end = _stage_window(stage)

    behavior_stmt = select(LearningBehaviorEvent).where(
        LearningBehaviorEvent.user_id.in_(student_ids),
        LearningBehaviorEvent.created_at >= start,
        LearningBehaviorEvent.created_at <= end,
    )
    if kp_ids:
        behavior_stmt = behavior_stmt.where(or_(LearningBehaviorEvent.course_id == int(course.id), LearningBehaviorEvent.kp_id.in_(kp_ids)))
    else:
        behavior_stmt = behavior_stmt.where(LearningBehaviorEvent.course_id == int(course.id))
    behavior_rows = session.exec(behavior_stmt).all()

    expression_stmt = select(ExpressionEvent).where(
        ExpressionEvent.user_id.in_(student_ids),
        ExpressionEvent.created_at >= start,
        ExpressionEvent.created_at <= end,
    )
    if kp_ids:
        expression_stmt = expression_stmt.where(ExpressionEvent.kp_id.in_(kp_ids))
    else:
        expression_stmt = expression_stmt.where(false())
    expression_rows = session.exec(expression_stmt).all()

    behavior_map: dict[int, list[LearningBehaviorEvent]] = {}
    for row in behavior_rows:
        behavior_map.setdefault(int(row.user_id), []).append(row)
    expression_map: dict[int, list[ExpressionEvent]] = {}
    for row in expression_rows:
        expression_map.setdefault(int(row.user_id), []).append(row)
    stage_snapshot_rows = session.exec(
        select(StageEvaluationSnapshot).where(
            StageEvaluationSnapshot.user_id.in_(student_ids),
            StageEvaluationSnapshot.stage_id == int(stage.id),
        )
    ).all()
    stage_snapshot_map = {int(row.user_id): row for row in stage_snapshot_rows if row.user_id is not None}
    profile_rows = session.exec(
        select(LearnerProfileSnapshot).where(
            LearnerProfileSnapshot.user_id.in_(student_ids),
            LearnerProfileSnapshot.subject == stage.subject,
            LearnerProfileSnapshot.grade == stage.grade,
        )
    ).all()
    profile_map = {int(row.user_id): row for row in profile_rows if row.user_id is not None}

    rows: list[dict[str, object]] = []
    for student_id in student_ids:
        student = student_map.get(student_id)
        if student is None:
            continue
        behaviors = behavior_map.get(student_id, [])
        expressions = expression_map.get(student_id, [])
        active_days = set()
        positive_total = 0.0
        negative_total = 0.0
        confidence_values: list[float] = []
        behavior_counter: Counter[str] = Counter()

        for row in behaviors:
            payload = _json_payload(row.value_json)
            active_days.add(row.created_at.date())
            behavior_counter[row.event_type] += 1
            positive, negative = _behavior_signal_score(event_type=row.event_type, payload=payload)
            positive_total += positive
            negative_total += negative
            if payload.get("confidence") is not None:
                try:
                    confidence_values.append(float(payload.get("confidence") or 0.0))
                except Exception:
                    pass

        expression_focus = 0
        expression_distracted = 0
        for row in expressions:
            payload = {"label": row.label, "difficulty": row.difficulty, "confidence": row.confidence}
            active_days.add(row.created_at.date())
            behavior_counter[f"expression:{row.label}"] += 1
            positive, negative = _behavior_signal_score(
                event_type=row.label,
                payload=payload,
                expression_label=row.label,
                confidence=float(row.confidence or 0.0),
            )
            positive_total += positive
            negative_total += negative
            label = str(row.label or "").strip().lower()
            if label in _BEHAVIOR_POSITIVE_LABELS:
                expression_focus += 1
            elif label in _BEHAVIOR_NEGATIVE_LABELS:
                expression_distracted += 1
            if row.confidence is not None:
                confidence_values.append(float(row.confidence))

        total_events = len(behaviors) + len(expressions)
        active_day_count = len(active_days)
        signal_balance = positive_total - negative_total
        behavior_score = _clamp01(0.46 + 0.09 * (positive_total ** 0.5) + 0.025 * min(active_day_count, 7) - 0.10 * negative_total)
        if total_events == 0:
            behavior_score = 0.0
        avg_confidence = sum(confidence_values) / len(confidence_values) if confidence_values else 0.0
        stage_snapshot = stage_snapshot_map.get(student_id)
        profile_snapshot = profile_map.get(student_id)
        if stage_snapshot is not None:
            dynamic_score = float(stage_snapshot.dynamic_score or 0.0)
        elif profile_snapshot is not None:
            dynamic_score = float(profile_snapshot.dynamic_score or 0.0)
        else:
            dynamic_score = behavior_score
        rows.append(
            {
                "user_id": student_id,
                "username": student.username,
                "student_no": student.student_no or "",
                "full_name": student.full_name or "",
                "class_name": student.class_name or "",
                "behavior_events": total_events,
                "active_days": active_day_count,
                "positive_events": round(positive_total, 2),
                "negative_events": round(negative_total, 2),
                "expression_events": len(expressions),
                "expression_focus": expression_focus,
                "expression_distracted": expression_distracted,
                "avg_confidence": round(avg_confidence, 4),
                "behavior_score": round(behavior_score, 4),
                "signal_balance": round(signal_balance, 2),
                "dominant_signal": "积极" if behavior_score >= 0.65 and signal_balance >= 0 else ("分心" if negative_total > positive_total else "观察"),
                "dynamic_score": round(dynamic_score, 4),
                "risk_level": _risk_level(dynamic_score),
            }
        )

    summary = {
        "course_id": int(course.id),
        "stage_id": int(stage.id),
        "stage_title": stage.title,
        "student_count": len(rows),
        "behavior_students": len([row for row in rows if int(row["behavior_events"]) > 0]),
        "expression_students": len([row for row in rows if int(row["expression_events"]) > 0]),
        "positive_events": int(round(sum(float(row["positive_events"]) for row in rows))),
        "negative_events": int(round(sum(float(row["negative_events"]) for row in rows))),
    }
    return summary, rows


def _apply_behavior_stage_rows(
    session: Session,
    *,
    course: Course,
    stage: CourseStage,
    user: User,
) -> tuple[StageImportBatch, int]:
    enrollments = session.exec(
        select(Enrollment).where(
            Enrollment.course_id == int(course.id),
            Enrollment.status == EnrollmentStatus.active,
        )
    ).all()
    student_ids = [int(item.student_id) for item in enrollments if item.student_id is not None]
    if not student_ids:
        batch = StageImportBatch(
            course_id=int(course.id),
            stage_id=int(stage.id),
            subject=stage.subject,
            grade=stage.grade,
            metric_type=StageMetricType.participation,
            file_name="system_behavior_summary",
            uploaded_by=user.username,
            total_rows=0,
            success_rows=0,
            failed_rows=0,
            error_json="[]",
        )
        session.add(batch)
        session.commit()
        session.refresh(batch)
        return batch, 0

    students = session.exec(select(User).where(User.id.in_(student_ids))).all()
    student_map = {int(item.id): item for item in students if item.id is not None}
    kp_ids = [
        int(item)
        for item in session.exec(
            select(KnowledgePoint.id).where(
                KnowledgePoint.subject == stage.subject,
                KnowledgePoint.grade == stage.grade,
            )
        ).all()
        if item is not None
    ]
    start, end = _stage_window(stage)

    behavior_stmt = select(LearningBehaviorEvent).where(
        LearningBehaviorEvent.user_id.in_(student_ids),
        LearningBehaviorEvent.created_at >= start,
        LearningBehaviorEvent.created_at <= end,
    )
    if kp_ids:
        behavior_stmt = behavior_stmt.where(or_(LearningBehaviorEvent.course_id == int(course.id), LearningBehaviorEvent.kp_id.in_(kp_ids)))
    else:
        behavior_stmt = behavior_stmt.where(LearningBehaviorEvent.course_id == int(course.id))
    behavior_rows = session.exec(behavior_stmt).all()

    expression_stmt = select(ExpressionEvent).where(
        ExpressionEvent.user_id.in_(student_ids),
        ExpressionEvent.created_at >= start,
        ExpressionEvent.created_at <= end,
    )
    if kp_ids:
        expression_stmt = expression_stmt.where(ExpressionEvent.kp_id.in_(kp_ids))
    else:
        expression_stmt = expression_stmt.where(false())
    expression_rows = session.exec(expression_stmt).all()

    per_student_day: dict[tuple[int, datetime.date], dict[str, object]] = {}

    def bucket(user_id: int, day: datetime.date) -> dict[str, object]:
        key = (user_id, day)
        item = per_student_day.setdefault(
            key,
            {
                "behavior_events": 0,
                "active_days": 0,
                "positive_events": 0.0,
                "negative_events": 0.0,
                "expression_events": 0,
                "expression_focus": 0,
                "expression_distracted": 0,
                "confidence_values": [],
                "signal_counter": Counter(),
            },
        )
        item["active_days"] = 1
        return item

    for row in behavior_rows:
        payload = _json_payload(row.value_json)
        day = row.created_at.date()
        item = bucket(int(row.user_id), day)
        item["behavior_events"] = int(item["behavior_events"]) + 1
        item["signal_counter"][row.event_type] += 1
        positive, negative = _behavior_signal_score(event_type=row.event_type, payload=payload)
        item["positive_events"] = float(item["positive_events"]) + positive
        item["negative_events"] = float(item["negative_events"]) + negative
        if payload.get("confidence") is not None:
            try:
                item["confidence_values"].append(float(payload.get("confidence") or 0.0))
            except Exception:
                pass

    for row in expression_rows:
        day = row.created_at.date()
        item = bucket(int(row.user_id), day)
        payload = {"label": row.label, "difficulty": row.difficulty, "confidence": row.confidence}
        item["behavior_events"] = int(item["behavior_events"]) + 1
        item["expression_events"] = int(item["expression_events"]) + 1
        item["signal_counter"][f"expression:{row.label}"] += 1
        positive, negative = _behavior_signal_score(
            event_type=row.label,
            payload=payload,
            expression_label=row.label,
            confidence=float(row.confidence or 0.0),
        )
        item["positive_events"] = float(item["positive_events"]) + positive
        item["negative_events"] = float(item["negative_events"]) + negative
        label = str(row.label or "").strip().lower()
        if label in _BEHAVIOR_POSITIVE_LABELS:
            item["expression_focus"] = int(item["expression_focus"]) + 1
        elif label in _BEHAVIOR_NEGATIVE_LABELS:
            item["expression_distracted"] = int(item["expression_distracted"]) + 1
        if row.confidence is not None:
            item["confidence_values"].append(float(row.confidence))

    existing_rows = session.exec(
        select(StageImportRecord).where(
            StageImportRecord.stage_id == int(stage.id),
            StageImportRecord.status == "behavior_auto",
        )
    ).all()
    for row in existing_rows:
        session.delete(row)
    session.commit()

    batch = StageImportBatch(
        course_id=int(course.id),
        stage_id=int(stage.id),
        subject=stage.subject,
        grade=stage.grade,
        metric_type=StageMetricType.participation,
        file_name="system_behavior_summary",
        uploaded_by=user.username,
        total_rows=len(per_student_day),
        success_rows=0,
        failed_rows=0,
        error_json="[]",
    )
    session.add(batch)
    session.commit()
    session.refresh(batch)

    affected_users: set[int] = set()
    success = 0
    for (user_id, day), item in sorted(per_student_day.items(), key=lambda x: (x[0][0], x[0][1])):
        student = student_map.get(user_id)
        if student is None:
            continue
        behavior_events = int(item["behavior_events"])
        active_days = int(item["active_days"])
        positive_events = float(item["positive_events"])
        negative_events = float(item["negative_events"])
        expression_events = int(item["expression_events"])
        confidence_values = list(item["confidence_values"])
        behavior_score = _clamp01(0.46 + 0.09 * (positive_events ** 0.5) + 0.025 * min(1, active_days) - 0.10 * negative_events)
        signal_balance = positive_events - negative_events
        status = "distracted" if negative_events > positive_events else "observed"
        if behavior_score >= 0.72 and negative_events <= positive_events:
            status = "engaged"
        record = StageImportRecord(
            batch_id=int(batch.id),
            course_id=int(course.id),
            stage_id=int(stage.id),
            user_id=user_id,
            subject=stage.subject,
            grade=stage.grade,
            metric_type=StageMetricType.participation,
            score_value=round(behavior_score * 100, 4),
            completion_value=_clamp01(behavior_events / 12.0),
            duration_minutes=float(min(60.0, behavior_events * 1.5)),
            status=status,
            note=(
                f"系统行为导入：{behavior_events} 条事件，"
                f"{active_days} 天活跃，{expression_events} 条表情/行为信号"
            ),
            happened_at=datetime.combine(day, datetime.min.time()) + timedelta(hours=12),
            raw_json=json.dumps(
                {
                    "source": "behavior_auto",
                    "day": day.isoformat(),
                    "user_id": user_id,
                    "row": {
                        "behavior_events": behavior_events,
                        "positive_events": positive_events,
                        "negative_events": negative_events,
                        "expression_events": expression_events,
                        "expression_focus": int(item["expression_focus"]),
                        "expression_distracted": int(item["expression_distracted"]),
                        "avg_confidence": (sum(confidence_values) / len(confidence_values)) if confidence_values else 0.0,
                        "signal_balance": signal_balance,
                    },
                },
                ensure_ascii=False,
            ),
        )
        session.add(record)
        success += 1
        affected_users.add(user_id)

    batch.success_rows = success
    batch.failed_rows = 0
    session.add(batch)
    session.commit()

    if affected_users:
        recalculate_stage_snapshots_for_stage(
            session,
            stage_id=int(stage.id),
            user_ids=sorted(affected_users),
            persist=True,
        )
    return batch, len(affected_users)


def _apply_internal_stage_rows(
    session: Session,
    *,
    course: Course,
    stage: CourseStage,
    user: User,
    include_video: bool = True,
    include_practice: bool = True,
    include_mastery: bool = True,
) -> tuple[StageImportBatch, int]:
    _, rows = _internal_stage_rows(session, course=course, stage=stage)
    student_ids = [int(row["user_id"]) for row in rows if row.get("user_id") is not None]
    enabled_groups = sum(1 for enabled in (include_video, include_practice, include_mastery) if enabled)

    existing_rows = session.exec(
        select(StageImportRecord).where(
            StageImportRecord.stage_id == int(stage.id),
            StageImportRecord.status == "internal_auto",
        )
    ).all()
    for row in existing_rows:
        session.delete(row)
    session.commit()

    batch = StageImportBatch(
        course_id=int(course.id),
        stage_id=int(stage.id),
        subject=stage.subject,
        grade=stage.grade,
        metric_type=StageMetricType.task,
        file_name="system_internal_summary",
        uploaded_by=user.username,
        total_rows=len(rows) * enabled_groups,
        success_rows=0,
        failed_rows=0,
        error_json="[]",
    )
    session.add(batch)
    session.commit()
    session.refresh(batch)

    success = 0
    for row in rows:
        user_id = int(row["user_id"])
        watched_minutes = float(row.get("watched_minutes") or 0.0)
        avg_video_completion = float(row.get("avg_video_completion") or 0.0)
        practice_attempts = int(row.get("practice_attempts") or 0)
        practice_accuracy = float(row.get("practice_accuracy") or 0.0)
        recommendation_count = int(row.get("recommendation_count") or 0)
        questionnaire_updates = int(row.get("questionnaire_updates") or 0)
        course_mastery = float(row.get("course_mastery") or 0.0)

        generated_records: list[StageImportRecord] = []
        if include_video:
            generated_records.append(
                StageImportRecord(
                    batch_id=int(batch.id),
                    course_id=int(course.id),
                    stage_id=int(stage.id),
                    user_id=user_id,
                    subject=stage.subject,
                    grade=stage.grade,
                    metric_type=StageMetricType.video,
                    score_value=round(avg_video_completion * 100, 4),
                    completion_value=avg_video_completion,
                    duration_minutes=watched_minutes,
                    status="internal_auto",
                    note="系统自动汇总：视频学习",
                    happened_at=stage.ends_at or datetime.utcnow(),
                    raw_json=json.dumps({"source": "internal_auto", "kind": "video", "row": row}, ensure_ascii=False),
                )
            )
        if include_practice:
            generated_records.append(
                StageImportRecord(
                    batch_id=int(batch.id),
                    course_id=int(course.id),
                    stage_id=int(stage.id),
                    user_id=user_id,
                    subject=stage.subject,
                    grade=stage.grade,
                    metric_type=StageMetricType.assignment,
                    score_value=round(practice_accuracy * 100, 4),
                    completion_value=min(1.0, practice_attempts / 5.0) if practice_attempts > 0 else 0.0,
                    submitted_on_time=practice_attempts > 0,
                    status="internal_auto",
                    note="系统自动汇总：练习表现",
                    happened_at=stage.ends_at or datetime.utcnow(),
                    raw_json=json.dumps({"source": "internal_auto", "kind": "practice", "row": row}, ensure_ascii=False),
                )
            )
        if include_mastery:
            generated_records.append(
                StageImportRecord(
                    batch_id=int(batch.id),
                    course_id=int(course.id),
                    stage_id=int(stage.id),
                    user_id=user_id,
                    subject=stage.subject,
                    grade=stage.grade,
                    metric_type=StageMetricType.task,
                    score_value=round(course_mastery * 100, 4),
                    completion_value=(
                        min(1.0, (recommendation_count + questionnaire_updates) / 4.0)
                        if (recommendation_count + questionnaire_updates) > 0
                        else course_mastery
                    ),
                    status="internal_auto",
                    note="系统自动汇总：掌握与推荐推进",
                    happened_at=stage.ends_at or datetime.utcnow(),
                    raw_json=json.dumps({"source": "internal_auto", "kind": "mastery", "row": row}, ensure_ascii=False),
                )
            )
        for record in generated_records:
            session.add(record)
            success += 1

    batch.success_rows = success
    batch.failed_rows = 0
    session.add(batch)
    session.commit()

    if student_ids:
        recalculate_stage_snapshots_for_stage(
            session,
            stage_id=int(stage.id),
            user_ids=sorted(set(student_ids)),
            persist=True,
        )
    return batch, len(set(student_ids))


@router.get("/courses/{course_id}", response_model=list[CourseStageOut])
def list_stages(
    course_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    rows = session.exec(
        select(CourseStage).where(CourseStage.course_id == course_id).order_by(CourseStage.stage_order, CourseStage.id)
    ).all()
    return [stage_support.course_out(row, course) for row in rows]


@router.post("/courses/{course_id}", response_model=CourseStageOut)
def create_stage(
    course_id: int,
    payload: CourseStageIn,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    _assert_course_stage_editable(course)
    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="title required")
    exists = session.exec(
        select(CourseStage).where(CourseStage.course_id == course_id, CourseStage.stage_order == payload.stage_order)
    ).first()
    if exists is not None:
        raise HTTPException(status_code=400, detail="stage_order already exists in this course")
    try:
        starts_at = stage_support.parse_dt(payload.starts_at)
        ends_at = stage_support.parse_dt(payload.ends_at)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    stage = CourseStage(
        course_id=course_id,
        subject=course.title,
        grade=(payload.grade or "通用").strip() or "通用",
        title=title,
        stage_order=max(1, int(payload.stage_order)),
        starts_at=starts_at,
        ends_at=ends_at,
        description=payload.description.strip(),
    )
    session.add(stage)
    session.commit()
    session.refresh(stage)
    stage_support.log_action(session, user, "course_stage_create", f"course_id={course_id} stage_id={stage.id}")
    return stage_support.course_out(stage)


@router.put("/{stage_id}", response_model=CourseStageOut)
def update_stage(
    stage_id: int,
    payload: CourseStageUpdateIn,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    stage = _get_stage_or_403(session, user, stage_id)
    course = _get_course_or_403(session, user, stage.course_id)
    _assert_course_stage_editable(course)
    if payload.title is not None:
        title = payload.title.strip()
        if not title:
            raise HTTPException(status_code=400, detail="title required")
        stage.title = title
    if payload.grade is not None:
        stage.grade = payload.grade.strip() or "通用"
    if payload.stage_order is not None:
        order_value = max(1, int(payload.stage_order))
        exists = session.exec(
            select(CourseStage).where(
                CourseStage.course_id == stage.course_id,
                CourseStage.stage_order == order_value,
                CourseStage.id != stage_id,
            )
        ).first()
        if exists is not None:
            raise HTTPException(status_code=400, detail="stage_order already exists in this course")
        stage.stage_order = order_value
    try:
        if payload.starts_at is not None:
            stage.starts_at = stage_support.parse_dt(payload.starts_at)
        if payload.ends_at is not None:
            stage.ends_at = stage_support.parse_dt(payload.ends_at)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if payload.description is not None:
        stage.description = payload.description.strip()
    session.add(stage)
    session.commit()
    session.refresh(stage)
    stage_support.log_action(session, user, "course_stage_update", f"stage_id={stage_id}")
    return stage_support.course_out(stage)


@router.delete("/{stage_id}")
def delete_stage(
    stage_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    stage = _get_stage_or_403(session, user, stage_id)
    course = _get_course_or_403(session, user, stage.course_id)
    _assert_course_stage_editable(course)
    has_batch = session.exec(select(StageImportBatch).where(StageImportBatch.stage_id == stage_id)).first()
    if has_batch is not None:
        raise HTTPException(status_code=400, detail="Stage already has imported data and cannot be deleted")
    session.delete(stage)
    session.commit()
    stage_support.log_action(session, user, "course_stage_delete", f"stage_id={stage_id}")
    return {"ok": True}


@router.get("/template")
def download_template(
    metric_type: str,
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _ = user
    try:
        metric = StageMetricType(metric_type)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid metric_type") from exc
    return Response(
        content="\ufeff" + stage_support.template_csv(metric),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="stage_template_{metric.value}.csv"'},
    )


@router.get("/metric-guides", response_model=list[StageMetricGuideOut])
def list_metric_guides(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _ = user
    return [stage_support.metric_guide(metric) for metric in StageMetricType]


@router.get("/imports", response_model=list[StageImportBatchOut])
def list_import_batches(
    course_id: int,
    stage_id: int | None = None,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    _get_course_or_403(session, user, course_id)
    query = select(StageImportBatch).where(StageImportBatch.course_id == course_id)
    if stage_id is not None:
        query = query.where(StageImportBatch.stage_id == stage_id)
    rows = session.exec(query.order_by(StageImportBatch.created_at.desc())).all()
    stage_map = {
        int(stage.id): stage.title
        for stage in session.exec(select(CourseStage).where(CourseStage.course_id == course_id)).all()
        if stage.id is not None
    }
    return [
        StageImportBatchOut(
            id=int(row.id),
            course_id=row.course_id,
            stage_id=row.stage_id,
            stage_title=stage_map.get(row.stage_id, f"阶段#{row.stage_id}"),
            subject=row.subject,
            grade=row.grade,
            metric_type=_metric_type_value(row.metric_type),
            file_name=row.file_name,
            uploaded_by=row.uploaded_by,
            total_rows=row.total_rows,
            success_rows=row.success_rows,
            failed_rows=row.failed_rows,
            error_preview=list(json.loads(row.error_json or "[]"))[:5],
            created_at=row.created_at.isoformat(),
        )
        for row in rows
    ]


@router.get("/internal-summary")
def internal_stage_summary(
    course_id: int,
    stage_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    summary, rows = _internal_stage_rows(session, course=course, stage=stage)
    return {
        "summary": summary,
        "rows": rows,
        "columns": [
            "username",
            "student_no",
            "full_name",
            "class_name",
            "watched_minutes",
            "avg_video_completion",
            "practice_attempts",
            "practice_accuracy",
            "recommendation_count",
            "questionnaire_updates",
            "course_mastery",
            "dynamic_score",
            "risk_level",
        ],
    }


@router.get("/internal-summary/export")
def export_internal_stage_summary(
    course_id: int,
    stage_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    _, rows = _internal_stage_rows(session, course=course, stage=stage)
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "账号",
            "学号",
            "姓名",
            "班级",
            "视频学习分钟",
            "平均视频完成率",
            "练习次数",
            "练习正确率",
            "推荐推进次数",
            "问卷更新次数",
            "课程掌握度",
            "动态评分",
            "风险等级",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["username"],
                row["student_no"],
                row["full_name"],
                row["class_name"],
                row["watched_minutes"],
                f'{round(float(row["avg_video_completion"]) * 100, 1)}%',
                row["practice_attempts"],
                f'{round(float(row["practice_accuracy"]) * 100, 1)}%',
                row["recommendation_count"],
                row["questionnaire_updates"],
                f'{round(float(row["course_mastery"]) * 100, 1)}%',
                f'{round(float(row["dynamic_score"]) * 100, 1)}%',
                row["risk_level"],
            ]
        )
    filename = f"stage_internal_summary_{course_id}_{stage_id}.csv"
    return Response(
        content="\ufeff" + buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/internal-behavior-summary")
def internal_behavior_summary(
    course_id: int,
    stage_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    _assert_course_stage_editable(course)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    summary, rows = _behavior_stage_rows(session, course=course, stage=stage)
    return {
        "summary": summary,
        "rows": rows,
        "columns": [
            "username",
            "student_no",
            "full_name",
            "class_name",
            "behavior_events",
            "active_days",
            "positive_events",
            "negative_events",
            "expression_events",
            "behavior_score",
            "dynamic_score",
            "risk_level",
        ],
    }


@router.get("/internal-behavior-summary/export")
def export_internal_behavior_summary(
    course_id: int,
    stage_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    _assert_course_stage_editable(course)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    _, rows = _behavior_stage_rows(session, course=course, stage=stage)
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "账号",
            "学号",
            "姓名",
            "班级",
            "行为事件数",
            "活跃天数",
            "正向信号",
            "负向信号",
            "表情信号数",
            "行为得分",
            "当前画像分",
            "风险等级",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["username"],
                row["student_no"],
                row["full_name"],
                row["class_name"],
                row["behavior_events"],
                row["active_days"],
                row["positive_events"],
                row["negative_events"],
                row["expression_events"],
                f'{round(float(row["behavior_score"]) * 100, 1)}%',
                f'{round(float(row["dynamic_score"]) * 100, 1)}%',
                row["risk_level"],
            ]
        )
    filename = f"stage_behavior_summary_{course_id}_{stage_id}.csv"
    return Response(
        content="\ufeff" + buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/internal-summary/apply", response_model=StageImportResultOut)
def apply_internal_stage_summary(
    course_id: int = Form(...),
    stage_id: int = Form(...),
    include_video: bool = Form(True),
    include_practice: bool = Form(True),
    include_mastery: bool = Form(True),
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    if not (include_video or include_practice or include_mastery):
        raise HTTPException(status_code=400, detail="请至少选择一个系统汇总映射项")
    batch, recalculated_users = _apply_internal_stage_rows(
        session,
        course=course,
        stage=stage,
        user=user,
        include_video=include_video,
        include_practice=include_practice,
        include_mastery=include_mastery,
    )
    affected_indicators: list[str] = []
    if include_video:
        affected_indicators.append("视频学习投入")
    if include_practice:
        affected_indicators.append("练习表现")
    if include_mastery:
        affected_indicators.append("掌握与推荐推进")
    stage_support.log_action(
        session,
        user,
        "stage_internal_summary_apply",
        "course_id=%s stage_id=%s success=%s include_video=%s include_practice=%s include_mastery=%s"
        % (course_id, stage_id, batch.success_rows, include_video, include_practice, include_mastery),
    )
    return StageImportResultOut(
        batch_id=int(batch.id),
        metric_type="internal_auto",
        total_rows=int(batch.total_rows),
        success_rows=int(batch.success_rows),
        failed_rows=int(batch.failed_rows),
        errors=[],
        affected_dimensions=["学习行为与过程", "知识与认知状态", "情感与社会性发展"],
        affected_indicators=affected_indicators,
        recalculated_users=recalculated_users,
        next_action="系统已按当前映射将平台内学习数据写入该阶段记录。后续外部补充导入会继续叠加到同一阶段，并再次重算学生画像。",
    )


@router.post("/internal-behavior-summary/apply", response_model=StageImportResultOut)
def apply_internal_behavior_summary(
    course_id: int = Form(...),
    stage_id: int = Form(...),
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    _assert_course_stage_editable(course)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    batch, recalculated_users = _apply_behavior_stage_rows(
        session,
        course=course,
        stage=stage,
        user=user,
    )
    stage_support.log_action(
        session,
        user,
        "stage_internal_behavior_apply",
        "course_id=%s stage_id=%s success=%s" % (course_id, stage_id, batch.success_rows),
    )
    return StageImportResultOut(
        batch_id=int(batch.id),
        metric_type="behavior_auto",
        total_rows=int(batch.total_rows),
        success_rows=int(batch.success_rows),
        failed_rows=int(batch.failed_rows),
        errors=[],
        affected_dimensions=["学习行为与过程", "情感与社会性发展"],
        affected_indicators=["学习动机与态度", "协作能力与社交网络", "自主调节与专注度"],
        recalculated_users=recalculated_users,
        next_action="系统行为事件已导入到当前阶段画像，后续可继续叠加线下补充数据并再次重算。",
    )


@router.post("/one-click-import", response_model=StageImportResultOut)
def one_click_stage_import(
    course_id: int = Form(...),
    stage_id: int = Form(...),
    include_video: bool = Form(True),
    include_practice: bool = Form(True),
    include_mastery: bool = Form(True),
    include_behavior: bool = Form(True),
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    _assert_course_stage_editable(course)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    if not (include_video or include_practice or include_mastery or include_behavior):
        raise HTTPException(status_code=400, detail="at least one import source must be enabled")

    total_rows = 0
    success_rows = 0
    failed_rows = 0
    batch_ids: list[int] = []
    recalculated_users = 0
    affected_dimensions: list[str] = []
    affected_indicators: list[str] = []

    if include_video or include_practice or include_mastery:
        batch, summary_recalculated = _apply_internal_stage_rows(
            session,
            course=course,
            stage=stage,
            user=user,
            include_video=include_video,
            include_practice=include_practice,
            include_mastery=include_mastery,
        )
        total_rows += int(batch.total_rows)
        success_rows += int(batch.success_rows)
        failed_rows += int(batch.failed_rows)
        recalculated_users = max(recalculated_users, summary_recalculated)
        if batch.id is not None:
            batch_ids.append(int(batch.id))
        affected_dimensions.extend(["learning_process", "knowledge_cognition", "emotion_social"])
        if include_video:
            affected_indicators.append("video_learning")
        if include_practice:
            affected_indicators.append("practice_performance")
        if include_mastery:
            affected_indicators.append("mastery_progress")

    if include_behavior:
        behavior_batch, behavior_recalculated = _apply_behavior_stage_rows(
            session,
            course=course,
            stage=stage,
            user=user,
        )
        total_rows += int(behavior_batch.total_rows)
        success_rows += int(behavior_batch.success_rows)
        failed_rows += int(behavior_batch.failed_rows)
        recalculated_users = max(recalculated_users, behavior_recalculated)
        if behavior_batch.id is not None:
            batch_ids.append(int(behavior_batch.id))
        affected_dimensions.extend(["learning_process", "emotion_social"])
        affected_indicators.extend(["motivation_attitude", "collaboration_social", "self_regulation"])

    unique_dimensions = list(dict.fromkeys(affected_dimensions))
    unique_indicators = list(dict.fromkeys(affected_indicators))
    latest_batch_id = batch_ids[-1] if batch_ids else 0
    stage_support.log_action(
        session,
        user,
        "stage_one_click_import",
        "course_id=%s stage_id=%s include_video=%s include_practice=%s include_mastery=%s include_behavior=%s total=%s"
        % (course_id, stage_id, include_video, include_practice, include_mastery, include_behavior, success_rows),
    )
    import_summary = stage_support.build_import_summary(
        course=course,
        stage=stage,
        metric_type="one_click_auto",
        total_rows=total_rows,
        success_rows=success_rows,
        failed_rows=failed_rows,
        recalculated_users=recalculated_users,
        source_mode="one_click_auto",
    )
    import_summary["batch_ids"] = batch_ids
    import_summary["enabled_sources"] = {
        "video": include_video,
        "practice": include_practice,
        "mastery": include_mastery,
        "behavior": include_behavior,
    }
    return StageImportResultOut(
        batch_id=latest_batch_id,
        metric_type="one_click_auto",
        total_rows=total_rows,
        success_rows=success_rows,
        failed_rows=failed_rows,
        errors=[],
        affected_dimensions=unique_dimensions,
        affected_indicators=unique_indicators,
        recalculated_users=recalculated_users,
        next_action="One-click import completed. The current stage has been refreshed with system summary data and behavior signals.",
        import_summary=import_summary,
    )


@router.post("/imports/upload", response_model=StageImportResultOut)
def upload_stage_data(
    course_id: int = Form(...),
    stage_id: int = Form(...),
    metric_type: str = Form(...),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
):
    course = _get_course_or_403(session, user, course_id)
    stage = _get_stage_or_403(session, user, stage_id)
    if stage.course_id != course_id:
        raise HTTPException(status_code=400, detail="stage does not belong to the course")
    try:
        metric = StageMetricType(metric_type)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid metric_type") from exc

    payload = file.file.read()
    rows = stage_support.rows_from_upload(file, payload)
    batch = StageImportBatch(
        course_id=course_id,
        stage_id=stage_id,
        subject=stage.subject,
        grade=stage.grade,
        metric_type=metric,
        file_name=file.filename or "",
        uploaded_by=user.username,
        total_rows=len(rows),
    )
    session.add(batch)
    session.commit()
    session.refresh(batch)

    errors: list[str] = []
    success = 0
    affected_user_ids: set[int] = set()
    for index, row in enumerate(rows, start=2):
        try:
            student = stage_support.find_user(session, row)
            if student.role != UserRole.student:
                raise ValueError("import user is not a student")
            enrollment = session.exec(
                select(Enrollment.id).where(
                    Enrollment.student_id == int(student.id),
                    Enrollment.course_id == course_id,
                    Enrollment.status == EnrollmentStatus.active,
                )
            ).first()
            if enrollment is None:
                raise ValueError("student is not enrolled in this course")
            kp = stage_support.find_kp(session, subject=stage.subject, grade=stage.grade, row=row)
            score_value = stage_support.to_float(row.get("score"), default=0.0)
            completion_value = stage_support.to_ratio(row.get("completion_ratio") or row.get("completion_value"))
            duration_minutes = stage_support.to_float(row.get("duration_minutes") or row.get("watched_minutes"), default=0.0)
            attendance_value = stage_support.to_float(row.get("attendance_value"), default=0.0)
            if metric == StageMetricType.attendance and not row.get("attendance_value"):
                attendance_value = 1.0 if (row.get("status") or "").strip().lower() in {"present", "attended", "on_time"} else 0.0
            happened_at = stage_support.parse_dt(row.get("happened_at")) or datetime.utcnow()
            record = StageImportRecord(
                batch_id=int(batch.id),
                course_id=course_id,
                stage_id=stage_id,
                user_id=int(student.id),
                kp_id=int(kp.id) if kp is not None else None,
                subject=stage.subject,
                grade=stage.grade,
                metric_type=metric,
                score_value=score_value,
                completion_value=completion_value,
                duration_minutes=duration_minutes,
                attendance_value=attendance_value,
                submitted_on_time=stage_support.to_bool(row.get("submitted_on_time")),
                status=(row.get("status") or "").strip(),
                note=(row.get("note") or "").strip(),
                happened_at=happened_at,
                raw_json=json.dumps(row, ensure_ascii=False),
            )
            session.add(record)
            success += 1
            if student.id is not None:
                affected_user_ids.add(int(student.id))
        except Exception as exc:
            errors.append(f"第 {index} 行：{exc}")

    batch.success_rows = success
    batch.failed_rows = len(rows) - success
    batch.error_json = json.dumps(errors[:20], ensure_ascii=False)
    session.add(batch)
    session.commit()
    if affected_user_ids:
        try:
            recalculate_stage_snapshots_for_stage(
                session,
                stage_id=stage_id,
                user_ids=sorted(affected_user_ids),
                persist=True,
            )
        except Exception as exc:
            logger.exception("stage snapshot recalc failed: stage_id=%s", stage_id)
            errors.append(f"阶段评价重算失败：{exc}")
            batch.error_json = json.dumps(errors[:20], ensure_ascii=False)
            session.add(batch)
            session.commit()
    stage_support.log_action(
        session,
        user,
        "stage_data_import",
        f"course_id={course_id} stage_id={stage_id} metric={metric.value} success={success} failed={batch.failed_rows}",
    )
    return StageImportResultOut(
        batch_id=int(batch.id),
        metric_type=metric.value,
        total_rows=len(rows),
        success_rows=success,
        failed_rows=len(rows) - success,
        errors=errors[:20],
        affected_dimensions=list(STAGE_IMPORT_GUIDES[metric]["affected_dimensions"]),
        affected_indicators=list(STAGE_IMPORT_GUIDES[metric]["affected_indicators"]),
        recalculated_users=len(affected_user_ids),
        next_action=str(STAGE_IMPORT_GUIDES[metric]["next_action"]),
    )
